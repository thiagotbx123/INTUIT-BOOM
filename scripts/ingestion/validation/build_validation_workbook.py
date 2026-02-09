"""
Build the Validation Specification Workbook (XLSX).
This is the single source of truth for ALL ingestion validation rules,
reference data, and CSV manifest. Both humans and Claude can use it.

Architecture:
  Sheet 1: RULES         - Every validation rule as a structured row
  Sheet 2: FK_MAP        - Foreign key mappings (source -> target)
  Sheet 3: EMPLOYEES     - Valid employee IDs + metadata
  Sheet 4: PROJECTS      - Valid project IDs + date ranges
  Sheet 5: PRODUCTS      - Valid product IDs + purchasable flag
  Sheet 6: ENUMS         - Valid enum values per column
  Sheet 7: CSV_MANIFEST  - All 15 CSVs with metadata
  Sheet 8: DATASETS      - Dataset configs
  Sheet 9: CHANGELOG     - Bug history
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
import os

OUTPUT = r"C:\Users\adm_r\Downloads\PARA_AUGUSTO\VALIDATION_SPEC_CONSTRUCTION.xlsx"
DB_PATH = r"C:\Users\adm_r\Clients\intuit-boom\qbo_database (1).db"
DATASET_ID = "321c6fa0-a4ee-4e05-b085-7b4d51473495"

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BLOCKER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
            cell.border = THIN_BORDER
        if lengths:
            width = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = width + 2


def build_rules_sheet(wb):
    ws = wb.create_sheet("RULES")
    headers = [
        "rule_id",
        "category",
        "severity",
        "table",
        "column",
        "validation_type",
        "condition",
        "ref_sheet",
        "ref_column",
        "description",
        "origin",
        "bug_ref",
        "added_date",
    ]
    ws.append(headers)

    rules = [
        # QB rules - BLOCKER
        [
            "QB_001",
            "QB",
            "BLOCKER",
            "ALL",
            "dataset_id",
            "EXACT_MATCH",
            f"== '{DATASET_ID}'",
            "",
            "",
            "All rows must have correct dataset_id",
            "Augusto",
            "",
            "2026-01",
        ],
        [
            "QB_002",
            "QB",
            "BLOCKER",
            "ALL",
            "company_type",
            "ENUM",
            "IN ref",
            "ENUMS",
            "company_type",
            "Must be parent/main_child/secondary_child",
            "Augusto",
            "",
            "2026-01",
        ],
        [
            "QB_003",
            "QB",
            "BLOCKER",
            "ALL",
            "*_relative|*_date",
            "RANGE",
            "interval_days <= 365",
            "",
            "",
            "All relative dates must be <= 365 days",
            "Augusto",
            "",
            "2026-01",
        ],
        [
            "QB_004",
            "QB",
            "BLOCKER",
            "ALL",
            "*_relative",
            "COVERAGE",
            "covers months 1-12",
            "",
            "",
            "Dates must cover full year (12 months)",
            "Augusto",
            "",
            "2026-02-07",
        ],
        [
            "QB_005",
            "QB",
            "BLOCKER",
            "ALL",
            "base_date",
            "EXACT_MATCH",
            "== 'first_of_year'",
            "",
            "",
            "base_date always 'first_of_year', never YYYY-MM-DD",
            "Augusto",
            "",
            "2026-01",
        ],
        [
            "QB_006",
            "QB",
            "BLOCKER",
            "ALL",
            "*_relative",
            "FORMAT",
            "regex: ^\\d+ mons? \\d+ days?",
            "",
            "",
            "PostgreSQL interval format only",
            "bug#6",
            "",
            "2026-01",
        ],
        [
            "QB_007",
            "QB",
            "BLOCKER",
            "ALL",
            "id",
            "GLOBAL_UNIQUE",
            "no collision cross-dataset",
            "",
            "",
            "IDs must be globally unique (query MAX(id) from DB first)",
            "collision",
            "",
            "2026-01",
        ],
        [
            "QB_008",
            "QB",
            "BLOCKER",
            "ALL",
            "id",
            "SEQUENTIAL",
            "no gaps in sequence",
            "",
            "",
            "IDs sequential without gaps",
            "",
            "",
            "2026-01",
        ],
        [
            "QB_009",
            "QB",
            "BLOCKER",
            "ALL",
            "*_id (FK)",
            "FK_EXISTS",
            "IN ref_sheet.id",
            "FK_MAP",
            "",
            "All FKs must reference existing records in DB",
            "bug#21,22",
            "#21,#22",
            "2026-02-09",
        ],
        [
            "QB_010",
            "QB",
            "BLOCKER",
            "invoices,bills",
            "parent->LI",
            "HAS_CHILDREN",
            "count(LI) >= 1",
            "",
            "",
            "Every parent must have at least 1 line item",
            "",
            "",
            "2026-01",
        ],
        [
            "QB_011",
            "QB",
            "BLOCKER",
            "bills",
            "paid_date",
            "NOT_NULL",
            "!= NULL, != ''",
            "",
            "",
            "Bills paid_date must be filled",
            "Augusto",
            "",
            "2026-01",
        ],
        [
            "QB_012",
            "QB",
            "BLOCKER",
            "bank_txns",
            "account",
            "LOOKUP",
            "parent=5, child=6",
            "",
            "",
            "Bank account by company type",
            "Augusto",
            "",
            "2026-01",
        ],
        [
            "QB_013",
            "QB",
            "BLOCKER",
            "invoice_li,bill_li",
            "qty",
            "RANGE",
            ">= 1",
            "",
            "",
            "Line item qty must be >= 1",
            "Augusto",
            "#10",
            "2026-02-07",
        ],
        [
            "QB_014",
            "QB",
            "BLOCKER",
            "invoices,bills",
            "terms+date",
            "FORMULA",
            "terms_days + relative_days <= 365",
            "",
            "",
            "Terms + date cannot overflow 365 days",
            "bug#10",
            "#10",
            "2026-02-07",
        ],
        [
            "QB_015",
            "QB",
            "BLOCKER",
            "bill_li",
            "product_service_id",
            "FK_FILTERED",
            "product.purchasable == true",
            "PRODUCTS",
            "purchasable",
            "Bills only reference purchasable products",
            "Lucas T",
            "",
            "2026-02-07",
        ],
        [
            "QB_016",
            "QB",
            "BLOCKER",
            "invoice_li,bill_li",
            "amount",
            "FORMULA",
            "amount == rate * qty (2 decimal)",
            "",
            "",
            "Amount = UnitPrice x Qty exactly",
            "web",
            "",
            "2026-02-08",
        ],
        [
            "QB_017",
            "QB",
            "HIGH",
            "invoices,bills",
            "doc_number",
            "MAX_LENGTH",
            "len <= 21",
            "",
            "",
            "DocNumber max 21 chars (QBO truncates silently)",
            "web",
            "",
            "2026-02-08",
        ],
        [
            "QB_018",
            "QB",
            "BLOCKER",
            "estimates",
            "target_status",
            "ENUM",
            "IN ref",
            "ENUMS",
            "estimate_status",
            "Only 'accepted' or 'cancelled'",
            "DB analysis",
            "",
            "2026-02-08",
        ],
        [
            "QB_019",
            "QB",
            "BLOCKER",
            "est_li_class,po_li_class",
            "per_line_item",
            "COUNT",
            "== 4 per LI",
            "",
            "",
            "Exactly 4 classification entries per line item",
            "DB analysis",
            "",
            "2026-02-08",
        ],
        [
            "QB_020",
            "QB",
            "BLOCKER",
            "expenses",
            "employee_id",
            "FK_EXISTS",
            "IN ref_sheet.id",
            "EMPLOYEES",
            "id",
            "employee_id must exist in employees table (has gaps!)",
            "bug#21",
            "#21",
            "2026-02-09",
        ],
        [
            "QB_021",
            "QB",
            "BLOCKER",
            "time_entries",
            "project_id",
            "FK_FILTERED",
            "project.status == 'In progress'",
            "PROJECTS",
            "status",
            "Only reference active (In progress) projects",
            "bug#22",
            "#22",
            "2026-02-09",
        ],
        [
            "QB_022",
            "QB",
            "BLOCKER",
            "time_entries",
            "start_date",
            "CROSS_REF",
            "month BETWEEN project.start_month AND project.end_month",
            "PROJECTS",
            "start_month,end_month",
            "Time entry date must fall within project active period",
            "bug#24",
            "#24",
            "2026-02-09",
        ],
        # BIZ rules
        [
            "BIZ_001",
            "BIZ",
            "HIGH",
            "ALL_FINANCIAL",
            "P&L",
            "FORMULA",
            "income - costs > 0 per CT per month",
            "",
            "",
            "P&L must be positive per company type per month",
            "",
            "",
            "2026-01",
        ],
        [
            "BIZ_002",
            "BIZ",
            "HIGH",
            "invoices",
            "COGS",
            "AWARENESS",
            "FIFO for 38 inventory products",
            "",
            "",
            "COGS generated on invoice creation (not payment)",
            "web",
            "",
            "2026-02-08",
        ],
        [
            "BIZ_003",
            "BIZ",
            "MEDIUM",
            "ALL_FINANCIAL",
            "budget",
            "CROSS_TABLE",
            "costs < income per CT",
            "",
            "",
            "Expense budget caps enforced",
            "",
            "#4",
            "2026-01",
        ],
        [
            "BIZ_004",
            "BIZ",
            "HIGH",
            "invoices",
            "payroll",
            "AWARENESS",
            "main_child has $1.78M hidden payroll",
            "",
            "",
            "Account for payroll in P&L calculations",
            "",
            "#7",
            "2026-01",
        ],
        # RL rules
        [
            "RL_001",
            "RL",
            "MEDIUM",
            "ALL",
            "notes",
            "DISTRIBUTION",
            "5-8 templates, 40% null",
            "",
            "",
            "Notes variation for realism",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_002",
            "RL",
            "MEDIUM",
            "invoices,bills",
            "terms",
            "DISTRIBUTION",
            "Net30:40%, Net45:25%, Net60:20%, DOR:10%, Net90:5%",
            "ENUMS",
            "terms",
            "Terms distribution for realism",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_003",
            "RL",
            "MEDIUM",
            "invoice_li",
            "count_per_parent",
            "DISTRIBUTION",
            "log-normal 1-25",
            "",
            "",
            "Line item count follows log-normal distribution",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_004",
            "RL",
            "MEDIUM",
            "invoices",
            "amount_range",
            "RANGE",
            "500 - 200000",
            "",
            "",
            "Invoice amount realistic range",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_005",
            "RL",
            "MEDIUM",
            "invoices",
            "customer-project",
            "AFFINITY",
            "1-2 projects per customer",
            "",
            "",
            "Customer-project affinity maintained",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_006",
            "RL",
            "MEDIUM",
            "invoice_li",
            "product per invoice",
            "UNIQUE_PER_PARENT",
            "no duplicates",
            "",
            "",
            "No duplicate products per invoice",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_007",
            "RL",
            "MEDIUM",
            "bills",
            "bill-PO link",
            "PERCENTAGE",
            "~22% linked",
            "",
            "",
            "Bills linked to POs at realistic rate",
            "",
            "#19,#20",
            "2026-02",
        ],
        [
            "RL_008",
            "RL",
            "MEDIUM",
            "expenses",
            "expense-estimate",
            "PERCENTAGE",
            "~7-10% linked",
            "",
            "",
            "Expenses linked to estimates at realistic rate",
            "",
            "#19",
            "2026-02",
        ],
        [
            "RL_009",
            "RL",
            "LOW",
            "bills",
            "vendor-product",
            "AFFINITY",
            "consistent specialization",
            "",
            "",
            "Vendor-product affinity for realism",
            "",
            "",
            "2026-01",
        ],
        [
            "RL_010",
            "RL",
            "LOW",
            "bills",
            "status",
            "CONVENTION",
            "empty string when paid",
            "",
            "",
            "Bill status = '' when paid (QBO convention)",
            "",
            "#16",
            "2026-01",
        ],
    ]

    for rule in rules:
        ws.append(rule)

    style_header(ws, len(headers))

    # Color by severity
    for row in range(2, ws.max_row + 1):
        sev = ws.cell(row=row, column=3).value
        fill = (
            BLOCKER_FILL if sev == "BLOCKER" else HIGH_FILL if sev == "HIGH" else None
        )
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return ws


def build_fk_map_sheet(wb):
    ws = wb.create_sheet("FK_MAP")
    headers = [
        "source_table",
        "source_column",
        "target_table",
        "target_column",
        "filter",
        "ref_sheet",
        "notes",
    ]
    ws.append(headers)

    fk_rows = [
        [
            "expenses",
            "employee_id",
            "employees",
            "id",
            "dataset_id match",
            "EMPLOYEES",
            "Has gaps: 4,6,38,42,46,48 missing",
        ],
        [
            "time_entries",
            "employee_id",
            "employees",
            "id",
            "dataset_id match",
            "EMPLOYEES",
            "11 unique IDs in v5",
        ],
        [
            "time_entries",
            "project_id",
            "projects",
            "id",
            "status='In progress'",
            "PROJECTS",
            "Only active projects",
        ],
        [
            "time_entries",
            "customer_id",
            "customers",
            "id",
            "dataset_id match",
            "",
            "50 customers",
        ],
        [
            "time_entries",
            "product_service_id",
            "products",
            "id",
            "dataset_id match",
            "PRODUCTS",
            "",
        ],
        [
            "invoices",
            "customer_id",
            "customers",
            "id",
            "dataset_id match",
            "",
            "50 customers",
        ],
        ["invoice_li", "invoice_id", "invoices", "id", "same CSV", "", "Parent-child"],
        [
            "invoice_li",
            "product_service_id",
            "products",
            "id",
            "dataset_id match",
            "PRODUCTS",
            "",
        ],
        ["bills", "vendor_id", "vendors", "id", "dataset_id match", "", "33 vendors"],
        ["bill_li", "bill_id", "bills", "id", "same CSV", "", "Parent-child"],
        [
            "bill_li",
            "product_service_id",
            "products",
            "id",
            "purchasable=true",
            "PRODUCTS",
            "Must be purchasable",
        ],
        [
            "expenses",
            "vendor_id",
            "vendors",
            "id",
            "dataset_id match",
            "",
            "Optional FK",
        ],
        [
            "expenses",
            "customer_id",
            "customers",
            "id",
            "dataset_id match",
            "",
            "Optional FK",
        ],
        ["expense_li", "expense_id", "expenses", "id", "same CSV", "", "Parent-child"],
        ["estimates", "customer_id", "customers", "id", "dataset_id match", "", ""],
        [
            "estimate_li",
            "estimate_id",
            "estimates",
            "id",
            "same CSV",
            "",
            "Parent-child",
        ],
        [
            "est_li_class",
            "line_item_id",
            "estimate_li",
            "id",
            "same CSV",
            "",
            "4 per LI",
        ],
        ["purchase_orders", "vendor_id", "vendors", "id", "dataset_id match", "", ""],
        [
            "po_li",
            "purchase_order_id",
            "purchase_orders",
            "id",
            "same CSV",
            "",
            "Parent-child",
        ],
        ["po_li_class", "line_item_id", "po_li", "id", "same CSV", "", "4 per LI"],
        [
            "bank_txns",
            "customer_id",
            "customers",
            "id",
            "dataset_id match",
            "",
            "Optional",
        ],
    ]

    for row in fk_rows:
        ws.append(row)

    style_header(ws, len(headers))
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_employees_sheet(wb):
    ws = wb.create_sheet("EMPLOYEES")
    headers = [
        "id",
        "first_name",
        "last_name",
        "status",
        "pay_type",
        "company_type",
        "assigned_project_id",
        "dataset_id",
    ]
    ws.append(headers)

    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, first_name, last_name, status, pay_type, company_type, assigned_project_id, dataset_id
            FROM employees
            WHERE dataset_id = ?
            ORDER BY id
        """,
            (DATASET_ID,),
        )
        for row in cur.fetchall():
            ws.append(list(row))
        conn.close()
        print(f"  EMPLOYEES: {ws.max_row - 1} rows from DB")
    except Exception as e:
        print(f"  EMPLOYEES: DB error - {e}, using hardcoded")
        valid_ids = [
            1,
            2,
            3,
            4,
            5,
            7,
            8,
            9,
            10,
            11,
            12,
            13,
            14,
            15,
            16,
            17,
            18,
            19,
            20,
            21,
            22,
            23,
            24,
            25,
            26,
            27,
            28,
            29,
            30,
            31,
            32,
            33,
            34,
            35,
            36,
            37,
            39,
            40,
            41,
            43,
            44,
            45,
            47,
            49,
            50,
        ]
        for eid in valid_ids:
            ws.append([eid, "", "", "Active", "", "all", None, DATASET_ID])

    style_header(ws, len(headers))
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_projects_sheet(wb):
    ws = wb.create_sheet("PROJECTS")
    headers = [
        "id",
        "name",
        "status",
        "relative_start_date",
        "project_length",
        "start_month",
        "end_month",
        "valid_for_TE",
        "dataset_id",
    ]
    ws.append(headers)

    projects = [
        [
            19,
            "Azure Pines - Playground",
            "In progress",
            "5 mons",
            "9 mons",
            6,
            12,
            "YES",
            DATASET_ID,
        ],
        [
            20,
            "GaleGuardian - Turbine",
            "In progress",
            "2 mons",
            "6 mons",
            3,
            8,
            "YES",
            DATASET_ID,
        ],
        [21, "Intuit Dome", "In progress", "1 mon", "1 year", 2, 12, "YES", DATASET_ID],
        [
            22,
            "Leap Labs - Solar Array",
            "In progress",
            "4 mons",
            "6 mons",
            5,
            10,
            "YES",
            DATASET_ID,
        ],
        [
            23,
            "BMH Landscaping",
            "In progress",
            "3 mons",
            "9 mons",
            4,
            12,
            "YES",
            DATASET_ID,
        ],
        [
            24,
            "TidalWave - Farmer's Market",
            "In progress",
            "4 mons",
            "1 year",
            5,
            12,
            "YES",
            DATASET_ID,
        ],
        [
            25,
            "Rowe Hotel - Parking Lot",
            "Canceled",
            "6 mons",
            "",
            None,
            None,
            "NO",
            DATASET_ID,
        ],
        [
            26,
            "Arbor Dunes - Garden Walk",
            "Canceled",
            "8 mons",
            "",
            None,
            None,
            "NO",
            DATASET_ID,
        ],
    ]

    for row in projects:
        ws.append(row)

    style_header(ws, len(headers))
    # Color cancelled rows
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=3).value == "Canceled":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = BLOCKER_FILL
        else:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = GREEN_FILL

    auto_width(ws)
    ws.freeze_panes = "A2"


def build_products_sheet(wb):
    ws = wb.create_sheet("PRODUCTS")
    headers = ["id", "name", "type", "purchasable", "cost", "dataset_id"]
    ws.append(headers)

    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        # Try different table names
        for tbl in ["products", "product_services", "quickbooks_product_services"]:
            try:
                cur.execute(
                    f"SELECT id, name, type, purchasable, cost, dataset_id FROM {tbl} WHERE dataset_id = ? ORDER BY id",
                    (DATASET_ID,),
                )
                rows = cur.fetchall()
                if rows:
                    for row in rows:
                        ws.append(list(row))
                    print(f"  PRODUCTS: {len(rows)} rows from {tbl}")
                    break
            except Exception:
                continue
        conn.close()
    except Exception as e:
        print(f"  PRODUCTS: DB error - {e}")
        ws.append([0, "DB_ERROR", "", "", "", ""])

    style_header(ws, len(headers))
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_enums_sheet(wb):
    ws = wb.create_sheet("ENUMS")
    headers = ["table", "column", "valid_values", "distribution", "notes"]
    ws.append(headers)

    enums = [
        [
            "ALL",
            "company_type",
            "parent|main_child|secondary_child",
            "",
            "3 valid types",
        ],
        ["ALL", "base_date", "first_of_year", "", "Always this value"],
        [
            "estimates",
            "target_status",
            "accepted|cancelled",
            "97%|3%",
            "Only 2 valid values",
        ],
        [
            "invoices",
            "terms",
            "Net 30|Net 45|Net 60|Due on Receipt|Net 90",
            "40%|25%|20%|10%|5%",
            "Distribution for realism",
        ],
        [
            "bills",
            "terms",
            "Net 30|Net 45|Net 60|Due on Receipt|Net 90",
            "40%|25%|20%|10%|5%",
            "Same as invoices",
        ],
        [
            "bills",
            "status",
            "(empty string)",
            "100%",
            "Empty when paid (QBO convention)",
        ],
        ["time_entries", "entry_type", "Start and End Date", "100%", ""],
        ["time_entries", "timezone", "Pacific Time", "100%", ""],
        [
            "employees",
            "status",
            "Active|Terminated|Paid leave|Unpaid leave|Not on payroll",
            "71%|11%|7%|7%|4%",
            "",
        ],
        [
            "employees",
            "pay_type",
            "Hourly|Salary",
            "53%|47%",
            "One record has lowercase 'hourly'",
        ],
        [
            "projects",
            "status",
            "In progress|Canceled",
            "75%|25%",
            "Only In progress valid for TEs",
        ],
    ]

    for row in enums:
        ws.append(row)

    style_header(ws, len(headers))
    auto_width(ws)
    ws.freeze_panes = "A2"


def build_csv_manifest_sheet(wb):
    ws = wb.create_sheet("CSV_MANIFEST")
    headers = [
        "file_num",
        "filename",
        "target_table",
        "rows",
        "id_start",
        "id_end",
        "delimiter",
        "operation",
        "depends_on",
        "employee_fk",
        "project_fk",
        "status",
        "notes",
    ]
    ws.append(headers)

    manifest = [
        [
            1,
            "INGESTION_INVOICES_v12.csv",
            "quickbooks_invoices",
            335,
            45135,
            45469,
            ",",
            "REPLACE",
            "customers,products",
            "NO",
            "NO",
            "READY",
            "",
        ],
        [
            2,
            "INGESTION_INVOICE_LINE_ITEMS_v8.csv",
            "quickbooks_invoice_line_items",
            1686,
            119728,
            121413,
            ";",
            "REPLACE",
            "file_01",
            "NO",
            "NO",
            "READY",
            "SEMICOLON delimiter",
        ],
        [
            3,
            "INGESTION_EXPENSES_v4.csv",
            "quickbooks_expenses",
            650,
            9645,
            10294,
            ",",
            "REPLACE",
            "customers,vendors,employees",
            "YES",
            "NO",
            "FIXED(#21)",
            "8 employee IDs remapped",
        ],
        [
            4,
            "INGESTION_EXPENSE_LINE_ITEMS_v4.csv",
            "quickbooks_expense_line_items",
            1006,
            8016,
            9021,
            ",",
            "REPLACE",
            "file_03",
            "NO",
            "NO",
            "READY",
            "",
        ],
        [
            5,
            "INGESTION_BILLS_v5.csv",
            "quickbooks_bills",
            96,
            4903,
            4998,
            ",",
            "REPLACE",
            "vendors,products",
            "NO",
            "NO",
            "FIXED(#20)",
            "TXN-13131 bill-PO fix",
        ],
        [
            6,
            "INGESTION_BILLS_LINE_ITEMS_v5.csv",
            "quickbooks_bills_line_items",
            149,
            5567,
            5715,
            ",",
            "REPLACE",
            "file_05,products",
            "NO",
            "NO",
            "READY",
            "purchasable products only",
        ],
        [
            7,
            "INGESTION_ESTIMATES_v1.csv",
            "quickbooks_estimates",
            75,
            766,
            840,
            ",",
            "REPLACE",
            "customers",
            "NO",
            "NO",
            "READY",
            "",
        ],
        [
            8,
            "INGESTION_ESTIMATE_LINE_ITEMS_v1.csv",
            "quickbooks_estimate_line_items",
            410,
            4079,
            4488,
            ",",
            "REPLACE",
            "file_07",
            "NO",
            "NO",
            "READY",
            "",
        ],
        [
            9,
            "INGESTION_ESTIMATE_LI_CLASSIFICATIONS_v1.csv",
            "quickbooks_estimate_line_item_classifications",
            1640,
            5626,
            7265,
            ",",
            "REPLACE",
            "file_08",
            "NO",
            "NO",
            "READY",
            "4 per LI",
        ],
        [
            10,
            "INGESTION_BANK_TRANSACTIONS_v1.csv",
            "quickbooks_bank_transactions",
            155,
            1321,
            1475,
            ",",
            "REPLACE",
            "customers",
            "NO",
            "NO",
            "READY",
            "",
        ],
        [
            11,
            "PLA-3261_TIME_ENTRIES_v5_supplement.csv",
            "quickbooks_time_entries",
            1575,
            73993,
            77988,
            ";",
            "APPEND",
            "employees,projects,products",
            "YES",
            "YES",
            "FIXED(#22,#24)",
            "Was 3996, now 1575",
        ],
        [
            12,
            "INGESTION_PROJECT_TASKS_v1.csv",
            "quickbooks_project_tasks",
            65,
            121,
            185,
            ",",
            "INSERT",
            "projects",
            "NO",
            "YES",
            "READY",
            "",
        ],
        [
            13,
            "INGESTION_PURCHASE_ORDERS_v1.csv",
            "quickbooks_purchase_orders",
            48,
            280,
            327,
            ",",
            "INSERT",
            "vendors",
            "NO",
            "NO",
            "READY",
            "Table may need creation",
        ],
        [
            14,
            "INGESTION_PO_LINE_ITEMS_v1.csv",
            "quickbooks_purchase_order_line_items",
            104,
            1148,
            1251,
            ",",
            "INSERT",
            "file_13,products",
            "NO",
            "NO",
            "READY",
            "",
        ],
        [
            15,
            "INGESTION_PO_LI_CLASSIFICATIONS_v1.csv",
            "quickbooks_purchase_order_line_item_classifications",
            416,
            3381,
            3796,
            ",",
            "INSERT",
            "file_14",
            "NO",
            "NO",
            "READY",
            "4 per LI",
        ],
    ]

    for row in manifest:
        ws.append(row)

    style_header(ws, len(headers))

    # Color by status
    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=12).value or ""
        if "FIXED" in status:
            fill = HIGH_FILL
        else:
            fill = GREEN_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_datasets_sheet(wb):
    ws = wb.create_sheet("DATASETS")
    headers = [
        "dataset_id",
        "name",
        "db_host",
        "db_port",
        "db_user",
        "db_name",
        "total_csvs",
        "total_rows",
        "pl_income",
        "pl_net",
        "pl_margin",
        "notes",
    ]
    ws.append(headers)

    ws.append(
        [
            DATASET_ID,
            "Keystone Construction",
            "tbx-postgres-staging.internal",
            5433,
            "unstable",
            "unstable",
            15,
            9406,
            "$10.93M",
            "+$2.93M",
            "26.8%",
            "parent=Keystone Construction, main_child=BlueCraft, secondary_child=Terra+5",
        ]
    )

    style_header(ws, len(headers))
    auto_width(ws)
    ws.freeze_panes = "A2"


def build_changelog_sheet(wb):
    ws = wb.create_sheet("CHANGELOG")
    headers = [
        "bug_id",
        "date",
        "file_affected",
        "issue",
        "root_cause",
        "fix",
        "rows_affected",
        "rule_added",
    ]
    ws.append(headers)

    bugs = [
        [
            "#1",
            "2026-01",
            "invoices",
            "Date clustering",
            "All dates in same week",
            "Seasonal redistribution",
            "335",
            "",
        ],
        [
            "#2",
            "2026-01",
            "invoices",
            "Customer ID collision",
            "IDs 51-75 clashed",
            "Remapped to 10616-10640",
            "335",
            "",
        ],
        [
            "#4",
            "2026-01",
            "ALL",
            "P&L negative ($16.3M costs vs $6.7M income)",
            "No budget cap",
            "Budget cap per CT",
            "ALL",
            "BIZ_003",
        ],
        [
            "#5",
            "2026-01",
            "ALL",
            "Dates > 1 year",
            "No 365-day limit",
            "Compressed to 365d",
            "ALL",
            "QB_003",
        ],
        [
            "#7",
            "2026-01",
            "invoices",
            "Payroll not accounted ($1.78M)",
            "Hidden main_child payroll",
            "Income boosted",
            "335",
            "BIZ_004",
        ],
        [
            "#8",
            "2026-01",
            "bills",
            "Cross-dataset FK contamination",
            "Referenced wrong dataset IDs",
            "Regenerated with valid IDs only",
            "96",
            "QB_009",
        ],
        [
            "#10",
            "2026-01",
            "invoices",
            "Monthly P&L negative Jan-Jun",
            "Uneven distribution",
            "v12 deterministic per-month",
            "335",
            "QB_014",
        ],
        [
            "#11",
            "2026-01",
            "invoices",
            "COGS not accounted",
            "38 inventory products ignored",
            "Service bias + income +25%",
            "335",
            "BIZ_002",
        ],
        [
            "#14",
            "2026-02",
            "time_entries",
            "Projects 23-26 had 0 TEs",
            "v4 missed them",
            "v5 supplement created",
            "3996",
            "",
        ],
        [
            "#16",
            "2026-02",
            "bills",
            "status='open' inconsistent",
            "Should be empty when paid",
            "Patched to empty string",
            "96",
            "RL_010",
        ],
        [
            "#17",
            "2026-02",
            "estimates",
            "Invalid statuses",
            "Had 'rejected','pending'",
            "Only accepted/cancelled",
            "75",
            "QB_018",
        ],
        [
            "#19",
            "2026-02",
            "bills,expenses",
            "Missing cross-table links",
            "No bill-PO or exp-est links",
            "31% bill-PO, 7.8% exp-est",
            "96+650",
            "RL_007,RL_008",
        ],
        [
            "#20",
            "2026-02-08",
            "bills",
            "TXN-13131 bill-PO product mismatch",
            "Bills linked by vendor not product",
            "6 reassigned, 9 unlinked, 3 improved",
            "18",
            "",
        ],
        [
            "#21",
            "2026-02-09",
            "expenses",
            "Employee FK violation",
            "Used range(1,51) but 6 gaps exist",
            "8 rows remapped to valid IDs",
            "8",
            "QB_020",
        ],
        [
            "#22",
            "2026-02-09",
            "time_entries",
            "Cancelled projects 25,26",
            "Script hardcoded [23,24,25,26]",
            "1800 rows removed",
            "1800",
            "QB_021",
        ],
        [
            "#24",
            "2026-02-09",
            "time_entries",
            "Dates outside project period",
            "No project date range check",
            "621 rows removed",
            "621",
            "QB_022",
        ],
    ]

    for row in bugs:
        ws.append(row)

    style_header(ws, len(headers))
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def main():
    print("=" * 60)
    print("BUILDING VALIDATION SPECIFICATION WORKBOOK")
    print("=" * 60)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("\nBuilding sheets...")
    build_rules_sheet(wb)
    print("  [1/9] RULES - validation rules")
    build_fk_map_sheet(wb)
    print("  [2/9] FK_MAP - foreign key mappings")
    build_employees_sheet(wb)
    print("  [3/9] EMPLOYEES - valid employee IDs")
    build_projects_sheet(wb)
    print("  [4/9] PROJECTS - project date ranges")
    build_products_sheet(wb)
    print("  [5/9] PRODUCTS - product catalog")
    build_enums_sheet(wb)
    print("  [6/9] ENUMS - valid enum values")
    build_csv_manifest_sheet(wb)
    print("  [7/9] CSV_MANIFEST - all 15 CSVs")
    build_datasets_sheet(wb)
    print("  [8/9] DATASETS - dataset config")
    build_changelog_sheet(wb)
    print("  [9/9] CHANGELOG - bug history")

    wb.save(OUTPUT)
    size = os.path.getsize(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Size: {size:,} bytes")
    print(f"Sheets: {len(wb.sheetnames)}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
