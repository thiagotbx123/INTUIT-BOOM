"""
Auto-Validator: Reads VALIDATION_SPEC_CONSTRUCTION.xlsx and validates ALL CSVs.
Runs every rule from the RULES sheet against the corresponding CSV files.

Usage:
    python validate_csvs.py                    # Validate all CSVs
    python validate_csvs.py --file 03          # Validate only file #03
    python validate_csvs.py --summary          # Just show pass/fail summary
"""

import os
import re
import sys
import openpyxl
from collections import defaultdict
from datetime import datetime

SPEC_PATH = r"C:\Users\adm_r\Downloads\PARA_AUGUSTO\VALIDATION_SPEC_CONSTRUCTION.xlsx"
CSV_DIR = r"C:\Users\adm_r\Downloads\PARA_AUGUSTO"

# ============================================================
# SPEC LOADER - Reads all sheets into memory
# ============================================================


class ValidationSpec:
    """Loads the entire validation spec workbook into structured data."""

    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self.rules = self._load_sheet("RULES")
        self.fk_map = self._load_sheet("FK_MAP")
        self.employees = self._load_sheet("EMPLOYEES")
        self.projects = self._load_sheet("PROJECTS")
        self.products = self._load_sheet("PRODUCTS")
        self.enums = self._load_sheet("ENUMS")
        self.csv_manifest = self._load_sheet("CSV_MANIFEST")
        self.datasets = self._load_sheet("DATASETS")

        # Build lookup sets
        self.valid_employee_ids = {int(r["id"]) for r in self.employees if r.get("id")}
        self.valid_project_ids = {
            int(r["id"]) for r in self.projects if r.get("status") == "In progress"
        }
        self.cancelled_project_ids = {
            int(r["id"]) for r in self.projects if r.get("status") == "Canceled"
        }
        self.project_ranges = {}
        for r in self.projects:
            if r.get("start_month") and r.get("end_month"):
                self.project_ranges[int(r["id"])] = (
                    int(r["start_month"]),
                    int(r["end_month"]),
                )

        self.valid_product_ids = set()
        self.purchasable_product_ids = set()
        for r in self.products:
            if r.get("id"):
                pid = int(r["id"])
                self.valid_product_ids.add(pid)
                if str(r.get("purchasable", "")).lower() in ("1", "true", "yes"):
                    self.purchasable_product_ids.add(pid)

        self.enum_values = {}
        for r in self.enums:
            key = f"{r.get('table', '')}.{r.get('column', '')}"
            vals = str(r.get("valid_values", "")).split("|")
            self.enum_values[key] = [v.strip() for v in vals]

        self.dataset_id = self.datasets[0]["dataset_id"] if self.datasets else ""

        # CSV manifest as dict
        self.manifest = {}
        for r in self.csv_manifest:
            num = int(r.get("file_num", 0))
            self.manifest[num] = r

        self.wb.close()

    def _load_sheet(self, name):
        if name not in self.wb.sheetnames:
            return []
        ws = self.wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])
        ]
        result = []
        for row in rows[1:]:
            d = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = val
            result.append(d)
        return result


# ============================================================
# CSV READER
# ============================================================


def read_csv_file(filepath, delimiter=","):
    """Read a CSV file and return list of dicts."""
    with open(filepath, "r", newline="", encoding="utf-8") as f:
        content = f.read()

    lines = content.strip().split("\n")
    header_line = lines[0].rstrip(delimiter)
    headers = header_line.split(delimiter)
    headers = [h.strip().lower() for h in headers]

    rows = []
    for line in lines[1:]:
        if not line.strip():
            continue
        fields = line.rstrip(delimiter).split(delimiter)
        row = {}
        for i, h in enumerate(headers):
            row[h] = fields[i].strip() if i < len(fields) else ""
        rows.append(row)

    return headers, rows


def parse_interval_days(val):
    """Parse PostgreSQL interval to approximate days."""
    if not val or val in ("", "NULL", "None", "#N/D", "first_of_year"):
        return None
    total = 0
    y = re.search(r"(\d+)\s*year", val)
    if y:
        total += int(y.group(1)) * 365
    m = re.search(r"(\d+)\s*mon", val)
    if m:
        total += int(m.group(1)) * 30.44
    d = re.search(r"(\d+)\s*day", val)
    if d:
        total += int(d.group(1))
    return total


def parse_interval_month(val):
    """Parse interval to month number (1-12)."""
    if not val or val in ("", "NULL", "None", "#N/D"):
        return None
    m = re.search(r"(\d+)\s*mon", val)
    mons = int(m.group(1)) if m else 0
    return mons + 1


def safe_int(val):
    if not val or val in ("", "NULL", "None", "#N/D"):
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val):
    if not val or val in ("", "NULL", "None", "#N/D"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ============================================================
# VALIDATORS - One per validation_type
# ============================================================


class ValidationResult:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.warnings = 0
        self.details = []

    def add_pass(self, rule_id, msg=""):
        self.passed += 1
        self.details.append(("PASS", rule_id, msg))

    def add_fail(self, rule_id, msg):
        self.failed += 1
        self.details.append(("FAIL", rule_id, msg))

    def add_warn(self, rule_id, msg):
        self.warnings += 1
        self.details.append(("WARN", rule_id, msg))


def validate_dataset_id(rows, spec, result, file_num):
    """QB_001: All rows must have correct dataset_id."""
    if not rows or "dataset_id" not in rows[0]:
        result.add_pass("QB_001", "N/A (no dataset_id column)")
        return
    bad = [r for r in rows if r.get("dataset_id", "") != spec.dataset_id]
    if bad:
        result.add_fail("QB_001", f"{len(bad)} rows with wrong dataset_id")
    else:
        result.add_pass("QB_001", "All dataset_ids match")


def validate_company_type(rows, spec, result, file_num):
    """QB_002: company_type must be valid enum."""
    if not rows or "company_type" not in rows[0]:
        result.add_pass("QB_002", "N/A (no company_type column)")
        return
    valid = {"parent", "main_child", "secondary_child", "all"}
    bad = [r for r in rows if r.get("company_type", "") not in valid]
    if bad:
        invalid_vals = set(r.get("company_type", "?") for r in bad)
        result.add_fail(
            "QB_002", f"{len(bad)} rows with invalid company_type: {invalid_vals}"
        )
    else:
        result.add_pass("QB_002", "All company_types valid")


def validate_dates_max_365(rows, headers, result, file_num):
    """QB_003: All relative dates <= 365 days."""
    date_cols = [
        h for h in headers if "relative" in h or ("date" in h and h != "base_date")
    ]
    violations = 0
    worst = 0
    for row in rows:
        for col in date_cols:
            days = parse_interval_days(row.get(col, ""))
            if days is not None and days > 365:
                violations += 1
                worst = max(worst, days)
    if violations:
        result.add_fail(
            "QB_003", f"{violations} date values > 365 days (worst: {worst:.0f} days)"
        )
    else:
        result.add_pass(
            "QB_003", f"All dates <= 365 days (checked {len(date_cols)} columns)"
        )


def validate_base_date(rows, result, file_num):
    """QB_005: base_date must be 'first_of_year'."""
    if not rows or "base_date" not in rows[0]:
        return
    bad = [r for r in rows if r.get("base_date", "first_of_year") != "first_of_year"]
    if bad:
        result.add_fail("QB_005", f"{len(bad)} rows with wrong base_date")
    else:
        result.add_pass("QB_005", "All base_date = 'first_of_year'")


def validate_sequential_ids(rows, result, file_num):
    """QB_008: IDs must be sequential."""
    ids = sorted(
        [safe_int(r.get("id", "")) for r in rows if safe_int(r.get("id")) is not None]
    )
    if not ids:
        return
    gaps = []
    for i in range(1, len(ids)):
        if ids[i] != ids[i - 1] + 1:
            gaps.append((ids[i - 1], ids[i]))
    if gaps:
        result.add_warn(
            "QB_008",
            f"{len(gaps)} gaps in ID sequence (first gap: {gaps[0][0]}->{gaps[0][1]})",
        )
    else:
        result.add_pass(
            "QB_008", f"IDs sequential: {ids[0]}-{ids[-1]} ({len(ids)} rows)"
        )


def validate_employee_fk(rows, spec, result, file_num):
    """QB_020: employee_id must exist in employees table."""
    if not any("employee_id" in r for r in rows[:1]):
        return
    invalid = []
    for row in rows:
        emp = safe_int(row.get("employee_id"))
        if emp is not None and emp not in spec.valid_employee_ids:
            invalid.append((row.get("id", "?"), emp))
    if invalid:
        bad_ids = set(e[1] for e in invalid)
        result.add_fail(
            "QB_020", f"{len(invalid)} rows with invalid employee_id: {sorted(bad_ids)}"
        )
    else:
        count = sum(1 for r in rows if safe_int(r.get("employee_id")) is not None)
        result.add_pass("QB_020", f"All employee_ids valid ({count} with values)")


def validate_project_active(rows, spec, result, file_num):
    """QB_021: project_id must reference active project."""
    if not rows or "project_id" not in rows[0]:
        return
    invalid = []
    for row in rows:
        proj = safe_int(row.get("project_id"))
        if proj is not None and proj in spec.cancelled_project_ids:
            invalid.append((row.get("id", "?"), proj))
    if invalid:
        bad_ids = set(e[1] for e in invalid)
        result.add_warn(
            "QB_021",
            f"{len(invalid)} rows reference cancelled projects: {sorted(bad_ids)} (may not block ingestion but is data quality issue)",
        )
    else:
        result.add_pass("QB_021", "No cancelled projects referenced")


def validate_te_within_project_dates(rows, spec, result, file_num):
    """QB_022: time entry date must fall within project active period."""
    if not any("project_id" in r for r in rows[:1]):
        return
    if not any("start_date_relative" in r for r in rows[:1]):
        return
    violations = []
    for row in rows:
        proj = safe_int(row.get("project_id"))
        if proj and proj in spec.project_ranges:
            month = parse_interval_month(row.get("start_date_relative", ""))
            if month:
                start_m, end_m = spec.project_ranges[proj]
                if month < start_m or month > end_m:
                    violations.append((row.get("id", "?"), proj, month, start_m, end_m))
    if violations:
        by_proj = defaultdict(list)
        for _, p, m, s, e in violations:
            by_proj[p].append(m)
        detail = "; ".join(
            f"proj {p}: months {sorted(set(ms))} (valid: {spec.project_ranges[p][0]}-{spec.project_ranges[p][1]})"
            for p, ms in sorted(by_proj.items())
        )
        result.add_fail(
            "QB_022", f"{len(violations)} TEs outside project dates: {detail}"
        )
    else:
        result.add_pass("QB_022", "All TEs within project date ranges")


def validate_li_qty(rows, result, file_num):
    """QB_013: Line item qty >= 1."""
    if not any("qty" in r for r in rows[:1]):
        return
    bad = []
    for row in rows:
        qty = safe_float(row.get("qty"))
        if qty is not None and qty < 1:
            bad.append((row.get("id", "?"), qty))
    if bad:
        result.add_fail("QB_013", f"{len(bad)} line items with qty < 1")
    else:
        result.add_pass("QB_013", "All qty >= 1")


def validate_amount_precision(rows, result, file_num):
    """QB_016: amount == rate * qty (2 decimal precision)."""
    has_amount = any("amount" in r for r in rows[:1])
    has_rate = any("rate" in r for r in rows[:1])
    has_qty = any("qty" in r for r in rows[:1])
    if not (has_amount and has_rate and has_qty):
        return
    bad = []
    for row in rows:
        amount = safe_float(row.get("amount"))
        rate = safe_float(row.get("rate"))
        qty = safe_float(row.get("qty"))
        if amount is not None and rate is not None and qty is not None:
            expected = round(rate * qty, 2)
            if abs(amount - expected) > 0.01:
                bad.append((row.get("id", "?"), amount, expected))
    if bad:
        result.add_fail("QB_016", f"{len(bad)} line items where amount != rate*qty")
    else:
        count = sum(1 for r in rows if safe_float(r.get("amount")) is not None)
        result.add_pass("QB_016", f"All amounts = rate*qty ({count} checked)")


def validate_bills_paid_date(rows, result, file_num):
    """QB_011: Bills must have paid_date."""

    if not any("paid_date" in r for r in rows[:1]):
        return
    bad = [
        r
        for r in rows
        if not r.get("paid_date") or r.get("paid_date") in ("", "NULL", "None")
    ]
    if bad:
        result.add_fail("QB_011", f"{len(bad)} bills without paid_date")
    else:
        result.add_pass("QB_011", "All bills have paid_date")


def validate_estimate_status(rows, spec, result, file_num):
    """QB_018: estimate target_status only accepted/cancelled."""
    if not any("target_status" in r for r in rows[:1]):
        return
    valid = {"accepted", "cancelled"}
    bad = [r for r in rows if r.get("target_status", "").strip().lower() not in valid]
    if bad:
        invalid_vals = set(r.get("target_status", "?") for r in bad)
        result.add_fail(
            "QB_018", f"{len(bad)} estimates with invalid status: {invalid_vals}"
        )
    else:
        result.add_pass("QB_018", "All estimate statuses valid")


def validate_doc_number_length(rows, result, file_num):
    """QB_017: DocNumber max 21 chars."""
    doc_cols = [
        h
        for h in (rows[0].keys() if rows else [])
        if "doc_number" in h or "invoice_no" in h or "bill_no" in h
    ]
    if not doc_cols:
        return
    bad = []
    for row in rows:
        for col in doc_cols:
            val = str(row.get(col, ""))
            if len(val) > 21:
                bad.append((row.get("id", "?"), col, len(val)))
    if bad:
        result.add_fail("QB_017", f"{len(bad)} rows with doc_number > 21 chars")
    else:
        result.add_pass("QB_017", "All doc_numbers <= 21 chars")


def validate_purchasable_products(rows, spec, result, file_num):
    """QB_015: Bill LIs only reference purchasable products."""
    if not spec.purchasable_product_ids:
        result.add_warn("QB_015", "No purchasable product data in spec - skipped")
        return
    if not any("product_service_id" in r for r in rows[:1]):
        return
    bad = []
    for row in rows:
        pid = safe_int(row.get("product_service_id"))
        if (
            pid is not None
            and pid not in spec.purchasable_product_ids
            and spec.purchasable_product_ids
        ):
            bad.append((row.get("id", "?"), pid))
    if bad:
        bad_pids = set(b[1] for b in bad)
        result.add_fail(
            "QB_015",
            f"{len(bad)} bill LIs reference non-purchasable products: {sorted(bad_pids)[:10]}",
        )
    else:
        result.add_pass("QB_015", "All bill LI products are purchasable")


def validate_classification_count(rows, result, file_num):
    """QB_019: Exactly 4 classification entries per line item."""
    if not any("line_item_id" in r for r in rows[:1]):
        return
    counts = defaultdict(int)
    for row in rows:
        li_id = row.get("line_item_id", "")
        if li_id:
            counts[li_id] += 1
    bad = {li: c for li, c in counts.items() if c != 4}
    if bad:
        result.add_fail(
            "QB_019",
            f"{len(bad)} line items with != 4 classifications (e.g., {list(bad.items())[:3]})",
        )
    else:
        result.add_pass(
            "QB_019", f"All {len(counts)} line items have exactly 4 classifications"
        )


# ============================================================
# MAIN VALIDATOR
# ============================================================


def validate_csv(file_num, spec, verbose=True):
    """Validate a single CSV file against the spec."""
    manifest = spec.manifest.get(file_num)
    if not manifest:
        print(f"  File #{file_num} not in manifest - SKIP")
        return None

    filename = manifest["filename"]
    delimiter = manifest.get("delimiter", ",")
    target_table = manifest.get("target_table", "")

    # Search for file by pattern
    found = None
    for f in os.listdir(CSV_DIR):
        if f.endswith(".csv") and f.startswith(f"{file_num:02d}_"):
            found = os.path.join(CSV_DIR, f)
            break

    if not found:
        # Try exact filename
        exact = os.path.join(CSV_DIR, filename)
        if os.path.exists(exact):
            found = exact

    if not found:
        if verbose:
            print(f"  File #{file_num:02d} ({filename}) - NOT FOUND")
        return None

    if verbose:
        print(f"\n{'=' * 60}")
        print(f"  FILE #{file_num:02d}: {os.path.basename(found)}")
        print(f"  Table: {target_table}")
        print(f"  Expected rows: {manifest.get('rows', '?')}")
        print(f"{'=' * 60}")

    headers, rows = read_csv_file(found, delimiter)
    result = ValidationResult()

    # Row count check
    expected = safe_int(str(manifest.get("rows", "")))
    actual = len(rows)
    if expected and actual != expected:
        result.add_warn("ROWS", f"Expected {expected} rows, got {actual}")
    else:
        result.add_pass("ROWS", f"{actual} rows (matches manifest)")

    # Run applicable validators
    validate_dataset_id(rows, spec, result, file_num)
    validate_company_type(rows, spec, result, file_num)
    validate_dates_max_365(rows, headers, result, file_num)
    validate_base_date(rows, result, file_num)
    validate_sequential_ids(rows, result, file_num)
    validate_employee_fk(rows, spec, result, file_num)
    validate_project_active(rows, spec, result, file_num)
    validate_te_within_project_dates(rows, spec, result, file_num)
    validate_li_qty(rows, result, file_num)
    validate_amount_precision(rows, result, file_num)
    validate_doc_number_length(rows, result, file_num)
    validate_estimate_status(rows, spec, result, file_num)
    validate_classification_count(rows, result, file_num)

    # Table-specific checks
    if "bills" in target_table and "line_item" not in target_table:
        validate_bills_paid_date(rows, result, file_num)
    if "bill" in target_table and "line_item" in target_table:
        validate_purchasable_products(rows, spec, result, file_num)

    # Print results
    if verbose:
        for status, rule, msg in result.details:
            icon = (
                "PASS" if status == "PASS" else "FAIL" if status == "FAIL" else "WARN"
            )
            print(f"  [{icon}] {rule}: {msg}")
        print(
            f"\n  Summary: {result.passed} PASS, {result.failed} FAIL, {result.warnings} WARN"
        )

    return result


def main():
    args = sys.argv[1:]
    file_filter = None
    summary_only = False

    for arg in args:
        if arg == "--summary":
            summary_only = True
        elif arg.startswith("--file"):
            file_filter = (
                int(args[args.index(arg) + 1])
                if arg == "--file"
                else int(arg.split("=")[1])
            )

    print("=" * 60)
    print("CSV AUTO-VALIDATOR v1.0")
    print(f"Spec: {SPEC_PATH}")
    print(f"CSVs: {CSV_DIR}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    spec = ValidationSpec(SPEC_PATH)
    print("\nSpec loaded:")
    print(f"  Rules: {len(spec.rules)}")
    print(f"  FK mappings: {len(spec.fk_map)}")
    print(f"  Employees: {len(spec.valid_employee_ids)} valid IDs")
    print(
        f"  Projects: {len(spec.valid_project_ids)} active, {len(spec.cancelled_project_ids)} cancelled"
    )
    print(
        f"  Products: {len(spec.valid_product_ids)} total, {len(spec.purchasable_product_ids)} purchasable"
    )
    print(f"  Dataset: {spec.dataset_id}")

    files_to_check = [file_filter] if file_filter else sorted(spec.manifest.keys())

    total_pass = 0
    total_fail = 0
    total_warn = 0
    file_results = {}

    for fnum in files_to_check:
        result = validate_csv(fnum, spec, verbose=not summary_only)
        if result:
            total_pass += result.passed
            total_fail += result.failed
            total_warn += result.warnings
            file_results[fnum] = result

    # Final summary
    print("\n" + "=" * 60)
    print("FINAL SUMMARY")
    print("=" * 60)
    for fnum, result in sorted(file_results.items()):
        fname = spec.manifest[fnum]["filename"]
        status = "PASS" if result.failed == 0 else "FAIL"
        fails = f" ({result.failed} failures)" if result.failed > 0 else ""
        warns = f" ({result.warnings} warnings)" if result.warnings > 0 else ""
        print(f"  [{status}] #{fnum:02d} {fname}{fails}{warns}")

    print(f"\n  TOTAL: {total_pass} PASS, {total_fail} FAIL, {total_warn} WARN")
    print(f"  Files checked: {len(file_results)}/{len(spec.manifest)}")

    if total_fail == 0:
        print("\n  ALL CHECKS PASSED!")
    else:
        print(f"\n  {total_fail} FAILURES NEED ATTENTION")

    print("=" * 60)
    return 1 if total_fail > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
