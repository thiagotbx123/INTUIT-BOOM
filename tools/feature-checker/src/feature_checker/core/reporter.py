"""Report generation for health check results."""

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..utils.config import get_config


class Reporter:
    """
    Generates reports from health check results.

    Supports:
    - Excel spreadsheets
    - JSON files
    - Markdown reports
    """

    def __init__(self, product: str, project: str = None):
        """
        Initialize reporter.

        Args:
            product: Product name
            project: Project name
        """
        self.config = get_config()
        self.product = product
        self.project = project or "default"
        self.output_dir = self.config.reports_dir

    def generate(self, results: List[Any], format: str = "excel") -> Path:
        """
        Generate report in specified format.

        Args:
            results: List of CheckResult objects
            format: Output format (excel, json, markdown)

        Returns:
            Path to generated report
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"health_check_{self.product}_{self.project}_{timestamp}"

        if format == "excel":
            return self._generate_excel(results, filename)
        elif format == "json":
            return self._generate_json(results, filename)
        elif format == "markdown":
            return self._generate_markdown(results, filename)
        else:
            raise ValueError(f"Unknown format: {format}")

    def _generate_excel(self, results: List[Any], filename: str) -> Path:
        """Generate Excel report."""
        filepath = self.output_dir / f"{filename}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Health Check Results"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        pass_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        fail_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        partial_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "#",
            "Check ID",
            "Name",
            "Status",
            "Message",
            "Duration (ms)",
            "Screenshot",
            "Timestamp",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for i, result in enumerate(results, 1):
            row = i + 1

            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=result.check_id).border = thin_border
            ws.cell(row=row, column=3, value=result.name).border = thin_border

            status_cell = ws.cell(row=row, column=4, value=result.status)
            status_cell.border = thin_border
            status_cell.alignment = Alignment(horizontal="center")
            if result.status == "PASS":
                status_cell.fill = pass_fill
                status_cell.font = Font(color="FFFFFF", bold=True)
            elif result.status == "FAIL":
                status_cell.fill = fail_fill
                status_cell.font = Font(color="FFFFFF", bold=True)
            elif result.status == "PARTIAL":
                status_cell.fill = partial_fill
                status_cell.font = Font(color="FFFFFF", bold=True)

            ws.cell(row=row, column=5, value=result.message).border = thin_border
            ws.cell(row=row, column=6, value=result.duration_ms).border = thin_border

            screenshot_cell = ws.cell(row=row, column=7)
            if result.screenshot:
                screenshot_cell.value = result.screenshot.name
                screenshot_cell.hyperlink = str(result.screenshot)
            screenshot_cell.border = thin_border

            ws.cell(row=row, column=8, value=result.timestamp).border = thin_border

        # Adjust column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 40
        ws.column_dimensions["H"].width = 25

        # Summary sheet
        summary_ws = wb.create_sheet("Summary")

        summary_data = self._calculate_summary(results)

        summary_ws.cell(row=1, column=1, value="Health Check Summary").font = Font(
            bold=True, size=14
        )
        summary_ws.cell(row=3, column=1, value="Product:")
        summary_ws.cell(row=3, column=2, value=self.product)
        summary_ws.cell(row=4, column=1, value="Project:")
        summary_ws.cell(row=4, column=2, value=self.project)
        summary_ws.cell(row=5, column=1, value="Generated:")
        summary_ws.cell(row=5, column=2, value=datetime.now().isoformat())

        summary_ws.cell(row=7, column=1, value="Results").font = Font(bold=True)
        summary_ws.cell(row=8, column=1, value="Total Checks:")
        summary_ws.cell(row=8, column=2, value=summary_data["total"])
        summary_ws.cell(row=9, column=1, value="Passed:")
        summary_ws.cell(row=9, column=2, value=summary_data["PASS"])
        summary_ws.cell(row=10, column=1, value="Failed:")
        summary_ws.cell(row=10, column=2, value=summary_data["FAIL"])
        summary_ws.cell(row=11, column=1, value="Partial:")
        summary_ws.cell(row=11, column=2, value=summary_data["PARTIAL"])

        wb.save(filepath)
        return filepath

    def _generate_json(self, results: List[Any], filename: str) -> Path:
        """Generate JSON report."""
        filepath = self.output_dir / f"{filename}.json"

        data = {
            "metadata": {
                "product": self.product,
                "project": self.project,
                "generated": datetime.now().isoformat(),
                "total_checks": len(results),
            },
            "summary": self._calculate_summary(results),
            "results": [
                {
                    "check_id": r.check_id,
                    "name": r.name,
                    "status": r.status,
                    "message": r.message,
                    "duration_ms": r.duration_ms,
                    "screenshot": str(r.screenshot) if r.screenshot else None,
                    "timestamp": r.timestamp,
                }
                for r in results
            ],
        }

        with open(filepath, "w") as f:
            json.dump(data, f, indent=2)

        return filepath

    def _generate_markdown(self, results: List[Any], filename: str) -> Path:
        """Generate Markdown report."""
        filepath = self.output_dir / f"{filename}.md"

        summary = self._calculate_summary(results)

        lines = [
            "# Health Check Report",
            "",
            f"**Product:** {self.product}",
            f"**Project:** {self.project}",
            f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## Summary",
            "",
            "| Metric | Value |",
            "|--------|-------|",
            f"| Total Checks | {summary['total']} |",
            f"| Passed | {summary['PASS']} |",
            f"| Failed | {summary['FAIL']} |",
            f"| Partial | {summary['PARTIAL']} |",
            "",
            "## Results",
            "",
            "| Status | Check | Message |",
            "|--------|-------|---------|",
        ]

        for r in results:
            icon = {"PASS": "✅", "FAIL": "❌", "PARTIAL": "⚠️", "SKIP": "⏭️"}.get(r.status, "❓")
            lines.append(f"| {icon} {r.status} | {r.name} | {r.message} |")

        # Failed checks detail
        failed = [r for r in results if r.status == "FAIL"]
        if failed:
            lines.extend(
                [
                    "",
                    "## Failed Checks Detail",
                    "",
                ]
            )
            for r in failed:
                lines.extend(
                    [
                        f"### {r.name}",
                        "",
                        f"- **Check ID:** {r.check_id}",
                        f"- **Message:** {r.message}",
                        f"- **Duration:** {r.duration_ms}ms",
                        "",
                    ]
                )

        with open(filepath, "w") as f:
            f.write("\n".join(lines))

        return filepath

    def _calculate_summary(self, results: List[Any]) -> Dict[str, int]:
        """Calculate summary statistics."""
        summary = {"total": len(results), "PASS": 0, "FAIL": 0, "PARTIAL": 0, "SKIP": 0}
        for r in results:
            summary[r.status] = summary.get(r.status, 0) + 1
        return summary
