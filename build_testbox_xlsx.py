"""Build TESTBOX_ACCOUNTS.xlsx from parsed markdown data — database-style normalized tables."""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MD_PATH = r"C:\Users\adm_r\Clients\intuit-boom\knowledge-base\access\TESTBOX_ACCOUNTS.md"
OUT_PATH = r"C:\Users\adm_r\Downloads\TESTBOX_ACCOUNTS.xlsx"


def parse_md():
    with open(MD_PATH, "r", encoding="utf-8") as f:
        text = f.read()

    accounts = []
    companies_rows = []

    # Split by ### headers (each account)
    blocks = re.split(r"\n### ", text)

    for block in blocks[1:]:  # skip everything before first ###
        lines = block.strip().split("\n")
        account_name = lines[0].strip()

        # Extract field-value pairs from the table
        fields = {}
        for line in lines:
            m = re.match(r"\|\s*\*\*(.+?)\*\*\s*\|\s*`?(.+?)`?\s*\|", line)
            if m:
                key = m.group(1).strip()
                val = m.group(2).strip()
                fields[key] = val

        email = fields.get("Email", "")
        password = fields.get("Password", "")
        totp = fields.get("TOTP Token", "")
        env = fields.get("Environment", "")
        state = fields.get("State", "")
        created = fields.get("Created", "")
        failed = fields.get("Failed", "")
        attempts = fields.get("Attempts", "")
        yoy = fields.get("YoY Growth", "")
        new_ingest = fields.get("New Ingest", "")
        telephone = fields.get("Telephone", "")
        sso = fields.get("SSO", "")

        # Determine dataset from section header
        # Look backwards from block position to find ## header
        dataset = ""
        block_pos = text.find("### " + account_name)
        if block_pos > 0:
            before = text[:block_pos]
            headers = re.findall(r"\n## (.+)", before)
            if headers:
                dataset = headers[-1].strip()
                if dataset in ("Summary",):
                    dataset = ""
                if dataset == "(No Dataset)":
                    dataset = ""

        # Parse companies
        company_list = []
        in_companies = False
        for line in lines:
            if "Company ID" in line and "Role" in line:
                in_companies = True
                continue
            if in_companies and line.startswith("|"):
                if "---" in line:
                    continue
                parts = [p.strip().strip("`") for p in line.split("|") if p.strip()]
                if len(parts) == 2:
                    company_list.append({"id": parts[0], "role": parts[1]})

        # Count parent/child
        parent_ids = [c["id"] for c in company_list if c["role"] == "parent"]
        child_ids = [c["id"] for c in company_list if c["role"] == "child"]
        other_ids = [c["id"] for c in company_list if c["role"] not in ("parent", "child")]

        account_id = len(accounts) + 1

        accounts.append(
            {
                "id": account_id,
                "dataset": dataset,
                "environment": env,
                "email": email,
                "password": password,
                "totp_token": totp,
                "state": state,
                "created": created,
                "failed": failed,
                "attempts": attempts,
                "num_companies": len(company_list),
                "parent_id": ", ".join(parent_ids) if parent_ids else "",
                "child_ids": ", ".join(child_ids) if child_ids else "",
                "other_ids": ", ".join(other_ids) if other_ids else "",
                "yoy_growth": yoy,
                "new_ingest": new_ingest,
                "telephone": telephone,
                "sso": sso,
            }
        )

        for comp in company_list:
            companies_rows.append(
                {
                    "account_id": account_id,
                    "email": email,
                    "dataset": dataset,
                    "company_id": comp["id"],
                    "role": comp["role"],
                }
            )

    return accounts, companies_rows


def style_header(ws, row, num_cols):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="1F4E79"))


def style_data(ws, start_row, end_row, num_cols):
    data_font = Font(name="Segoe UI", size=9)
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    thin = Side(style="thin", color="E0E0E0")
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if (row - start_row) % 2 == 1:
                cell.fill = alt_fill


def auto_width(ws, num_cols, max_width=45):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, max_width)


def build_xlsx(accounts, companies_rows):
    wb = Workbook()

    # === Sheet 1: ACCOUNTS (master table) ===
    ws1 = wb.active
    ws1.title = "Accounts"

    headers1 = [
        "ID",
        "Dataset",
        "Environment",
        "Email",
        "Password",
        "TOTP Token",
        "State",
        "Created",
        "Failed",
        "Attempts",
        "# Companies",
        "Parent ID",
        "Child IDs",
        "YoY Growth",
        "New Ingest",
        "Telephone",
        "SSO",
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    for i, acc in enumerate(accounts, 2):
        ws1.cell(row=i, column=1, value=acc["id"])
        ws1.cell(row=i, column=2, value=acc["dataset"])
        ws1.cell(row=i, column=3, value=acc["environment"])
        ws1.cell(row=i, column=4, value=acc["email"])
        ws1.cell(row=i, column=5, value=acc["password"])
        ws1.cell(row=i, column=6, value=acc["totp_token"])
        ws1.cell(row=i, column=7, value=acc["state"])
        ws1.cell(row=i, column=8, value=acc["created"])
        ws1.cell(row=i, column=9, value=acc["failed"])
        ws1.cell(row=i, column=10, value=acc["attempts"])
        ws1.cell(row=i, column=11, value=acc["num_companies"])
        ws1.cell(row=i, column=12, value=acc["parent_id"])
        ws1.cell(row=i, column=13, value=acc["child_ids"])
        ws1.cell(row=i, column=14, value=acc["yoy_growth"])
        ws1.cell(row=i, column=15, value=acc["new_ingest"])
        ws1.cell(row=i, column=16, value=acc["telephone"])
        ws1.cell(row=i, column=17, value=acc["sso"])

    style_header(ws1, 1, len(headers1))
    style_data(ws1, 2, len(accounts) + 1, len(headers1))
    auto_width(ws1, len(headers1))
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(accounts) + 1}"
    ws1.freeze_panes = "A2"

    # === Sheet 2: COMPANIES (normalized — 1 row per company per account) ===
    ws2 = wb.create_sheet("Companies")

    headers2 = ["Account ID", "Email", "Dataset", "Company ID", "Role"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)

    for i, comp in enumerate(companies_rows, 2):
        ws2.cell(row=i, column=1, value=comp["account_id"])
        ws2.cell(row=i, column=2, value=comp["email"])
        ws2.cell(row=i, column=3, value=comp["dataset"])
        ws2.cell(row=i, column=4, value=comp["company_id"])
        ws2.cell(row=i, column=5, value=comp["role"])

    style_header(ws2, 1, len(headers2))
    style_data(ws2, 2, len(companies_rows) + 1, len(headers2))
    auto_width(ws2, len(headers2))
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(companies_rows) + 1}"
    ws2.freeze_panes = "A2"

    # === Sheet 3: SUMMARY (pivot by dataset) ===
    ws3 = wb.create_sheet("Summary")

    # Count by dataset
    dataset_counts = {}
    for acc in accounts:
        ds = acc["dataset"] or "(no dataset)"
        dataset_counts[ds] = dataset_counts.get(ds, 0) + 1

    headers3 = ["Dataset", "Accounts", "Status"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)

    sorted_ds = sorted(dataset_counts.items(), key=lambda x: (-x[1], x[0]))
    for i, (ds, count) in enumerate(sorted_ds, 2):
        ws3.cell(row=i, column=1, value=ds)
        ws3.cell(row=i, column=2, value=count)
        # Count active (Wait For Next Activity) vs failed
        active = sum(
            1 for a in accounts if (a["dataset"] or "(no dataset)") == ds and a["state"] == "Wait For Next Activity"
        )
        ws3.cell(row=i, column=3, value=f"{active} active / {count - active} other")

    total_row = len(sorted_ds) + 2
    ws3.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Segoe UI", bold=True, size=10)
    ws3.cell(row=total_row, column=2, value=len(accounts)).font = Font(name="Segoe UI", bold=True, size=10)

    style_header(ws3, 1, len(headers3))
    style_data(ws3, 2, total_row, len(headers3))
    auto_width(ws3, len(headers3))

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"  Accounts: {len(accounts)} rows")
    print(f"  Companies: {len(companies_rows)} rows")
    print(f"  Datasets: {len(dataset_counts)}")


if __name__ == "__main__":
    accounts, companies = parse_md()
    build_xlsx(accounts, companies)
