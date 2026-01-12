"""
Winter Release FY26 - Validation Matrix v4
- All 29 features in ALL environment sheets
- Calibri 11, centered
- Columns: Ref, Feature, Category, Client Notes, Evidence Link, Status
- Step by Step: single rich manual with all details and alt checks
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Styles
CALIBRI_11 = Font(name="Calibri", size=11)
CALIBRI_11_BOLD = Font(name="Calibri", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# All 29 features
FEATURES = [
    {
        "ref": "WR-001",
        "name": "Accounting AI",
        "category": "AI Agents",
        "client_note": "Accounting AI is active with auto-categorization enabled.",
        "step_by_step": """ACCOUNTING AI - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
AI helps categorize bank transactions automatically. Look for sparkle icons next to suggested categories.

NAVIGATION PATH:
Left Menu > Transactions > Bank transactions

PRIMARY VALIDATION:
1. Click "Transactions" in the left menu
2. Click "Bank transactions"
3. Wait for page to fully load (10-15 seconds)
4. Look at the transaction list
5. Find transactions with a SPARKLE ICON (small star shape)
6. This sparkle means AI is suggesting a category
7. Take screenshot showing sparkle icons

ALTERNATIVE CHECK 1:
Check Settings > Bank transactions > Auto-categorization enabled

ALTERNATIVE CHECK 2:
Run SQL on database: SELECT COUNT(*) FROM bank_transactions WHERE ai_suggested=1

ALTERNATIVE CHECK 3:
Open Intuit Assist and type: "categorize my transactions"

WHAT TO LOOK FOR:
- Sparkle icon next to transaction category = AI active
- "Suggested" label = AI recommendation
- Auto-categorized transactions = AI working

SCREENSHOT CHECKLIST:
[ ] Bank transactions page visible
[ ] At least one sparkle icon visible
[ ] Transaction list showing

PASS CRITERIA:
- Sparkle icons visible on transactions = PASS
- AI suggestions appearing = PASS
- No sparkles but page works = Check if there are uncategorized transactions""",
    },
    {
        "ref": "WR-002",
        "name": "Sales Tax AI",
        "category": "AI Agents",
        "client_note": "Sales Tax AI provides pre-filing checks and discrepancy alerts.",
        "step_by_step": """SALES TAX AI - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
AI checks for errors before you file sales tax. Shows warnings and suggestions.

NAVIGATION PATH:
Left Menu > Taxes > Sales tax

PRIMARY VALIDATION:
1. Click "Taxes" in the left menu
2. Click "Sales tax"
3. Wait for Sales Tax Center to load
4. Look for "Filing Pre-Check" section
5. Look for AI warnings (usually in blue or yellow boxes)
6. Look for "Discrepancy" alerts
7. Take screenshot of pre-check area

ALTERNATIVE CHECK 1:
Check if 'Automated Sales Tax' is enabled in Settings > Sales > Sales tax

ALTERNATIVE CHECK 2:
Look for tax rate suggestions when creating an invoice

ALTERNATIVE CHECK 3:
Check Taxes > Payments for AI-generated recommendations

WHAT TO LOOK FOR:
- "Pre-Check" or "Review before filing" = AI check
- Warning icons with suggestions = AI working
- Discrepancy alerts = AI finding issues

SCREENSHOT CHECKLIST:
[ ] Sales Tax Center visible
[ ] Pre-check section visible
[ ] Any warnings or alerts visible

PASS CRITERIA:
- Pre-check features visible = PASS
- AI warnings showing = PASS""",
    },
    {
        "ref": "WR-003",
        "name": "Project Management AI",
        "category": "AI Agents",
        "client_note": "Project Management AI tracks profitability and provides insights.",
        "step_by_step": """PROJECT MANAGEMENT AI - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
AI helps manage projects by showing profitability, suggesting budgets, and providing insights.

NAVIGATION PATH:
Left Menu > Projects > [Select any project]

PRIMARY VALIDATION:
1. Click "Projects" in the left menu
2. Wait for project list to load
3. Click on any project name to open it
4. Wait for project details (10-20 seconds)
5. Look for:
   - Profitability percentage or graph
   - "Insights" section
   - Sparkle icons for AI suggestions
   - Budget vs Actual comparison
6. Scroll down to see all sections
7. Take screenshot showing project insights

ALTERNATIVE CHECK 1:
Check Projects list for profitability column showing percentages

ALTERNATIVE CHECK 2:
Go to Reports and look for 'Project Profitability' report

ALTERNATIVE CHECK 3:
Run SQL: SELECT status, COUNT(*) FROM projects GROUP BY status

WHAT TO LOOK FOR:
- Profitability % = AI calculated
- Sparkle icons = AI suggestions
- "Insights" tab/section = AI analysis
- Budget recommendations = AI working

SCREENSHOT CHECKLIST:
[ ] Project name visible at top
[ ] Profitability info visible
[ ] Any AI insights or sparkles visible

PASS CRITERIA:
- Profitability tracking visible = PASS
- AI insights showing = PASS""",
    },
    {
        "ref": "WR-004",
        "name": "Finance AI",
        "category": "AI Agents",
        "client_note": "Finance AI provides automated financial insights and recommendations.",
        "step_by_step": """FINANCE AI - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
AI provides financial summaries, insights, and recommendations.

NAVIGATION PATH:
Dashboard (Homepage) OR Left Menu > Reports

PRIMARY VALIDATION:
1. Go to Dashboard (click QuickBooks logo top left)
2. Look for "Intuit Assist" icon (chat bubble, usually bottom right)
3. Look for "Business overview" section with insights
4. If not on Dashboard, go to Reports
5. Look for AI-enhanced report sections
6. Check for "Insights" or "Recommendations" boxes
7. Take screenshot of any AI features

ALTERNATIVE CHECK 1:
Check Dashboard widgets for AI-generated insights cards

ALTERNATIVE CHECK 2:
Open Intuit Assist and ask: "Show my revenue trend"

ALTERNATIVE CHECK 3:
Go to Reports > Business Snapshot and look for AI summary

WHAT TO LOOK FOR:
- Intuit Assist chat icon
- Financial insights on Dashboard
- AI recommendations

SCREENSHOT CHECKLIST:
[ ] Dashboard or Reports visible
[ ] Intuit Assist icon or chat visible
[ ] Any AI insights visible

PASS CRITERIA:
- Intuit Assist icon visible = PASS
- Financial insights showing = PASS""",
    },
    {
        "ref": "WR-005",
        "name": "Solutions Specialist",
        "category": "AI Agents",
        "client_note": "Solutions Specialist provides personalized AI recommendations.",
        "step_by_step": """SOLUTIONS SPECIALIST - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
AI provides personalized recommendations in Business Feed based on your business.

NAVIGATION PATH:
Dashboard > Business Feed (right side)

PRIMARY VALIDATION:
1. Go to Dashboard (click QuickBooks logo)
2. Look at the RIGHT side of the Dashboard
3. Find "Business Feed" or "Activity Feed" section
4. Scroll through the feed items
5. Look for:
   - "Recommended for you" items
   - Dimension widgets
   - IES-exclusive suggestions
   - Personalized tips
6. Take screenshot of Business Feed

ALTERNATIVE CHECK 1:
Check for 'Discover' or 'Explore' section on Dashboard

ALTERNATIVE CHECK 2:
Look for industry-specific tips in the feed

ALTERNATIVE CHECK 3:
Check Apps page for recommended apps section

WHAT TO LOOK FOR:
- Personalized recommendations = AI working
- Dimension widgets = IES feature active
- "Suggested" items = AI curated

SCREENSHOT CHECKLIST:
[ ] Dashboard visible
[ ] Business Feed section visible
[ ] At least one recommendation visible

PASS CRITERIA:
- Business Feed with recommendations = PASS""",
    },
    {
        "ref": "WR-006",
        "name": "Customer Agent",
        "category": "AI Agents",
        "client_note": "Customer Agent allows importing leads from Gmail/Outlook.",
        "step_by_step": """CUSTOMER AGENT - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
AI helps manage customer leads by importing from email and suggesting actions.

NAVIGATION PATH:
Left Menu > Sales > Customers > Leads tab

PRIMARY VALIDATION:
1. Click "Sales" in the left menu
2. Click "Customers"
3. Wait for Customers page to load
4. Look for "Leads" tab at the top
5. Click on "Leads" tab
6. Look for email import options:
   - "Import from Gmail"
   - "Import from Outlook"
   - "Connect email" button
7. Take screenshot showing Leads features

ALTERNATIVE CHECK 1:
Check + Create menu for 'Lead' option

ALTERNATIVE CHECK 2:
Look for lead pipeline view on Customers page

ALTERNATIVE CHECK 3:
Run SQL: SELECT COUNT(*) FROM customers WHERE lead_status IS NOT NULL

WHAT TO LOOK FOR:
- Leads tab on Customers page
- Email import options (Gmail, Outlook)
- Lead management features

SCREENSHOT CHECKLIST:
[ ] Customers page visible
[ ] Leads tab visible (if exists)
[ ] Email import options visible

PASS CRITERIA:
- Leads tab visible = PASS
- Email import options = PASS""",
    },
    {
        "ref": "WR-007",
        "name": "Intuit Intelligence (Omni)",
        "category": "AI Agents",
        "client_note": "Intuit Intelligence provides conversational AI for business queries.",
        "step_by_step": """INTUIT INTELLIGENCE - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Conversational AI assistant (chatbot) that answers questions about your business.

NAVIGATION PATH:
Look for chat bubble icon (bottom right of screen)

PRIMARY VALIDATION:
1. Look at BOTTOM RIGHT corner of any QBO page
2. Find a chat bubble icon or "Intuit Assist" button
3. May also be a sparkle icon
4. Click the icon to open chat window
5. If chat opens, type: "What was my revenue last month?"
6. Wait for AI response
7. Take screenshot of chat window

ALTERNATIVE CHECK 1:
Check Help menu for 'Ask Intuit Assist' option

ALTERNATIVE CHECK 2:
Look for AI icon in the search bar

ALTERNATIVE CHECK 3:
Check Settings for Intuit Assist toggle/enable option

NOTE: This feature may be in beta. Not visible in all environments.

WHAT TO LOOK FOR:
- Intuit Assist chat bubble icon
- Chat window that opens
- AI responses to questions

SCREENSHOT CHECKLIST:
[ ] Intuit Assist icon visible OR
[ ] Chat window open
[ ] AI response showing (if tested)

PASS CRITERIA:
- Intuit Assist icon visible = PASS
- Chat window opens = PASS
- AI responds = PASS""",
    },
    {
        "ref": "WR-008",
        "name": "Conversational BI",
        "category": "AI Agents",
        "client_note": "Conversational BI allows natural language queries with visual reports.",
        "step_by_step": """CONVERSATIONAL BI - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Ask questions in natural language, get reports and charts as answers.

NAVIGATION PATH:
Intuit Assist chat > Ask data question

PRIMARY VALIDATION:
1. Open Intuit Assist (see WR-007 steps)
2. In the chat, type a data question:
   - "Show me revenue by month"
   - "What are my top 5 customers?"
   - "Compare expenses this year vs last"
3. Press Enter to send
4. Wait for response (10-20 seconds)
5. Look for CHARTS or TABLES in response
6. Take screenshot of question + visual response

ALTERNATIVE CHECK 1:
Try different question types (revenue, expenses, customers)

ALTERNATIVE CHECK 2:
Check if results can be saved as a report

ALTERNATIVE CHECK 3:
Test date range questions like "revenue in Q4"

WHAT TO LOOK FOR:
- Chart generated from question = BI working
- Table with data = BI working
- Text-only response = Partial (no visualization)

SCREENSHOT CHECKLIST:
[ ] Question visible in chat
[ ] Visual response (chart/table) visible
[ ] Data appears accurate

PASS CRITERIA:
- Visual response generated = PASS
- Data chart showing = PASS""",
    },
    {
        "ref": "WR-009",
        "name": "Custom KPIs",
        "category": "Reporting",
        "client_note": "Custom KPIs with pre-built metrics and custom formula capability.",
        "step_by_step": """CUSTOM KPIs - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Create and view Key Performance Indicators with custom formulas.

NAVIGATION PATH:
Left Menu > Reports > KPI Scorecard
OR Direct URL: qbo.intuit.com/app/business-intelligence/kpi-scorecard

PRIMARY VALIDATION:
1. Click "Reports" in left menu
2. Look for "KPIs" or "KPI Scorecard" in list
3. Click on it (or use direct URL above)
4. Wait for page to load (15-20 seconds)
5. You should see KPI tiles with values
6. Look for "Add KPI" or "Create Custom" button
7. Look for KPI library/gallery
8. Take screenshot of KPI Scorecard

ALTERNATIVE CHECK 1:
Check Dashboard for KPI widgets

ALTERNATIVE CHECK 2:
Run SQL: SELECT COUNT(*) FROM invoices (confirms data for KPIs exists)

ALTERNATIVE CHECK 3:
Look for KPIs in Dashboards section

WHAT TO LOOK FOR:
- KPI tiles showing values (Revenue, Margin, etc.)
- "Add KPI" button = can customize
- KPI library = multiple options available

SCREENSHOT CHECKLIST:
[ ] KPI Scorecard page visible
[ ] Multiple KPIs with values
[ ] Add/Create button visible

PASS CRITERIA:
- KPI Scorecard loads = PASS
- KPIs with values = PASS
- Custom option available = PASS""",
    },
    {
        "ref": "WR-010",
        "name": "Dashboards",
        "category": "Reporting",
        "client_note": "Dashboards provide visual analytics with pre-built templates.",
        "step_by_step": """DASHBOARDS - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Visual dashboards with charts and graphs for business analysis.

NAVIGATION PATH:
Left Menu > Reports > Dashboards
OR Direct URL: qbo.intuit.com/app/reportbuilder

PRIMARY VALIDATION:
1. Click "Reports" in left menu
2. Look for "Dashboards" option
3. Click on it
4. You should see dashboard templates/gallery
5. Look for options like:
   - Financial Dashboard
   - Sales Dashboard
   - Cash Flow Dashboard
6. Click on any dashboard to open
7. Wait for charts to load
8. Take screenshot of dashboard gallery OR open dashboard

ALTERNATIVE CHECK 1:
Try direct URL: qbo.intuit.com/app/reportbuilder

ALTERNATIVE CHECK 2:
Check for 'Custom Reports' section in Reports

ALTERNATIVE CHECK 3:
Look for dashboard sharing/export options

WHAT TO LOOK FOR:
- Dashboard gallery with templates
- Charts with data loading
- Interactive filters
- Drill-down options

SCREENSHOT CHECKLIST:
[ ] Dashboard gallery visible OR
[ ] Open dashboard with charts
[ ] Data populating charts

PASS CRITERIA:
- Dashboard gallery exists = PASS
- Charts showing data = PASS""",
    },
    {
        "ref": "WR-011",
        "name": "3P Data Integrations",
        "category": "Reporting",
        "client_note": "Third-party integrations allow connecting CRM and business apps.",
        "step_by_step": """3P DATA INTEGRATIONS - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Connect third-party apps (Salesforce, HubSpot) to bring external data into reports.

NAVIGATION PATH:
Left Menu > Apps

PRIMARY VALIDATION:
1. Click "Apps" in left menu
2. Wait for Apps page to load
3. Look for "My Apps" section (connected apps)
4. Look for "Find apps" or marketplace section
5. Search for integrations like:
   - Salesforce
   - HubSpot
   - Monday.com
   - Shopify
6. Check for "data" or "reporting" categories
7. Take screenshot of Apps page

ALTERNATIVE CHECK 1:
Check Settings > Connected apps

ALTERNATIVE CHECK 2:
Look for 'Data sync' option in Settings

ALTERNATIVE CHECK 3:
Search Apps marketplace for 'CRM'

WHAT TO LOOK FOR:
- Connected apps in "My Apps"
- Integration marketplace
- Data sync options

SCREENSHOT CHECKLIST:
[ ] Apps page visible
[ ] Integration options visible
[ ] My Apps section visible

PASS CRITERIA:
- Apps page loads = PASS
- Integration options available = PASS""",
    },
    {
        "ref": "WR-012",
        "name": "Calculated Fields",
        "category": "Reporting",
        "client_note": "Calculated Fields allows custom formulas in any report.",
        "step_by_step": """CALCULATED FIELDS - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Add custom calculations (formulas) to reports.

NAVIGATION PATH:
Reports > [Any Report] > Customize

PRIMARY VALIDATION:
1. Click "Reports" in left menu
2. Open any report (Profit and Loss recommended)
3. Click "Customize" button (top right)
4. Customization panel opens
5. Look for:
   - "Add calculated field"
   - "Custom column"
   - "Formula" option
6. May be under "Columns" or "Fields" section
7. Take screenshot of customization panel

ALTERNATIVE CHECK 1:
Try multiple report types for the option

ALTERNATIVE CHECK 2:
Check 'Columns' section in customize panel

ALTERNATIVE CHECK 3:
Look for 'Add column' button in report header

WHAT TO LOOK FOR:
- Calculated field option in customize panel
- Formula builder/editor
- Custom column option

SCREENSHOT CHECKLIST:
[ ] Report open
[ ] Customize panel visible
[ ] Calculated field option visible

PASS CRITERIA:
- Calculated field option exists = PASS
- Can add formula = PASS""",
    },
    {
        "ref": "WR-013",
        "name": "Management Reports",
        "category": "Reporting",
        "client_note": "Management Reports combines financial statements for board presentations.",
        "step_by_step": """MANAGEMENT REPORTS - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Combine multiple financial statements into professional reports for management.

NAVIGATION PATH:
Left Menu > Reports > Management Reports
OR Direct URL: qbo.intuit.com/app/managementreports

PRIMARY VALIDATION:
1. Click "Reports" in left menu
2. Look for "Management Reports" in the list
3. Click on it
4. Wait for page to load
5. Look for options to:
   - Combine reports
   - Add cover page
   - Add text/notes
   - Customize branding
6. Check for templates
7. Take screenshot

ALTERNATIVE CHECK 1:
Try direct URL: qbo.intuit.com/app/managementreports

ALTERNATIVE CHECK 2:
Check for 'Package' or 'Report Package' option in Reports

ALTERNATIVE CHECK 3:
Look in Reports favorites section

WHAT TO LOOK FOR:
- Management Reports page
- Templates available
- Combined report options
- Cover page customization

SCREENSHOT CHECKLIST:
[ ] Management Reports page visible
[ ] Creation options visible
[ ] Templates available

PASS CRITERIA:
- Page loads = PASS
- Can create combined reports = PASS""",
    },
    {
        "ref": "WR-014",
        "name": "Benchmarking",
        "category": "Reporting",
        "client_note": "Benchmarking compares metrics against industry standards.",
        "step_by_step": """BENCHMARKING - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Compare your business performance against industry benchmarks.

NAVIGATION PATH:
Dashboard OR Reports > Benchmarking

PRIMARY VALIDATION:
1. Check Dashboard for industry comparison widgets
2. Go to Reports
3. Look for "Benchmarking" or "Industry Comparison"
4. Check KPI Scorecard for benchmark data
5. Look for "vs industry" or "% compared to"
6. Check Performance Center if available

ALTERNATIVE CHECK 1:
Check Dashboard for 'How am I doing' widget

ALTERNATIVE CHECK 2:
Look in KPI Scorecard for industry % comparisons

ALTERNATIVE CHECK 3:
Search Reports for 'benchmark'

LOCATIONS TO CHECK:
- Dashboard widgets
- KPI Scorecard
- Reports section
- Performance Center

WHAT TO LOOK FOR:
- Industry comparison percentages
- "vs industry" labels
- Benchmark indicators (up/down arrows)

SCREENSHOT CHECKLIST:
[ ] Location where benchmarking found
[ ] Industry comparison visible
[ ] OR note where you searched

PASS CRITERIA:
- Industry comparison visible = PASS
- Benchmark data showing = PASS
- Not found = Document search locations""",
    },
    {
        "ref": "WR-015",
        "name": "Multi-Entity Reports",
        "category": "Reporting",
        "client_note": "Multi-Entity Reports provide consolidated financial reporting.",
        "step_by_step": """MULTI-ENTITY REPORTS - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
View consolidated reports across multiple companies in a group.

IMPORTANT: Must switch to CONSOLIDATED VIEW first!

NAVIGATION PATH:
Company Selector (top left) > Consolidated View > Reports

PRIMARY VALIDATION:
1. Look at TOP LEFT corner (company name)
2. Click on company name
3. Dropdown shows all companies
4. Look for "Consolidated" or "Parent" view
5. Click to switch to Consolidated
6. WAIT for view to change (10-15 seconds)
7. Verify header shows "Consolidated"
8. Go to Reports > Profit and Loss
9. Report should show combined data from ALL companies
10. Take screenshot showing header + report

ALTERNATIVE CHECK 1:
Check company selector for 'All companies' option

ALTERNATIVE CHECK 2:
Look for entity column in reports

ALTERNATIVE CHECK 3:
Run SQL: SELECT company_type, COUNT(*) FROM invoices GROUP BY company_type

WHAT TO LOOK FOR:
- Header confirms Consolidated view
- Report shows combined totals
- "All entities" or multi-company data

SCREENSHOT CHECKLIST:
[ ] Header shows Consolidated view
[ ] Report open with data
[ ] Combined entity data visible

PASS CRITERIA:
- Can switch to Consolidated = PASS
- Report shows multi-entity data = PASS""",
    },
    {
        "ref": "WR-016",
        "name": "Dimension Assignment v2",
        "category": "Dimensions",
        "client_note": "Dimension Assignment v2 provides AI-assisted categorization.",
        "step_by_step": """DIMENSION ASSIGNMENT v2 - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Assign dimension values to transactions with AI suggestions (sparkle icons).

NAVIGATION PATH:
Direct URL: qbo.intuit.com/app/dimensions/assignment
OR Settings (gear) > Dimensions

PRIMARY VALIDATION:
1. Try direct URL above
2. OR click Gear icon > look for "Dimensions"
3. Look for "Assignment" tab/section
4. Wait for page to load
5. You should see a table of items needing dimensions
6. Look for SPARKLE ICONS = AI suggestions
7. Look for "Suggested" labels
8. Take screenshot showing assignment table

ALTERNATIVE CHECK 1:
Check Settings > Account and Settings > Advanced > Dimensions

ALTERNATIVE CHECK 2:
Look for dimension fields when creating invoices

ALTERNATIVE CHECK 3:
Run SQL: SELECT COUNT(*) FROM classifications

WHAT TO LOOK FOR:
- Assignment table with products/services
- Sparkle icons on suggestions
- Dimension dropdown options

SCREENSHOT CHECKLIST:
[ ] Dimension Assignment page visible
[ ] Items listed in table
[ ] Sparkle icons visible (if present)

PASS CRITERIA:
- Assignment page loads = PASS
- AI sparkles visible = BONUS""",
    },
    {
        "ref": "WR-017",
        "name": "Hierarchical Dimension Reporting",
        "category": "Dimensions",
        "client_note": "Hierarchical Dimensions allow parent-child organization for roll-up reporting.",
        "step_by_step": """HIERARCHICAL DIMENSION REPORTING - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Report on dimensions with parent-child hierarchy (Region > City > Store).

OPTION A - DATABASE EVIDENCE (Recommended):
Run SQL on QBO database:
SELECT id, name, parent_id FROM classifications WHERE parent_id IS NOT NULL LIMIT 20;

If rows returned = hierarchy exists. Screenshot the query result.


OPTION B - UI VALIDATION:

NAVIGATION PATH:
Reports > [Any Report] > Customize > Filter by Dimension

PRIMARY VALIDATION:
1. Go to Reports
2. Open Profit and Loss or similar
3. Click "Customize"
4. Look for dimension filter options
5. When selecting dimension values, look for:
   - Indented items (children under parents)
   - Expandable sections
   - Roll-up totals
6. Take screenshot showing hierarchy

ALTERNATIVE CHECK 1:
Run SQL: SELECT COUNT(*) FROM classifications WHERE parent_id IS NOT NULL

ALTERNATIVE CHECK 2:
Check dimension setup in Settings for parent/child config

ALTERNATIVE CHECK 3:
Look for 'Sub-class' options when creating dimensions

WHAT TO LOOK FOR:
- Parent > Child structure in selection
- Indented dimension values
- Roll-up/drill-down capability

SCREENSHOT CHECKLIST:
[ ] Database query results OR
[ ] Report with hierarchical dimension filter
[ ] Parent-child relationship visible

PASS CRITERIA:
- Database shows parent_id relationships = PASS
- UI shows hierarchical selection = PASS""",
    },
    {
        "ref": "WR-018",
        "name": "Dimensions on Workflow",
        "category": "Dimensions",
        "client_note": "Workflow automation can use dimensions as rule conditions.",
        "step_by_step": """DIMENSIONS ON WORKFLOW - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Use dimensions as conditions in workflow automation rules.

NOTE: May return 404 in Construction environments - that's expected.

NAVIGATION PATH:
Settings (gear) > Workflow automation
OR Direct URL: qbo.intuit.com/app/workflowautomation

PRIMARY VALIDATION:
1. Click Gear icon (top right)
2. Look for "Workflow automation" or "Workflows"
3. Click on it
4. IF PAGE LOADS:
   - Look at existing rules
   - Click "Create rule" or edit a rule
   - In rule conditions, look for "Dimension" option
   - Check if you can filter by dimension values
5. IF 404 ERROR:
   - Screenshot the 404 page
   - Mark as "NOT AVAILABLE" (expected in some environments)
6. Take screenshot

ALTERNATIVE CHECK 1:
Try Settings > Manage workflows

ALTERNATIVE CHECK 2:
Check for 'Automation' in left menu

ALTERNATIVE CHECK 3:
Look for rule conditions mentioning Class/Location

WHAT TO LOOK FOR:
- Workflow automation page
- Dimension as rule condition option
- OR 404 error (document it)

SCREENSHOT CHECKLIST:
[ ] Workflow page visible OR 404 error
[ ] Dimension option in rules (if available)

PASS CRITERIA:
- Dimension in rule conditions = PASS
- 404 error = Mark NOT AVAILABLE""",
    },
    {
        "ref": "WR-019",
        "name": "Dimensions on Balance Sheet",
        "category": "Dimensions",
        "client_note": "Balance Sheet can be filtered by dimensions for segment analysis.",
        "step_by_step": """DIMENSIONS ON BALANCE SHEET - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Filter Balance Sheet reports by dimension values (Class, Location, etc).

NAVIGATION PATH:
Reports > Balance Sheet > Customize > Filter

PRIMARY VALIDATION:
1. Click "Reports" in left menu
2. Find "Balance Sheet" report
3. Click to open it
4. Wait for report to generate
5. Click "Customize" button (top right)
6. Look for "Filter" section
7. In filters, look for:
   - "Class" (type of dimension)
   - "Location"
   - "Department"
   - Any custom dimensions
8. Click on dimension dropdown to see values
9. Select a value and run report
10. Take screenshot of customization panel

ALTERNATIVE CHECK 1:
Check P&L report for same dimension filters

ALTERNATIVE CHECK 2:
Look for 'Segment' or 'Location' columns in report

ALTERNATIVE CHECK 3:
Try 'Balance Sheet by Class' report if available

WHAT TO LOOK FOR:
- Dimension filter options in customize
- Dropdown with dimension values
- Report updates when filtered

SCREENSHOT CHECKLIST:
[ ] Balance Sheet open
[ ] Customize panel visible
[ ] Dimension filter options shown

PASS CRITERIA:
- Dimension filters available = PASS
- Can filter report = PASS""",
    },
    {
        "ref": "WR-020",
        "name": "Parallel Approval",
        "category": "Workflow",
        "client_note": "Parallel Approval allows simultaneous review by multiple approvers.",
        "step_by_step": """PARALLEL APPROVAL - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Allow multiple approvers to review simultaneously (not sequentially).

NOTE: May return 404 in Construction environments.

NAVIGATION PATH:
Settings (gear) > Workflow automation
OR Direct URL: qbo.intuit.com/app/workflowautomation

PRIMARY VALIDATION:
1. Click Gear icon (top right)
2. Click "Workflow automation"
3. IF PAGE LOADS:
   - Look for approval workflow rules
   - Create or edit an approval rule
   - Look for MULTIPLE approver option
   - Check for "Parallel" or "Any approver" choice:
     * "All must approve" = sequential
     * "Any can approve" = parallel
4. IF 404 ERROR:
   - Screenshot the 404
   - Mark as NOT AVAILABLE
5. Take screenshot

ALTERNATIVE CHECK 1:
Look for 'Approval rules' in Settings

ALTERNATIVE CHECK 2:
Check for 'Approval required' option on bills/expenses

ALTERNATIVE CHECK 3:
Look for user role 'Approver' in user management

WHAT TO LOOK FOR:
- Multiple approver fields
- "Any can approve" option
- Parallel approval configuration

SCREENSHOT CHECKLIST:
[ ] Workflow page visible OR 404
[ ] Multiple approvers option (if available)
[ ] Parallel/Any option visible

PASS CRITERIA:
- Parallel approval option = PASS
- 404 = NOT AVAILABLE""",
    },
    {
        "ref": "WR-021",
        "name": "Desktop Migration",
        "category": "Migration",
        "client_note": "Desktop Migration allows importing from QuickBooks Desktop.",
        "step_by_step": """DESKTOP MIGRATION - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Migrate data from QuickBooks Desktop to QuickBooks Online.

NOTE: Full demo requires FRESH tenant. Can only verify option exists.

NAVIGATION PATH:
Settings (gear) > Import data

PRIMARY VALIDATION:
1. Click Gear icon (top right)
2. Look for "Import data" or "Import"
3. Click on it
4. Look for Desktop migration option:
   - "Import from QuickBooks Desktop"
   - "Desktop migration"
   - "Convert from Desktop"
5. May also be under "Tools" menu
6. Note what file types supported (.qbw, .qbb)
7. Take screenshot of import options

ALTERNATIVE CHECK 1:
Check Settings > Tools > Import

ALTERNATIVE CHECK 2:
Look for 'Migration' in Help menu

ALTERNATIVE CHECK 3:
Search Settings for 'Desktop'

WHAT TO LOOK FOR:
- Desktop import option
- Supported file types
- Migration workflow

SCREENSHOT CHECKLIST:
[ ] Import page visible
[ ] Desktop migration option visible

PASS CRITERIA:
- Desktop migration option exists = PASS

NOTE: Full migration requires fresh tenant.""",
    },
    {
        "ref": "WR-022",
        "name": "DFY Migration",
        "category": "Migration",
        "client_note": "Done-For-You Migration provides professional migration services.",
        "step_by_step": """DFY (DONE-FOR-YOU) MIGRATION - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Professional migration service - Intuit handles the data migration for you.

NOTE: Full demo requires FRESH tenant.

NAVIGATION PATH:
Settings (gear) > Import data

PRIMARY VALIDATION:
1. Click Gear icon (top right)
2. Look for "Import data"
3. Click on it
4. Look for DFY option:
   - "Done for you" migration
   - "Professional migration"
   - "We'll do it for you"
5. May be presented as premium option
6. Note service details if visible
7. Take screenshot

ALTERNATIVE CHECK 1:
Check for 'Professional services' link on import page

ALTERNATIVE CHECK 2:
Look in Help for migration services

ALTERNATIVE CHECK 3:
Contact Intuit for DFY availability

WHAT TO LOOK FOR:
- DFY/Professional migration option
- Service description
- Pricing information (if shown)

SCREENSHOT CHECKLIST:
[ ] Import options visible
[ ] DFY option visible (if exists)

PASS CRITERIA:
- DFY option visible = PASS
- Only self-service = Note subscription level""",
    },
    {
        "ref": "WR-023",
        "name": "Feature Compatibility",
        "category": "Migration",
        "client_note": "Feature Compatibility documentation available through Intuit resources.",
        "step_by_step": """FEATURE COMPATIBILITY - VALIDATION MANUAL

WHAT THIS IS:
Documentation showing which Desktop features are compatible with QBO.

THIS IS DOCUMENTATION - NO UI VALIDATION NEEDED

VALIDATION METHOD:
- Reference Intuit's official compatibility documentation
- No screenshot required
- Mark as N/A for UI validation

EVIDENCE:
Note: "Feature Compatibility is documentation only - refer to Intuit migration guide"

EXPECTED RESULT:
N/A - Documentation feature""",
    },
    {
        "ref": "WR-024",
        "name": "Certified Payroll Report",
        "category": "Construction",
        "client_note": "Certified Payroll generates WH-347 reports for government projects.",
        "step_by_step": """CERTIFIED PAYROLL REPORT - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Generate WH-347 certified payroll reports for government construction projects.

NOTE: CONSTRUCTION-SPECIFIC feature.

NAVIGATION PATH:
Reports > Payroll > Certified Payroll
OR Payroll > Reports > Certified Payroll

PRIMARY VALIDATION:
1. Go to Reports
2. Look in "Payroll" category
3. OR search for "Certified Payroll"
4. OR check Payroll > Reports
5. Find "Certified Payroll Report" or "WH-347"
6. Click to open
7. Select date range and project (if required)
8. Generate the report
9. Verify WH-347 format
10. Take screenshot

ALTERNATIVE CHECK 1:
Check Payroll > Tax forms for WH-347

ALTERNATIVE CHECK 2:
Look for 'Prevailing wage' reports

ALTERNATIVE CHECK 3:
Run SQL: SELECT company_type, COUNT(*) FROM employees GROUP BY company_type

WHAT TO LOOK FOR:
- WH-347 format
- Employee wages, hours, deductions
- Government compliance fields
- Project association

SCREENSHOT CHECKLIST:
[ ] Certified Payroll Report visible
[ ] WH-347 header/format visible
[ ] Employee data showing

PASS CRITERIA:
- WH-347 report available = PASS
- Government format correct = PASS""",
    },
    {
        "ref": "WR-025",
        "name": "Sales Order",
        "category": "Construction",
        "client_note": "Sales Orders support order-to-cash workflow before invoicing.",
        "step_by_step": """SALES ORDER - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Create sales orders before invoicing (order-to-invoice workflow).

NAVIGATION PATH:
+ (Create) button > Sales Order
OR Sales menu > Sales Orders

PRIMARY VALIDATION:
1. Click "+" (Create) button
2. Look for "Sales Order" option
3. OR check Sales menu > Sales Orders
4. OR check Settings > Sales settings
5. May need to enable in Settings:
   - Gear > Account and Settings > Sales
   - Look for "Sales Order" toggle
6. If found, click to view Sales Order form
7. Note available fields
8. Take screenshot

ALTERNATIVE CHECK 1:
Check Settings > Sales for order settings

ALTERNATIVE CHECK 2:
Look for 'Order' in + Create menu

ALTERNATIVE CHECK 3:
Check Sales transactions list for orders

WHAT TO LOOK FOR:
- Sales Order in Create menu
- Sales Order creation form
- Order-to-invoice workflow

SCREENSHOT CHECKLIST:
[ ] Sales Order option visible in menu
[ ] OR Sales Order form visible
[ ] OR Settings showing Sales Order option

PASS CRITERIA:
- Sales Order available = PASS
- May need enabling in Settings""",
    },
    {
        "ref": "WR-026",
        "name": "Multi-Entity Payroll Hub",
        "category": "Payroll",
        "client_note": "Multi-Entity Payroll Hub provides centralized employee management.",
        "step_by_step": """MULTI-ENTITY PAYROLL HUB - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Centralized employee management across multiple companies.

IMPORTANT: Must be in CONSOLIDATED VIEW!

NAVIGATION PATH:
Company Selector > Consolidated View > Payroll > Employees

PRIMARY VALIDATION:
1. Look at TOP LEFT (company name)
2. Click on company name
3. Select "Consolidated" or "Parent" view
4. WAIT for view to change
5. Verify header shows Consolidated
6. Click "Payroll" in left menu
7. Click "Employees" or "Employee Hub"
8. You should see employees from ALL companies
9. Look for entity/company column
10. Take screenshot

ALTERNATIVE CHECK 1:
Check for 'All entities' filter on employees page

ALTERNATIVE CHECK 2:
Look for entity dropdown in Payroll section

ALTERNATIVE CHECK 3:
Run SQL: SELECT company_type, COUNT(*) FROM employees GROUP BY company_type

WHAT TO LOOK FOR:
- Header confirms Consolidated view
- Employee list from multiple companies
- Entity column showing company names

SCREENSHOT CHECKLIST:
[ ] Header shows Consolidated
[ ] Payroll > Employees visible
[ ] Multiple entity employees visible

PASS CRITERIA:
- Multi-entity employee list = PASS
- Entity column visible = PASS""",
    },
    {
        "ref": "WR-027",
        "name": "Garnishments",
        "category": "Payroll",
        "client_note": "Garnishments feature allows configuring wage garnishments.",
        "step_by_step": """GARNISHMENTS - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Configure automatic wage garnishments (child support, tax levies) for employees.

NAVIGATION PATH:
Payroll > Employees > [Employee] > Deductions/Garnishments

PRIMARY VALIDATION:
1. Click "Payroll" in left menu
2. Click "Employees"
3. Wait for list to load
4. Click on any employee name
5. Wait for profile to load
6. Scroll through profile sections
7. Look for:
   - "Deductions" section
   - "Garnishments" section
   - "Withholdings" section
8. May be under "Pay" or "Taxes" tab
9. Check for garnishment types:
   - Child support
   - Tax levies
   - Creditor garnishments
10. Take screenshot

ALTERNATIVE CHECK 1:
Check employee 'Additional pay/deductions' section

ALTERNATIVE CHECK 2:
Look for 'Court-ordered' deductions

ALTERNATIVE CHECK 3:
Check Payroll Settings for garnishment setup

WHAT TO LOOK FOR:
- Garnishment section in profile
- Garnishment type options
- Configuration fields

SCREENSHOT CHECKLIST:
[ ] Employee profile open
[ ] Garnishments section visible
[ ] Options available

PASS CRITERIA:
- Garnishment section exists = PASS
- Can configure types = PASS""",
    },
    {
        "ref": "WR-028",
        "name": "Assignments in QBTime",
        "category": "Payroll",
        "client_note": "QBTime Assignments tracks time by project for labor costing.",
        "step_by_step": """ASSIGNMENTS IN QBTIME - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
Track time entries with project/job assignments for labor costing.

OPTION A - DATABASE EVIDENCE (Quick):
Run SQL: SELECT COUNT(*) as total_time_entries FROM time_entries;
Large number (e.g., 62,997) = time tracking active.


OPTION B - UI VALIDATION:

NAVIGATION PATH:
Left Menu > Time OR Payroll > Time

PRIMARY VALIDATION:
1. Click "Time" in left menu
2. OR Payroll > Time entries
3. Wait for time entries page
4. Look at the time entry list
5. Check for columns:
   - Employee
   - Customer/Project
   - Service item
   - Hours
6. Look for "Assignment" or "Project" column
7. Check for job costing info
8. Take screenshot

ALTERNATIVE CHECK 1:
Run SQL: SELECT COUNT(*) FROM time_entries

ALTERNATIVE CHECK 2:
Check for 'Timesheet' in left menu

ALTERNATIVE CHECK 3:
Look for time by project reports in Reports

WHAT TO LOOK FOR:
- Time entries with project assignments
- Job/customer association
- Hours by project

SCREENSHOT CHECKLIST:
[ ] Database query result OR
[ ] Time entries page visible
[ ] Assignments/projects visible

PASS CRITERIA:
- Database shows time entries = PASS
- UI shows assignments = PASS""",
    },
    {
        "ref": "WR-029",
        "name": "Enhanced Amendments (CA)",
        "category": "Payroll",
        "client_note": "Enhanced Amendments (CA) is California-specific. Requires CA tenant.",
        "step_by_step": """ENHANCED AMENDMENTS (CA) - VALIDATION MANUAL

WHAT THIS FEATURE DOES:
California-specific payroll amendment handling.

IMPORTANT: Requires CALIFORNIA tenant!

VALIDATION STATUS: NOT APPLICABLE
Current environments are not California-specific.

IF YOU HAVE CA TENANT:
1. Login to California tenant
2. Go to Payroll > Taxes
3. Look for "Amendments" or "Corrections"
4. Check for CA-specific workflow
5. Screenshot amendment options

FOR CURRENT ENVIRONMENTS:
Mark as N/A - Requires California tenant

EXPECTED RESULT:
N/A - CA tenant required""",
    },
]

# Complete login information
LOGINS = [
    {
        "environment": "TCO",
        "email": "quickbooks-tco-tbxdemo@tbxofficial.com",
        "password": "TestBox!23",
        "totp_secret": "RP4MFT45ZY5EYGMRIPEANJA6YFKHV5O7",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341455130122367",
                "name": "Traction Control Outfitters (Parent)",
                "type": "Parent",
                "priority": "P0",
            },
            {
                "cid": "9341455130196737",
                "name": "Global Tread Distributors",
                "type": "Child",
                "priority": "P0",
            },
            {
                "cid": "9341455130166501",
                "name": "Apex Tire & Auto Retail",
                "type": "Child",
                "priority": "P0",
            },
            {
                "cid": "9341455130182073",
                "name": "RoadReady Service Solutions",
                "type": "Child",
                "priority": "P0",
            },
        ],
        "notes": "TCO multi-entity demo. Parent for consolidated, Children for individual.",
    },
    {
        "environment": "CONSTRUCTION SALES",
        "email": "quickbooks-test-account@tbxofficial.com",
        "password": "TestBox123!",
        "totp_secret": "23CIWY3QYCOXJRKZYG6YKR7HUYVPLPEL",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341454156620895",
                "name": "Keystone Construction (Par.) - PRIMARY",
                "type": "Parent",
                "priority": "P0",
            },
            {
                "cid": "9341454156620204",
                "name": "Keystone Terra (Ch.) - PRIMARY",
                "type": "Child",
                "priority": "P0",
            },
            {
                "cid": "9341454156643813",
                "name": "Keystone Ironcraft - PRIMARY",
                "type": "Child",
                "priority": "P1",
            },
            {
                "cid": "9341454156621045",
                "name": "Keystone BlueCraft (Ch.)",
                "type": "Child",
                "priority": "P0",
            },
            {
                "cid": "9341454156640324",
                "name": "KeyStone Stonecraft",
                "type": "Child",
                "priority": "P1",
            },
            {
                "cid": "9341454156634620",
                "name": "KeyStone Canopy",
                "type": "Child",
                "priority": "P1",
            },
            {
                "cid": "9341454156644492",
                "name": "KeyStone Ecocraft",
                "type": "Child",
                "priority": "P1",
            },
            {
                "cid": "9341454156629550",
                "name": "KeyStone Volt",
                "type": "Child",
                "priority": "P1",
            },
        ],
        "notes": "Construction Sales. PRIMARY marked for main demos.",
    },
    {
        "environment": "CONSTRUCTION EVENTS",
        "email": "quickbooks-test-account-qsp@tbxofficial.com",
        "password": "TestBox123!",
        "totp_secret": "23CIWY3QYCOXJRKZYG6YKR7HUYVPLPEL",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341454796804202",
                "name": "Keystone Construction (Event) - PRIMARY",
                "type": "Parent",
                "priority": "P0",
            },
            {
                "cid": "9341454842738078",
                "name": "Keystone Terra (Event) - PRIMARY",
                "type": "Child",
                "priority": "P0",
            },
            {
                "cid": "9341454842756096",
                "name": "Keystone BlueCraft (Event)",
                "type": "Child",
                "priority": "P0",
            },
        ],
        "notes": "Construction Events demo environment.",
    },
    {
        "environment": "CONSTRUCTION BI",
        "email": "quickbooks-tbx-product-team-test@tbxofficial.com",
        "password": "TBD",
        "totp_secret": "TBD",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341455531293719",
                "name": "Keystone Construction (BI)",
                "type": "Parent",
                "priority": "",
            },
            {
                "cid": "9341455531294375",
                "name": "Keystone Terra (BI)",
                "type": "Child",
                "priority": "",
            },
            {
                "cid": "9341455531274694",
                "name": "Keystone BlueCraft (BI)",
                "type": "Child",
                "priority": "",
            },
        ],
        "notes": "Product team BI testing. Confirm credentials with team.",
    },
]


def create_workbook():
    wb = Workbook()

    # Sheet 1: TCO
    create_environment_sheet(wb, "TCO Validation", "TCO")

    # Sheet 2: Construction Sales
    ws_cs = wb.create_sheet("Construction Sales")
    create_environment_sheet_on_ws(ws_cs, "CONSTRUCTION SALES")

    # Sheet 3: Construction Events
    ws_ce = wb.create_sheet("Construction Events")
    create_environment_sheet_on_ws(ws_ce, "CONSTRUCTION EVENTS")

    # Sheet 4: Step by Step (the rich manual)
    create_step_by_step_sheet(wb)

    # Sheet 5: Logins
    create_logins_sheet(wb)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    return wb


def create_environment_sheet(wb, sheet_name, env_name):
    ws = wb.active
    ws.title = sheet_name
    create_environment_sheet_on_ws(ws, env_name)


def create_environment_sheet_on_ws(ws, env_name):
    # Headers: Ref, Feature, Category, Client Notes, Evidence Link, Status
    headers = [
        "Ref",
        "Feature Name",
        "Category",
        "Client Notes",
        "Evidence Link",
        "Status",
    ]

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{env_name} - WINTER RELEASE FY26 VALIDATION"
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = CENTER

    # Headers row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ALL 29 features
    row = 4
    for feature in FEATURES:
        cell_ref = ws.cell(row=row, column=1, value=feature["ref"])
        cell_ref.font = CALIBRI_11
        cell_ref.alignment = CENTER
        cell_ref.border = THIN_BORDER

        cell_name = ws.cell(row=row, column=2, value=feature["name"])
        cell_name.font = CALIBRI_11
        cell_name.alignment = CENTER
        cell_name.border = THIN_BORDER

        cell_cat = ws.cell(row=row, column=3, value=feature["category"])
        cell_cat.font = CALIBRI_11
        cell_cat.alignment = CENTER
        cell_cat.border = THIN_BORDER

        cell_notes = ws.cell(row=row, column=4, value=feature["client_note"])
        cell_notes.font = CALIBRI_11
        cell_notes.alignment = CENTER
        cell_notes.border = THIN_BORDER

        cell_evidence = ws.cell(row=row, column=5, value="")
        cell_evidence.font = CALIBRI_11
        cell_evidence.alignment = CENTER
        cell_evidence.border = THIN_BORDER

        cell_status = ws.cell(row=row, column=6, value="")
        cell_status.font = CALIBRI_11
        cell_status.alignment = CENTER
        cell_status.border = THIN_BORDER

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 12


def create_step_by_step_sheet(wb):
    ws = wb.create_sheet("Step by Step Guide")

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "VALIDATION MANUAL - COMPLETE STEP-BY-STEP INSTRUCTIONS"
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = CENTER

    # Headers
    headers = ["Ref", "Feature", "Complete Validation Instructions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data
    row = 4
    for feature in FEATURES:
        cell_ref = ws.cell(row=row, column=1, value=feature["ref"])
        cell_ref.font = CALIBRI_11
        cell_ref.alignment = CENTER
        cell_ref.border = THIN_BORDER

        cell_name = ws.cell(row=row, column=2, value=feature["name"])
        cell_name.font = CALIBRI_11
        cell_name.alignment = CENTER
        cell_name.border = THIN_BORDER

        cell_steps = ws.cell(row=row, column=3, value=feature["step_by_step"])
        cell_steps.font = CALIBRI_11
        cell_steps.alignment = LEFT_TOP
        cell_steps.border = THIN_BORDER

        ws.row_dimensions[row].height = 380
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 120


def create_logins_sheet(wb):
    ws = wb.create_sheet("Logins and Access")

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "COMPLETE LOGIN CREDENTIALS AND COMPANY IDs"
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = CENTER

    row = 3
    for login in LOGINS:
        # Environment header
        env_cell = ws.cell(row=row, column=1, value=login["environment"])
        env_cell.font = Font(name="Calibri", size=12, bold=True)
        env_cell.fill = SECTION_FILL
        env_cell.alignment = CENTER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        # Credentials
        ws.cell(row=row, column=1, value="Email:").font = CALIBRI_11_BOLD
        ws.cell(row=row, column=2, value=login["email"]).font = CALIBRI_11
        ws.cell(row=row, column=2).alignment = CENTER
        row += 1

        ws.cell(row=row, column=1, value="Password:").font = CALIBRI_11_BOLD
        ws.cell(row=row, column=2, value=login["password"]).font = CALIBRI_11
        ws.cell(row=row, column=2).alignment = CENTER
        row += 1

        ws.cell(row=row, column=1, value="TOTP Secret:").font = CALIBRI_11_BOLD
        ws.cell(row=row, column=2, value=login["totp_secret"]).font = CALIBRI_11
        ws.cell(row=row, column=2).alignment = CENTER
        row += 1

        ws.cell(row=row, column=1, value="URL:").font = CALIBRI_11_BOLD
        ws.cell(row=row, column=2, value=login["url"]).font = CALIBRI_11
        ws.cell(row=row, column=2).alignment = CENTER
        row += 1

        row += 1

        # Companies header
        company_headers = ["CID", "Company Name", "Type", "Priority"]
        for col, header in enumerate(company_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = CALIBRI_11_BOLD
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )
            cell.alignment = CENTER
        row += 1

        for company in login["companies"]:
            ws.cell(row=row, column=1, value=company["cid"]).font = CALIBRI_11
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value=company["name"]).font = CALIBRI_11
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=3, value=company["type"]).font = CALIBRI_11
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4, value=company["priority"]).font = CALIBRI_11
            ws.cell(row=row, column=4).alignment = CENTER
            row += 1

        # Notes
        row += 1
        ws.cell(row=row, column=1, value="Notes:").font = CALIBRI_11_BOLD
        notes_cell = ws.cell(row=row, column=2, value=login["notes"])
        notes_cell.font = CALIBRI_11
        notes_cell.alignment = CENTER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 2

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


if __name__ == "__main__":
    wb = create_workbook()
    output_path = (
        "C:/Users/adm_r/intuit-boom/docs/WINTER_RELEASE_VALIDATION_MATRIX_v4.xlsx"
    )
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(
        "Sheets: TCO Validation, Construction Sales, Construction Events, Step by Step Guide, Logins and Access"
    )
    print("\nv4 Changes:")
    print("- ALL 29 features in ALL environment sheets")
    print("- Calibri 11, centered")
    print("- Columns: Ref, Feature, Category, Client Notes, Evidence Link, Status")
    print("- Step by Step: single rich manual with all details + alt checks")
    print("- Removed Validator and Date columns")
