"""
Módulo para gerenciar cache de hyperlinks do Google Drive.
Aprende com planilhas existentes e reutiliza os mapeamentos.
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# Configurar encoding UTF-8 para stdout
sys.stdout.reconfigure(encoding="utf-8")

# Cache file location
CACHE_FILE = Path(__file__).parent.parent / "data" / "hyperlink_cache.json"


def load_cache():
    """Carrega o cache de hyperlinks do arquivo JSON."""
    if CACHE_FILE.exists():
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache):
    """Salva o cache de hyperlinks no arquivo JSON."""
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def learn_from_spreadsheet(spreadsheet_path):
    """
    Aprende mapeamento filename -> hyperlink de uma planilha existente.

    Args:
        spreadsheet_path: Caminho para a planilha com hyperlinks corretos

    Returns:
        dict: Mapeamento {filename: hyperlink_url}
    """
    cache = load_cache()
    learned = 0

    wb = load_workbook(spreadsheet_path)
    ws = wb.active

    # Encontrar colunas de Evidence e Link
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    evidence_col = headers.get("Evidence_file")
    link_cols = [col for name, col in headers.items() if "Link" in name]

    if not evidence_col or not link_cols:
        print(f"Colunas não encontradas: Evidence_file={evidence_col}, Links={link_cols}")
        wb.close()
        return cache

    # Extrair mapeamentos
    for row_idx in range(2, ws.max_row + 1):
        evidence_cell = ws.cell(row=row_idx, column=evidence_col)
        filename = evidence_cell.value

        if not filename or filename.strip() == "" or filename.startswith("N/A"):
            continue

        # Pode ter múltiplos filenames separados por newline
        filenames = [f.strip() for f in str(filename).split("\n") if f.strip()]

        for i, fname in enumerate(filenames):
            if i < len(link_cols):
                link_cell = ws.cell(row=row_idx, column=link_cols[i])
                if link_cell.hyperlink and link_cell.hyperlink.target:
                    url = link_cell.hyperlink.target
                    if url.startswith("https://drive.google.com"):
                        if fname not in cache:
                            cache[fname] = url
                            learned += 1
                            print(f"  Aprendido: {fname[:50]}... -> {url[:60]}...")

    wb.close()

    # Salvar cache atualizado
    save_cache(cache)
    print(f"\nTotal aprendido: {learned} novos mapeamentos")
    print(f"Cache total: {len(cache)} entradas")

    return cache


def get_hyperlink(filename):
    """
    Obtém hyperlink para um filename do cache.

    Args:
        filename: Nome do arquivo de screenshot

    Returns:
        str: URL do Google Drive ou None se não encontrado
    """
    cache = load_cache()
    return cache.get(filename)


def get_all_hyperlinks():
    """Retorna todo o cache de hyperlinks."""
    return load_cache()


if __name__ == "__main__":
    # Teste: aprender da planilha oficial
    OFFICIAL_SPREADSHEET = Path(r"G:\Meu Drive\TestBox\QBO-Evidence\fall_release_control.xlsx")

    print(f"Aprendendo hyperlinks de: {OFFICIAL_SPREADSHEET}")
    print("=" * 80)

    cache = learn_from_spreadsheet(OFFICIAL_SPREADSHEET)

    print("\n" + "=" * 80)
    print("Cache final:")
    for fname, url in list(cache.items())[:5]:
        print(f"  {fname[:50]}... -> {url[:60]}...")
    if len(cache) > 5:
        print(f"  ... e mais {len(cache) - 5} entradas")
