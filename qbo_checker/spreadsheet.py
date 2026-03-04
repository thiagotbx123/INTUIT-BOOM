# qbo_checker/spreadsheet.py
"""Excel spreadsheet module for Fall Release control sheet

FONTE DE VERDADE para geracao de planilhas de controle.
Usa features_rich.json para templates de notas e hyperlink_cache para links do Drive.
"""

import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .config import CONTROL_SHEET, EVIDENCE_DIR


# Paths
FEATURES_RICH_FILE = Path(__file__).parent / "features_rich.json"
HYPERLINK_CACHE_FILE = Path(__file__).parent.parent / "data" / "hyperlink_cache.json"


# Column definitions
CONTROL_COLUMNS = [
    "Ref",
    "Project",
    "Company",
    "Feature",
    "Status",
    "Intuit_notes",
    "Internal_TBX_notes",
    "Evidence_file",
    "Link",
]

# Status colors
STATUS_COLORS = {
    "YES": "90EE90",  # Light green
    "PARTIAL": "FFD700",  # Gold
    "NO": "FF6B6B",  # Light red
    "NA": "D3D3D3",  # Light gray
    "Testing": "87CEEB",  # Light blue
}


def load_features_rich():
    """Load enriched features from features_rich.json"""
    if FEATURES_RICH_FILE.exists():
        with open(FEATURES_RICH_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return {f["ref"]: f for f in data.get("features", [])}
    return {}


def load_hyperlink_cache():
    """Load hyperlink cache for Google Drive links"""
    if HYPERLINK_CACHE_FILE.exists():
        with open(HYPERLINK_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def get_drive_link(filename: str) -> str:
    """Get Google Drive link from cache, or local path as fallback"""
    cache = load_hyperlink_cache()
    if filename in cache:
        return cache[filename]
    # Fallback to local path
    return str(EVIDENCE_DIR / filename)


def generate_notes_from_template(feature_data: dict, observation: str = None) -> tuple:
    """
    Generate Intuit and Internal notes from feature template.

    Args:
        feature_data: Feature dict from features_rich.json
        observation: Optional real-time observation to incorporate

    Returns:
        tuple: (intuit_notes, internal_notes)
    """
    templates = feature_data.get("notes_template", {})

    # Get template or use last validation notes
    intuit_notes = templates.get("intuit", "")
    internal_notes = templates.get("internal", "")

    # If observation provided, enhance internal notes
    if observation:
        if internal_notes:
            internal_notes = f"{internal_notes}\n\nObservacao atual: {observation}"
        else:
            internal_notes = f"Observacao: {observation}"

    return intuit_notes, internal_notes


def create_control_sheet() -> Path:
    """Create new control spreadsheet with headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature Checks"

    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(CONTROL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Set column widths
    widths = [10, 10, 15, 40, 10, 60, 50, 55, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(CONTROL_SHEET)
    return CONTROL_SHEET


def load_or_create_control_sheet() -> openpyxl.Workbook:
    """Load existing control sheet or create new one."""
    if CONTROL_SHEET.exists():
        return openpyxl.load_workbook(CONTROL_SHEET)
    else:
        create_control_sheet()
        return openpyxl.load_workbook(CONTROL_SHEET)


def generate_full_spreadsheet(project: str = "TCO", company: str = "Apex Tire") -> Path:
    """
    Generate complete spreadsheet from features_rich.json with:
    - All features defined
    - Notes from templates
    - Hyperlinks from cache

    Returns:
        Path to generated spreadsheet
    """
    features = load_features_rich()
    cache = load_hyperlink_cache()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{project} Feature Control"

    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    link_font = Font(color="0563C1", underline="single")

    # Write headers
    for col, header in enumerate(CONTROL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write feature rows
    row_idx = 2
    stats = {"total": 0, "with_link": 0, "missing_link": 0}

    for ref, feature in sorted(features.items(), key=lambda x: x[0]):
        # Note: evidence_type pode ser usado futuramente para filtrar WFS backend features
        _ = feature.get("evidence", {}).get("type", "ui")  # noqa: F841

        # Get notes from template
        intuit_notes, internal_notes = generate_notes_from_template(feature)

        # Get screenshot filename
        filename = feature.get("evidence", {}).get("filename_pattern", "")

        # Get status from last validation
        status = feature.get("last_validation", {}).get("status", "Testing")

        # Data row
        data = [
            ref,  # Ref
            project,  # Project
            company,  # Company
            feature.get("name", ""),  # Feature
            status,  # Status
            intuit_notes,  # Intuit_notes
            internal_notes,  # Internal_TBX_notes
            filename,  # Evidence_file
            None,  # Link (will add separately)
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

            # Style status cell with color
            if col == 5 and status in STATUS_COLORS:
                cell.fill = PatternFill(
                    start_color=STATUS_COLORS[status],
                    end_color=STATUS_COLORS[status],
                    fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="center")

            # Wrap text for notes columns
            if col in [6, 7]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Add hyperlink for evidence file
        link_cell = ws.cell(row=row_idx, column=9)
        link_cell.border = thin_border

        if filename and filename in cache:
            link_cell.value = "View"
            link_cell.hyperlink = cache[filename]
            link_cell.font = link_font
            stats["with_link"] += 1
        elif filename and not filename.startswith("N/A"):
            link_cell.value = "N/A"
            stats["missing_link"] += 1
        else:
            link_cell.value = "-"

        stats["total"] += 1
        row_idx += 1

    # Set column widths
    widths = [10, 10, 15, 40, 10, 60, 50, 55, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    # Save with date
    today = datetime.now().strftime("%Y-%m-%d")
    output_path = EVIDENCE_DIR / f"fall_release_control_{project}_{today}.xlsx"
    wb.save(output_path)
    wb.close()

    print(f"\nPlanilha gerada: {output_path}")
    print(f"Total features: {stats['total']}")
    print(f"Com link: {stats['with_link']}")
    print(f"Sem link: {stats['missing_link']}")

    return output_path


def add_feature_result(
    project: str,
    company: str,
    ref: str,
    feature: str,
    status: str,
    intuit_notes: str,
    internal_notes: str,
    evidence_file: str,
) -> int:
    """Add a feature check result to the control sheet."""
    wb = load_or_create_control_sheet()
    ws = wb.active

    # Find next empty row
    next_row = ws.max_row + 1

    # Get Google Drive link from cache
    drive_link = get_drive_link(evidence_file)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write data
    data = [
        ref,  # Col 1: Ref
        project,  # Col 2: Project
        company,  # Col 3: Company
        feature,  # Col 4: Feature
        status,  # Col 5: Status
        intuit_notes,  # Col 6: Intuit_notes
        internal_notes,  # Col 7: Internal_TBX_notes
        evidence_file,  # Col 8: Evidence_file
        None,  # Col 9: Link (will add hyperlink)
    ]

    for col, value in enumerate(data, 1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.border = thin_border

        # Style status cell with color
        if col == 5 and status in STATUS_COLORS:
            cell.fill = PatternFill(
                start_color=STATUS_COLORS[status],
                end_color=STATUS_COLORS[status],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

        # Wrap text for notes columns
        if col in [6, 7]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Add hyperlink for evidence file (Google Drive or local)
    link_cell = ws.cell(row=next_row, column=9)
    link_cell.border = thin_border

    if drive_link.startswith("https://"):
        link_cell.value = "View"
        link_cell.hyperlink = drive_link
        link_cell.font = Font(color="0563C1", underline="single")
    else:
        link_cell.hyperlink = drive_link
        link_cell.value = drive_link
        link_cell.style = "Hyperlink"

    wb.save(CONTROL_SHEET)

    # Sort the sheet after adding
    sort_control_sheet()

    return next_row


def sort_control_sheet():
    """Sort control sheet by Ref."""
    wb = load_or_create_control_sheet()
    ws = wb.active

    if ws.max_row <= 1:
        return  # Nothing to sort

    # Get all data rows (skip header)
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, len(CONTROL_COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append(
                {
                    "value": cell.value,
                    "fill": cell.fill.copy() if cell.fill else None,
                    "alignment": cell.alignment.copy() if cell.alignment else None,
                    "font": cell.font.copy() if cell.font else None,
                    "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
                }
            )
        data_rows.append(row_data)

    # Sort by Ref (col 1)
    data_rows.sort(key=lambda x: x[0]["value"] or "")

    # Write sorted data back
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_data["value"]
            if cell_data["fill"]:
                cell.fill = cell_data["fill"]
            if cell_data["alignment"]:
                cell.alignment = cell_data["alignment"]
            if cell_data["font"]:
                cell.font = cell_data["font"]
            if cell_data["hyperlink"]:
                cell.hyperlink = cell_data["hyperlink"]

    wb.save(CONTROL_SHEET)


def find_existing_result(project: str, company: str, feature: str) -> int:
    """Find existing row for a feature result."""
    wb = load_or_create_control_sheet()
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if (
            ws.cell(row=row, column=2).value == project
            and ws.cell(row=row, column=3).value == company
            and ws.cell(row=row, column=4).value == feature
        ):
            return row

    return None


def update_feature_result(
    row: int,
    status: str = None,
    intuit_notes: str = None,
    internal_notes: str = None,
    evidence_file: str = None,
):
    """Update existing feature result row."""
    wb = load_or_create_control_sheet()
    ws = wb.active

    if status:
        cell = ws.cell(row=row, column=5, value=status)
        if status in STATUS_COLORS:
            cell.fill = PatternFill(
                start_color=STATUS_COLORS[status],
                end_color=STATUS_COLORS[status],
                fill_type="solid",
            )

    if intuit_notes:
        cell = ws.cell(row=row, column=6, value=intuit_notes)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    if internal_notes:
        cell = ws.cell(row=row, column=7, value=internal_notes)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    if evidence_file:
        ws.cell(row=row, column=8, value=evidence_file)
        # Update link with Google Drive link
        drive_link = get_drive_link(evidence_file)
        link_cell = ws.cell(row=row, column=9)

        if drive_link.startswith("https://"):
            link_cell.value = "View"
            link_cell.hyperlink = drive_link
            link_cell.font = Font(color="0563C1", underline="single")
        else:
            link_cell.hyperlink = drive_link
            link_cell.value = drive_link
            link_cell.style = "Hyperlink"

    wb.save(CONTROL_SHEET)


def get_summary() -> dict:
    """Get summary statistics from control sheet."""
    wb = load_or_create_control_sheet()
    ws = wb.active

    summary = {"YES": 0, "PARTIAL": 0, "NO": 0, "NA": 0, "Testing": 0, "total": 0}

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=5).value
        if status in summary:
            summary[status] += 1
            summary["total"] += 1

    return summary


# Main entry point for generating spreadsheet
if __name__ == "__main__":
    import sys

    sys.stdout.reconfigure(encoding="utf-8")

    print("Gerando planilha completa a partir de features_rich.json...")
    output = generate_full_spreadsheet("TCO", "Apex Tire")
    print(f"\nPlanilha salva em: {output}")
