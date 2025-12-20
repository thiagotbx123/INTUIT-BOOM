# -*- coding: utf-8 -*-
"""
WFS Project Plan Excel Generator
Creates comprehensive project plan with all insights
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_project_plan():
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(
        start_color="FFE66D", end_color="FFE66D", fill_type="solid"
    )
    low_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    done_fill = PatternFill(start_color="95D5B2", end_color="95D5B2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row=1, cols=10):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # ============================================================
    # SHEET 1: Project Overview
    # ============================================================
    ws1 = wb.active
    ws1.title = "Project Overview"

    ws1["A1"] = "WFS PROJECT OVERVIEW"
    ws1["A1"].font = Font(bold=True, size=16)
    ws1.merge_cells("A1:D1")

    overview_data = [
        ["Field", "Value", "Status", "Notes"],
        ["Project Name", "WFS (Workforce Solutions)", "Active", "Intuit HCM Platform"],
        ["Client", "Intuit", "Active", "QuickBooks Online"],
        ["TestBox Lead", "Katherine Lu", "Active", "Program Lead"],
        ["TSA Assigned", "Thiago Rodrigues", "Active", "Technical Solutions Architect"],
        ["TSA Lead", "Lucas Wakigawa (Waki)", "Active", "Team Lead"],
        ["Alpha Access Date", "December 12, 2025", "DONE", "GoQuick Alpha Retail"],
        ["SOW Target Date", "December 19, 2025", "IN PROGRESS", "Today!"],
        ["WFS Alpha Launch", "December 2025", "Upcoming", "~100 customers"],
        ["WFS Beta Launch", "February 2026", "Planned", "Early Access Feb 4"],
        ["WFS GA Launch", "May 1, 2026", "Planned", "Full launch"],
        [
            "Exploration Environment",
            "GoQuick Alpha Retail",
            "DONE",
            "Realm ID: 9341455848423964",
        ],
        ["Mineral HR Access", "Confirmed via SSO", "DONE", "PR_ELITE billing plan"],
        ["QBO Routes Mapped", "47 routes", "DONE", "100% functional"],
        ["WFS Reports Identified", "27 reports", "DONE", "All with URLs"],
        ["Screenshots Captured", "54+ screenshots", "DONE", "QBO + Mineral"],
    ]

    for row_idx, row_data in enumerate(overview_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
            elif "DONE" in str(value):
                cell.fill = done_fill
            elif "IN PROGRESS" in str(value):
                cell.fill = medium_fill

    auto_width(ws1)

    # ============================================================
    # SHEET 2: Timeline & Milestones
    # ============================================================
    ws2 = wb.create_sheet("Timeline")

    timeline_data = [
        ["Phase", "Milestone", "Target Date", "Status", "Owner", "Dependencies"],
        [
            "Exploration",
            "Alpha Access Granted",
            "Dec 12, 2025",
            "DONE",
            "Intuit",
            "None",
        ],
        [
            "Exploration",
            "QBO Complete Mapping",
            "Dec 19, 2025",
            "DONE",
            "Thiago",
            "Alpha Access",
        ],
        [
            "Exploration",
            "Mineral HR Mapping",
            "Dec 19, 2025",
            "DONE",
            "Thiago",
            "Alpha Access",
        ],
        [
            "Exploration",
            "Intuit Intelligence Discovery",
            "Dec 19, 2025",
            "DONE",
            "Thiago",
            "Alpha Access",
        ],
        [
            "SOW",
            "SOW Draft Complete",
            "Dec 19, 2025",
            "IN PROGRESS",
            "Thiago/Katherine",
            "Exploration",
        ],
        ["SOW", "Internal Review", "Dec 20, 2025", "Pending", "Katherine", "SOW Draft"],
        [
            "SOW",
            "SOW to Client",
            "Dec 23, 2025",
            "Pending",
            "Katherine",
            "Internal Review",
        ],
        ["SOW", "SOW Approval", "Jan 6, 2026", "Pending", "Intuit", "SOW to Client"],
        [
            "Winter Release",
            "Feature Analysis",
            "Dec 22, 2025",
            "Pending",
            "Thiago/Waki",
            "Doc Access",
        ],
        [
            "Winter Release",
            "Early Access",
            "Feb 4, 2026",
            "Planned",
            "Intuit",
            "SOW Approval",
        ],
        [
            "Alpha",
            "Wave 1 Build Start",
            "Jan 13, 2026",
            "Planned",
            "Thiago",
            "SOW Approval",
        ],
        [
            "Alpha",
            "Demo Stories Ready",
            "Jan 27, 2026",
            "Planned",
            "Thiago",
            "Wave 1 Build",
        ],
        [
            "Alpha",
            "Health Checks Automated",
            "Feb 3, 2026",
            "Planned",
            "Thiago",
            "Demo Stories",
        ],
        [
            "Beta",
            "Wave 2 Build Start",
            "Feb 10, 2026",
            "Planned",
            "Thiago",
            "Alpha Complete",
        ],
        [
            "Beta",
            "Enhanced Data Ready",
            "Mar 3, 2026",
            "Planned",
            "Thiago",
            "Wave 2 Build",
        ],
        [
            "GA",
            "Wave 3 Build Start",
            "Apr 7, 2026",
            "Planned",
            "Thiago",
            "Beta Complete",
        ],
        ["GA", "Production Ready", "May 1, 2026", "Planned", "Team", "Wave 3 Build"],
    ]

    for row_idx, row_data in enumerate(timeline_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "DONE" in str(value):
                cell.fill = done_fill
            elif "IN PROGRESS" in str(value):
                cell.fill = medium_fill

    auto_width(ws2)

    # ============================================================
    # SHEET 3: Action Items
    # ============================================================
    ws3 = wb.create_sheet("Action Items")

    action_data = [
        ["Priority", "Action Item", "Owner", "Due Date", "Status", "Blocker?", "Notes"],
        [
            "P0",
            "Finalize SOW draft with Mineral scope",
            "Thiago",
            "Dec 19",
            "In Progress",
            "No",
            "Include Mineral HR integration",
        ],
        [
            "P0",
            "Review Winter Release doc (24 features)",
            "Thiago/Waki",
            "Dec 22",
            "Pending",
            "No",
            "Identify data/automation needs",
        ],
        [
            "P0",
            "Share SOW internally for review",
            "Thiago",
            "Dec 20",
            "Pending",
            "No",
            "Get Katherine's feedback",
        ],
        [
            "P1",
            "Follow-up with Kev about click path",
            "Katherine",
            "Dec 23",
            "Pending",
            "Yes",
            "Blocking demo planning",
        ],
        [
            "P1",
            "Document Winter Release matrix",
            "Thiago",
            "Dec 22",
            "Pending",
            "No",
            "Feature vs effort analysis",
        ],
        [
            "P1",
            "Confirm final demo environment",
            "Katherine",
            "Jan 6",
            "Pending",
            "No",
            "GoQuick is temporary",
        ],
        [
            "P1",
            "Request PRDs from Intuit",
            "Katherine",
            "Jan 6",
            "Pending",
            "Yes",
            "Main blocker for scope",
        ],
        [
            "P2",
            "Test employee creation via UI",
            "Thiago",
            "Jan 13",
            "Pending",
            "No",
            "Part of Wave 1",
        ],
        [
            "P2",
            "Test time entry creation via UI",
            "Thiago",
            "Jan 13",
            "Pending",
            "No",
            "Part of Wave 1",
        ],
        [
            "P2",
            "Setup automated health checks",
            "Thiago",
            "Jan 27",
            "Pending",
            "No",
            "Daily automation",
        ],
        [
            "P2",
            "Validate demo stories with Intuit",
            "Katherine",
            "Jan 27",
            "Pending",
            "No",
            "Before Beta",
        ],
        [
            "P3",
            "Document Mineral API (if available)",
            "Thiago",
            "Feb 10",
            "Pending",
            "No",
            "Contact Mitratech if needed",
        ],
        [
            "P3",
            "Create enablement documentation",
            "Thiago",
            "Apr 28",
            "Pending",
            "No",
            "For GA launch",
        ],
    ]

    for row_idx, row_data in enumerate(action_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "P0" in str(value):
                cell.fill = high_fill
            elif "P1" in str(value):
                cell.fill = medium_fill
            elif "P2" in str(value) or "P3" in str(value):
                cell.fill = low_fill

    auto_width(ws3)

    # ============================================================
    # SHEET 4: Risks
    # ============================================================
    ws4 = wb.create_sheet("Risks")

    risk_data = [
        [
            "ID",
            "Risk",
            "Probability",
            "Impact",
            "Severity",
            "Mitigation",
            "Owner",
            "Status",
        ],
        [
            "R1",
            "PRDs not received from Intuit",
            "Medium",
            "High",
            "HIGH",
            "Modular SOW that can be refined",
            "Katherine",
            "Open",
        ],
        [
            "R2",
            "Final demo environment different from Alpha",
            "High",
            "Medium",
            "HIGH",
            "Document transition plan",
            "Katherine",
            "Open",
        ],
        [
            "R3",
            "Click path for reps undefined",
            "High",
            "High",
            "CRITICAL",
            "Follow-up with Kev urgently",
            "Katherine",
            "Open",
        ],
        [
            "R4",
            "Mineral integration details TBD",
            "High",
            "Medium",
            "HIGH",
            "Separate scope for Mineral",
            "Thiago",
            "Open",
        ],
        [
            "R5",
            "Winter Release overlap with WFS",
            "High",
            "Medium",
            "HIGH",
            "Plan capacity early",
            "Waki",
            "Open",
        ],
        [
            "R6",
            "Staged rollout impacts all sellers",
            "Medium",
            "High",
            "HIGH",
            "Careful implementation plan",
            "Katherine",
            "Open",
        ],
        [
            "R7",
            "License count unclear for Feb",
            "Medium",
            "Low",
            "MEDIUM",
            "Confirm with Christina",
            "Katherine",
            "Open",
        ],
        [
            "R8",
            "SOW deadline missed",
            "Low",
            "High",
            "MEDIUM",
            "Finalize today",
            "Thiago",
            "In Progress",
        ],
    ]

    for row_idx, row_data in enumerate(risk_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "CRITICAL" in str(value):
                cell.fill = high_fill
            elif "HIGH" in str(value):
                cell.fill = medium_fill

    auto_width(ws4)

    # ============================================================
    # SHEET 5: Data Architecture
    # ============================================================
    ws5 = wb.create_sheet("Data Architecture")

    arch_data = [
        ["Component", "Type", "Source", "Target", "Data Flow", "Status", "Notes"],
        [
            "Employee ID",
            "Golden Thread",
            "QBO Payroll",
            "All Modules",
            "Bidirectional",
            "Confirmed",
            "Links employee across system",
        ],
        [
            "Time Entries",
            "Transactional",
            "Time Module",
            "Payroll/Projects",
            "One-way",
            "Confirmed",
            "Flows to paycheck",
        ],
        [
            "PTO Requests",
            "Transactional",
            "Time Off",
            "Payroll",
            "One-way",
            "Confirmed",
            "Affects pay calculation",
        ],
        [
            "Job Costing",
            "Analytical",
            "Projects",
            "Reports",
            "Aggregated",
            "Confirmed",
            "For profitability analysis",
        ],
        [
            "Payroll Data",
            "Master",
            "Payroll",
            "Reports/Tax",
            "One-way",
            "Confirmed",
            "Core HR data",
        ],
        [
            "Worker Profile",
            "Master",
            "Employees",
            "All Modules",
            "Reference",
            "Confirmed",
            "Central employee record",
        ],
        [
            "Mineral HR Data",
            "External",
            "QBO",
            "Mineral",
            "SSO + Data",
            "Partial",
            "Need click path from Kev",
        ],
        [
            "Org Chart",
            "Structural",
            "Employees",
            "Directory",
            "Hierarchical",
            "Confirmed",
            "Shows reporting structure",
        ],
        [
            "Documents",
            "Content",
            "Team Module",
            "Employees",
            "Association",
            "Confirmed",
            "Employee documents",
        ],
        [
            "Benefits",
            "Master",
            "Team Module",
            "Payroll",
            "Reference",
            "Confirmed",
            "Deductions setup",
        ],
        [
            "Workers Comp",
            "Compliance",
            "Team Module",
            "Reports",
            "Calculated",
            "Confirmed",
            "State-specific",
        ],
    ]

    for row_idx, row_data in enumerate(arch_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "Partial" in str(value):
                cell.fill = medium_fill
            elif "Confirmed" in str(value):
                cell.fill = done_fill

    auto_width(ws5)

    # ============================================================
    # SHEET 6: Communications (with suggested messages)
    # ============================================================
    ws6 = wb.create_sheet("Communications")

    comm_data = [
        [
            "Context",
            "Recipient",
            "Channel",
            "Urgency",
            "Suggested Message (English)",
            "Notes",
        ],
        [
            "PRD Follow-up",
            "Katherine",
            "Slack DM",
            "High",
            "Hey Katherine, quick check on the PRDs from Intuit. Do we have any update on when we might get them? I want to make sure our SOW scope is as accurate as possible. Let me know if there's anything I can do to help move this along.",
            "Main blocker",
        ],
        [
            "Click Path Question",
            "Kev (via Katherine)",
            "Slack Channel",
            "High",
            "Hey Kev, following up on my earlier question. We got Mineral access working through QBO and everything looks good on our side. Just wanted to check - what's the click path that sales reps should follow to show data flowing from WFS into Mineral during demos? And what data should we expect to see in Mineral to prove the connection is working? Thanks!",
            "Blocking demo planning",
        ],
        [
            "SOW Internal Review",
            "Katherine",
            "Slack DM",
            "High",
            "Hi Katherine, I just finished the SOW draft with the Mineral scope included. Can you take a look when you have a chance? I tried to keep it modular so we can easily update it once we get the PRDs. Let me know if anything needs changes.",
            "Before holiday",
        ],
        [
            "Winter Release Analysis",
            "Katherine/Waki",
            "Slack Channel",
            "Medium",
            "Hey team, I'm starting to look at the Winter Release features doc. I'll create a matrix showing which features need new data or automations vs what we already have. Should have something ready by end of day tomorrow. Any specific features you want me to prioritize?",
            "Before Dec 22",
        ],
        [
            "Environment Confirmation",
            "Katherine (for Christina)",
            "Slack DM",
            "Medium",
            "Hey Katherine, quick question about the demo environment. I know GoQuick Alpha Retail is just for our exploration. Do we have any update on what the final demo environment will be for sales? Want to make sure we plan the transition properly.",
            "Important for Wave 1",
        ],
        [
            "License Count",
            "Katherine",
            "Slack DM",
            "Low",
            "Hi Katherine, when you get a chance, could you check with Christina about the exact number of licenses we'll have for February? I saw the PO mentions 60 additional, just want to confirm for our planning.",
            "Non-urgent",
        ],
        [
            "Weekly Status Update",
            "Katherine/Team",
            "Slack Channel",
            "Medium",
            "Weekly WFS Update:\n\n Status: On Track\n\n Done this week:\n- Completed QBO exploration (47 routes mapped)\n- Confirmed Mineral HR access via SSO\n- Identified 27 WFS-related reports\n- Discovered Intuit Intelligence (AI BETA)\n- Captured 54+ screenshots\n\n In Progress:\n- Finalizing SOW draft\n- Reviewing Winter Release features\n\n Blockers:\n- PRDs still pending\n- Click path for demos TBD\n\n Next Week:\n- Share SOW for internal review\n- Complete Winter Release analysis",
            "Template for weekly updates",
        ],
        [
            "Exploration Complete",
            "Katherine",
            "Slack DM",
            "Medium",
            "Good news! I finished the full exploration of both QBO and Mineral. Here's what we have:\n- 47 QBO routes mapped and working\n- 27 WFS reports identified with direct URLs\n- Mineral HR fully accessible via SSO\n- Found Intuit Intelligence (their AI assistant) in beta\n- 54+ screenshots captured\n\nAll documented and ready for the SOW. Let me know if you need anything else!",
            "Milestone communication",
        ],
        [
            "Kev Introduction",
            "Kev",
            "Slack DM",
            "Low",
            "Hey Kev, I'm Thiago from the TestBox team working on the WFS demo environments. Katherine mentioned you're the go-to person for Mineral integration questions. Would love to connect when you have a few minutes to chat about the data flow between WFS and Mineral. No rush, just whenever works for you.",
            "If direct contact needed",
        ],
        [
            "SOW to Client",
            "Katherine (for Intuit)",
            "Email draft",
            "High",
            "Subject: WFS Demo Environment SOW - Ready for Review\n\nHi team,\n\nAttached is the Statement of Work for the WFS demo environment build. Key highlights:\n\n- Wave 1 (Alpha): Demo environment setup, baseline data, 2 demo stories\n- Wave 2 (Beta): Enhanced data, additional demo stories, automation\n- Wave 3 (GA): Production reliability, monitoring, scale testing\n\nAssumptions and open items are documented in Section X. We'll refine the scope once we receive the PRDs.\n\nPlease let us know if you have any questions or need clarification on anything.\n\nBest,\n[Name]",
            "For Katherine to send",
        ],
    ]

    for row_idx, row_data in enumerate(comm_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "High" in str(row_data[3]) if col_idx == 4 else False:
                cell.fill = high_fill

    # Set column widths for readability
    ws6.column_dimensions["A"].width = 20
    ws6.column_dimensions["B"].width = 15
    ws6.column_dimensions["C"].width = 15
    ws6.column_dimensions["D"].width = 10
    ws6.column_dimensions["E"].width = 80
    ws6.column_dimensions["F"].width = 20

    # ============================================================
    # SHEET 7: Environment Mapping
    # ============================================================
    ws7 = wb.create_sheet("Environment Mapping")

    env_data = [
        ["Environment", "Realm ID", "Purpose", "Access", "Features", "Status", "Notes"],
        [
            "GoQuick Alpha Retail",
            "9341455848423964",
            "Exploration Only",
            "Thiago, Katherine, Saubhagya",
            "Full WFS, Mineral, AI",
            "Active",
            "NOT final demo env",
        ],
        [
            "Keystone Construction",
            "TBD",
            "Demo Stories",
            "TBD",
            "Job Costing, PTO",
            "Planned",
            "Better for demos",
        ],
        [
            "Final Demo Env",
            "TBD",
            "Sales Demos",
            "All reps",
            "Full WFS",
            "TBD",
            "Need confirmation",
        ],
        [
            "Mineral HR",
            "Via SSO",
            "HR Compliance",
            "Client Admins",
            "Templates, Training, Tools",
            "Active",
            "PR_ELITE plan",
        ],
    ]

    for row_idx, row_data in enumerate(env_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws7.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "Active" in str(value):
                cell.fill = done_fill
            elif "TBD" in str(value):
                cell.fill = medium_fill

    auto_width(ws7)

    # ============================================================
    # SHEET 8: WFS Reports
    # ============================================================
    ws8 = wb.create_sheet("WFS Reports")

    reports_data = [
        ["Category", "Report Name", "URL Path", "WFS Relevance", "Demo Story"],
        [
            "Time",
            "Unbilled Time",
            "/app/report/builder?token=UNBILLED_TIME",
            "High",
            "Job Costing",
        ],
        [
            "Time",
            "Time Activities by Customer",
            "/app/report/builder?token=TIME_ACTIVITIES_BY_CUST",
            "High",
            "Job Costing",
        ],
        [
            "Time",
            "Time Activities by Employee",
            "/app/report/builder?token=TIME_ACTIVITIES",
            "High",
            "Payroll Lifecycle",
        ],
        [
            "Time",
            "Recent Time Activities",
            "/app/report/builder?token=RECENT_TIME_ACTIVITIES",
            "Medium",
            "Time Tracking",
        ],
        [
            "Time",
            "Project Profitability",
            "/app/reportv2?rptId=PROJECT_PROFITABILITY_SUMMARY",
            "High",
            "Job Costing",
        ],
        [
            "Employee",
            "Employee Contact List",
            "/app/report/builder",
            "High",
            "HR Management",
        ],
        [
            "Employee",
            "Employee Directory",
            "/app/payroll/reports/employee-directory",
            "High",
            "HR Management",
        ],
        [
            "Employee",
            "Employee Details",
            "/app/payroll/reports/employee-details",
            "High",
            "HR Management",
        ],
        [
            "Payroll",
            "Payroll Summary",
            "/app/payroll/reports/payroll-summary",
            "High",
            "Payroll Lifecycle",
        ],
        [
            "Payroll",
            "Payroll by Employee",
            "/app/payroll/reports/payroll-summary?name=pse",
            "High",
            "Payroll Lifecycle",
        ],
        [
            "Payroll",
            "Payroll Details",
            "/app/payroll/reports/payroll-summary?name=details",
            "High",
            "Payroll Lifecycle",
        ],
        [
            "Payroll",
            "Paycheck History",
            "/app/payroll/reports/paycheck-history",
            "High",
            "Payroll Lifecycle",
        ],
        [
            "Payroll",
            "Deductions/Contributions",
            "/app/payroll/reports/deductions-contributions",
            "Medium",
            "Benefits",
        ],
        [
            "Payroll",
            "Payroll Item List",
            "/app/payroll/reports/payroll-item-list",
            "Medium",
            "Setup",
        ],
        [
            "Payroll",
            "Total Payroll Cost",
            "/app/payroll/reports/total-payroll-cost",
            "High",
            "Cost Analysis",
        ],
        [
            "Payroll",
            "Total Pay",
            "/app/payroll/reports/total-pay",
            "High",
            "Cost Analysis",
        ],
        [
            "Tax",
            "Tax Liability",
            "/app/payroll/reports/tax-liability",
            "Medium",
            "Compliance",
        ],
        [
            "Tax",
            "Tax Payments",
            "/app/payroll/reports/tax-payments",
            "Medium",
            "Compliance",
        ],
        [
            "Tax",
            "Tax and Wage Summary",
            "/app/payroll/reports/tax-wage-summary",
            "Medium",
            "Compliance",
        ],
        [
            "Benefits",
            "Vacation and Sick Leave",
            "/app/payroll/reports/vacation-sick-leave",
            "High",
            "Unified PTO",
        ],
        [
            "Benefits",
            "Retirement Plans",
            "/app/payroll/reports/retirement-plans",
            "Medium",
            "Benefits",
        ],
        [
            "Benefits",
            "State Mandated Retirement",
            "/app/payroll/reports/retirement-plans?name=statemandretplans",
            "Low",
            "Compliance",
        ],
        [
            "Compliance",
            "Workers Compensation",
            "/app/payroll/reports/workers-compensation",
            "Medium",
            "Compliance",
        ],
        [
            "Compliance",
            "Multiple Worksites",
            "/app/payroll/reports/multiple-worksite",
            "Low",
            "Compliance",
        ],
        [
            "Contractor",
            "1099 Balance Detail",
            "/app/report/builder?SubToken=contractorBalanceDetail",
            "Medium",
            "Contractors",
        ],
        [
            "Contractor",
            "1099 Balance Summary",
            "/app/report/builder?SubToken=contractorBalanceSummary",
            "Medium",
            "Contractors",
        ],
        [
            "Contractor",
            "Contractor Payments",
            "/app/payroll/reports/contractor-payments",
            "Medium",
            "Contractors",
        ],
    ]

    for row_idx, row_data in enumerate(reports_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws8.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "High" in str(value):
                cell.fill = done_fill
            elif "Medium" in str(value):
                cell.fill = low_fill

    auto_width(ws8)

    # ============================================================
    # SHEET 9: Metrics & KPIs
    # ============================================================
    ws9 = wb.create_sheet("Metrics KPIs")

    metrics_data = [
        ["Phase", "Metric", "Target", "Current", "Status", "Measurement"],
        ["Alpha", "SOW Approved", "Yes", "In Progress", "On Track", "Client sign-off"],
        ["Alpha", "Environment Mapped", "100%", "100%", "DONE", "47/47 routes"],
        [
            "Alpha",
            "Demo Stories Ready",
            "2+",
            "0",
            "Pending",
            "Unified PTO, Magic Docs",
        ],
        [
            "Alpha",
            "Health Check Automated",
            "Yes",
            "No",
            "Pending",
            "Daily script running",
        ],
        [
            "Beta",
            "Demo Stories Ready",
            "4+",
            "0",
            "Planned",
            "Add Job Costing, Hire to Fire",
        ],
        ["Beta", "Data Refresh Automated", "Yes", "No", "Planned", "Weekly refresh"],
        [
            "Beta",
            "Mineral Integration",
            "Documented",
            "Partial",
            "In Progress",
            "Need click path",
        ],
        [
            "Beta",
            "Winter Release Features",
            "Incorporated",
            "Pending",
            "Planned",
            "24 features analyzed",
        ],
        ["GA", "Demo Stories Ready", "6+", "0", "Planned", "Full suite"],
        ["GA", "Production Monitoring", "Yes", "No", "Planned", "Alerts configured"],
        ["GA", "Scale Test", "Passed", "Not Started", "Planned", "Multiple instances"],
        [
            "GA",
            "Enablement Docs",
            "Complete",
            "Not Started",
            "Planned",
            "Training materials",
        ],
    ]

    for row_idx, row_data in enumerate(metrics_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws9.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif "DONE" in str(value):
                cell.fill = done_fill
            elif "On Track" in str(value) or "In Progress" in str(value):
                cell.fill = low_fill
            elif "Pending" in str(value) or "Planned" in str(value):
                cell.fill = medium_fill

    auto_width(ws9)

    # Save the workbook
    output_path = "C:/Users/adm_r/intuit-boom/WFS_PROJECT_PLAN_COMPLETE.xlsx"
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    create_project_plan()
