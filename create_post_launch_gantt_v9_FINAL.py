"""
Post-Launch Requests - Winter/February Release 2026
VERSION 9 FINAL - Fixed date logic: Complete ≤ today, Target = 02-25
Today: 2026-01-20
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUTPUT_FILE = r"C:\Users\adm_r\intuit-boom\POST_LAUNCH_WINTER_RELEASE_2026_FINAL.xlsx"

# Cores
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GATE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
COMPLETE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
IN_PROGRESS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NOT_STARTED_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BLOCKED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
CE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GTM_FILL = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
CLIENT_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GATE_FONT = Font(bold=True, color="FFFFFF", size=10)
CATEGORY_FONT = Font(bold=True, color="FFFFFF", size=10, italic=True)
NORMAL_FONT = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_gantt():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Post-Launch Requests 2026"

    headers = ["Task / Deliverable", "Owner", "Status", "Start", "End"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

    tasks = [
        # ========== GATE 1: INTAKE & ALIGNMENT ==========
        {"task": "GATE 1: INTAKE & ALIGNMENT", "is_gate": True},
        {
            "task": "Sends feature request/scope for Winter Release",
            "owner": "Customer",
            "status": "Complete",
            "start": "2026-01-02",
            "end": "2026-01-06",
            "is_client": True,
        },
        {
            "task": "Review request and document understanding",
            "owner": "GTM + FDE",
            "status": "Complete",
            "start": "2026-01-06",
            "end": "2026-01-10",
        },
        {
            "task": "Validates our understanding of scope",
            "owner": "Customer",
            "status": "Complete",
            "start": "2026-01-10",
            "end": "2026-01-13",
            "is_client": True,
        },
        {
            "task": "Address customer questions/clarifications",
            "owner": "GTM + FDE",
            "status": "Complete",
            "start": "2026-01-13",
            "end": "2026-01-17",
        },
        {
            "task": "Approves final scope",
            "owner": "Customer",
            "status": "In Progress",
            "start": "2026-01-17",
            "end": "2026-01-21",
            "is_client": True,
        },
        {
            "task": "Create validation checklist per category",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-20",
            "end": "2026-01-23",
        },
        # ========== GATE 2: ENVIRONMENT CONFIRMATION (all Complete = end ≤ today) ==========
        {"task": "GATE 2: ENVIRONMENT CONFIRMATION", "is_gate": True},
        {
            "task": "Identify available CIDs in TestBox",
            "owner": "FDE + GTM",
            "status": "Complete",
            "start": "2026-01-06",
            "end": "2026-01-13",
        },
        {
            "task": "Confirms CIDs with customer",
            "owner": "GTM",
            "status": "Complete",
            "start": "2026-01-13",
            "end": "2026-01-15",
            "is_gtm": True,
        },
        {
            "task": "Confirms target CIDs for validation",
            "owner": "Customer",
            "status": "Complete",
            "start": "2026-01-15",
            "end": "2026-01-17",
            "is_client": True,
        },
        {
            "task": "Aligns with FDE which CIDs to validate",
            "owner": "GTM",
            "status": "Complete",
            "start": "2026-01-17",
            "end": "2026-01-17",
            "is_gtm": True,
        },
        {
            "task": "Validates TCO access (CID + credentials)",
            "owner": "FDE",
            "status": "Complete",
            "start": "2026-01-13",
            "end": "2026-01-20",
        },
        {
            "task": "Validates Construction Events access (CID + credentials)",
            "owner": "FDE",
            "status": "Complete",
            "start": "2026-01-13",
            "end": "2026-01-20",
        },
        {
            "task": "Validates Construction Sales access (CID + credentials)",
            "owner": "FDE",
            "status": "Complete",
            "start": "2026-01-13",
            "end": "2026-01-20",
        },
        # ========== GATE 3: FEATURE VALIDATION ==========
        {"task": "GATE 3: FEATURE VALIDATION", "is_gate": True},
        {
            "task": "[AI AGENTS] WR-001 Accounting AI - Bank transaction suggestions",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-20",
            "end": "2026-01-27",
        },
        {
            "task": "[AI AGENTS] WR-002 Sales Tax AI - Pre-file validation",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-20",
            "end": "2026-01-27",
        },
        {
            "task": "[AI AGENTS] WR-003 Project Management AI - Budget/profitability",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-20",
            "end": "2026-01-27",
        },
        {
            "task": "[AI AGENTS] WR-004 Finance AI - Dashboard KPIs + AI insights",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-20",
            "end": "2026-01-27",
        },
        {
            "task": "[AI AGENTS] WR-005 Solutions Specialist - Business Feed dimensions",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-21",
            "end": "2026-01-28",
        },
        {
            "task": "[AI AGENTS] WR-006 Customer Agent - Leads tab + email import",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-21",
            "end": "2026-01-28",
        },
        {
            "task": "[AI AGENTS] WR-007 Intuit Intelligence - Chat interface",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-21",
            "end": "2026-01-28",
        },
        {
            "task": "[AI AGENTS] WR-008 Conversational BI - Reports AI query",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-27",
            "end": "2026-01-30",
        },
        {
            "task": "[REPORTING] WR-009 KPIs Customizados - Scorecard tiles",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-22",
            "end": "2026-01-29",
        },
        {
            "task": "[REPORTING] WR-010 Dashboards - Pre-built + custom",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-22",
            "end": "2026-01-29",
        },
        {
            "task": "[REPORTING] WR-011 3P Data Integrations - Apps marketplace",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-23",
            "end": "2026-01-30",
        },
        {
            "task": "[REPORTING] WR-012 Calculated Fields - Custom report fields",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-23",
            "end": "2026-01-30",
        },
        {
            "task": "[REPORTING] WR-013 Management Reports - Full financials",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-24",
            "end": "2026-01-31",
        },
        {
            "task": "[REPORTING] WR-014 Benchmarking - Performance Center",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-24",
            "end": "2026-01-31",
        },
        {
            "task": "[REPORTING] WR-015 Multi-Entity Reports - Consolidated View",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-27",
            "end": "2026-02-03",
        },
        {
            "task": "[DIMENSIONS] WR-016 Dimension Assignment v2 - AI sparkles",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-27",
            "end": "2026-01-31",
        },
        {
            "task": "[DIMENSIONS] WR-017 Hierarchical Dimension Reporting",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-27",
            "end": "2026-01-31",
        },
        {
            "task": "[DIMENSIONS] WR-018 Dimensions on Workflow",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-28",
            "end": "2026-02-03",
        },
        {
            "task": "[DIMENSIONS] WR-019 Dimensions on Balance Sheet",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-28",
            "end": "2026-02-03",
        },
        {
            "task": "[WORKFLOW] WR-020 Parallel Approval - Multiple approvers config",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-29",
            "end": "2026-02-03",
        },
        {
            "task": "[MIGRATION] WR-021 Seamless Desktop Migration - Import option",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-29",
            "end": "2026-02-03",
        },
        {
            "task": "[MIGRATION] WR-022 DFY Migration Experience - Done-for-you wizard",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-30",
            "end": "2026-02-04",
        },
        {
            "task": "[MIGRATION] WR-023 Feature Compatibility - Documentation only",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-30",
            "end": "2026-01-30",
        },
        {
            "task": "[CONSTRUCTION] WR-024 Certified Payroll Report - WH-347 format",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-30",
            "end": "2026-02-04",
        },
        {
            "task": "[CONSTRUCTION] WR-025 Sales Order - Creation interface",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-31",
            "end": "2026-02-04",
        },
        {
            "task": "[PAYROLL] WR-026 Multi-Entity Payroll Hub - Cross-entity view",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-01-31",
            "end": "2026-02-04",
        },
        {
            "task": "[PAYROLL] WR-027 Garnishments Child Support - Employee config",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-02-02",
            "end": "2026-02-05",
        },
        {
            "task": "[PAYROLL] WR-028 Assignments in QBTime - Time entry fields",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-02-02",
            "end": "2026-02-05",
        },
        {
            "task": "[PAYROLL] WR-029 Enhanced Amendments CA - N/A (requires CA)",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-02-05",
            "end": "2026-02-05",
        },
        # ========== CONSOLIDATED: DATA PIPELINE ==========
        {"task": ">> CONSOLIDATED: DATA PIPELINE", "is_category": True},
        {
            "task": "Identify all data gaps across categories",
            "owner": "FDE",
            "status": "In Progress",
            "start": "2026-02-03",
            "end": "2026-02-05",
        },
        {
            "task": "Create tickets - synthetic data generation",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-05",
            "end": "2026-02-05",
        },
        {
            "task": "Generate synthetic data (all categories)",
            "owner": "Data Engineering",
            "status": "Not Started",
            "start": "2026-02-05",
            "end": "2026-02-10",
            "is_de": True,
        },
        {
            "task": "Verify synthetic data quality",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-10",
            "end": "2026-02-11",
        },
        {
            "task": "Create tickets - data ingestion",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-05",
            "end": "2026-02-05",
        },
        {
            "task": "Execute data ingestion (all categories)",
            "owner": "CE",
            "status": "Not Started",
            "start": "2026-02-05",
            "end": "2026-02-11",
            "is_ce": True,
        },
        {
            "task": "Verify data ingestion completed",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-11",
            "end": "2026-02-12",
        },
        # ========== CONSOLIDATED: EVIDENCE PIPELINE ==========
        {"task": ">> CONSOLIDATED: EVIDENCE PIPELINE", "is_category": True},
        {
            "task": "Capture screenshots - AI Agents (TCO, Events, Sales)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-10",
            "end": "2026-02-12",
        },
        {
            "task": "Capture screenshots - Reporting (TCO, Events, Sales)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-10",
            "end": "2026-02-12",
        },
        {
            "task": "Capture screenshots - Dimensions (all environments)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-11",
            "end": "2026-02-13",
        },
        {
            "task": "Capture screenshots - Workflow (all environments)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-11",
            "end": "2026-02-13",
        },
        {
            "task": "Capture screenshots - Migration (TCO)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-12",
            "end": "2026-02-13",
        },
        {
            "task": "Capture screenshots - Construction (Events, Sales)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-12",
            "end": "2026-02-14",
        },
        {
            "task": "Capture screenshots - Payroll (all environments)",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-13",
            "end": "2026-02-15",
        },
        {
            "task": "Re-capture problematic evidence",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-15",
            "end": "2026-02-16",
        },
        {
            "task": "Organize evidence by category and environment",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-16",
            "end": "2026-02-17",
        },
        {
            "task": "Upload all evidence to Google Drive",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-17",
            "end": "2026-02-18",
        },
        {
            "task": "Update validation matrix with final status",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-18",
            "end": "2026-02-19",
        },
        {
            "task": "Generate validation report for customer",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-19",
            "end": "2026-02-20",
        },
        # ========== GATE 4: INTERNAL REVIEW ==========
        {"task": "GATE 4: INTERNAL REVIEW", "is_gate": True},
        {
            "task": "FDE + GTM review validation report together",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-20",
            "end": "2026-02-20",
        },
        {
            "task": "Identify gaps or missing evidence",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-20",
            "end": "2026-02-20",
        },
        {
            "task": "Address gaps before customer presentation",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-20",
            "end": "2026-02-21",
        },
        {
            "task": "Prepare customer presentation deck",
            "owner": "GTM",
            "status": "Not Started",
            "start": "2026-02-20",
            "end": "2026-02-21",
            "is_gtm": True,
        },
        # ========== GATE 5: CUSTOMER REVIEW (ends on target 02-25) ==========
        {"task": "GATE 5: CUSTOMER REVIEW", "is_gate": True},
        {
            "task": "Schedule review meeting with customer",
            "owner": "GTM",
            "status": "Not Started",
            "start": "2026-02-18",
            "end": "2026-02-19",
            "is_gtm": True,
        },
        {
            "task": "Present validation results to customer",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-21",
            "end": "2026-02-22",
        },
        {
            "task": "Customer reviews evidence and provides feedback",
            "owner": "Customer",
            "status": "Not Started",
            "start": "2026-02-22",
            "end": "2026-02-23",
            "is_client": True,
        },
        {
            "task": "Document customer feedback and action items",
            "owner": "GTM",
            "status": "Not Started",
            "start": "2026-02-23",
            "end": "2026-02-23",
            "is_gtm": True,
        },
        {
            "task": "Implement feedback adjustments",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-23",
            "end": "2026-02-24",
        },
        {
            "task": "Customer validates adjustments",
            "owner": "Customer",
            "status": "Not Started",
            "start": "2026-02-24",
            "end": "2026-02-24",
            "is_client": True,
        },
        {
            "task": "Final sign-off from customer",
            "owner": "Customer",
            "status": "Not Started",
            "start": "2026-02-24",
            "end": "2026-02-25",
            "is_client": True,
        },
        # ========== GATE 6: DOCUMENTATION & DELIVERY (internal, can overlap) ==========
        {"task": "GATE 6: DOCUMENTATION & DELIVERY", "is_gate": True},
        {
            "task": "Create Winter Release notes summary",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-24",
            "end": "2026-02-25",
        },
        {
            "task": "Archive all evidence and documentation",
            "owner": "FDE",
            "status": "Not Started",
            "start": "2026-02-25",
            "end": "2026-02-25",
        },
        {
            "task": "Document final validation status",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-24",
            "end": "2026-02-25",
        },
        # ========== GATE 7: POST-LAUNCH (internal TB only, after 02-25) ==========
        {"task": "GATE 7: POST-LAUNCH MONITORING (Internal TB)", "is_gate": True},
        {
            "task": "Monitor post-launch issues (2 weeks)",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-26",
            "end": "2026-03-12",
        },
        {
            "task": "Address customer-reported issues",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-02-26",
            "end": "2026-03-12",
        },
        {
            "task": "Final project retrospective",
            "owner": "FDE + GTM",
            "status": "Not Started",
            "start": "2026-03-12",
            "end": "2026-03-13",
        },
    ]

    row = 2
    for task_data in tasks:
        is_gate = task_data.get("is_gate", False)
        is_category = task_data.get("is_category", False)
        is_ce = task_data.get("is_ce", False)
        is_de = task_data.get("is_de", False)
        is_gtm = task_data.get("is_gtm", False)
        is_client = task_data.get("is_client", False)

        cell_task = ws.cell(row=row, column=1, value=task_data["task"])
        cell_task.border = THIN_BORDER

        cell_owner = ws.cell(row=row, column=2, value=task_data.get("owner", ""))
        cell_owner.alignment = Alignment(horizontal="center")
        cell_owner.border = THIN_BORDER

        cell_status = ws.cell(row=row, column=3, value=task_data.get("status", ""))
        cell_status.alignment = Alignment(horizontal="center")
        cell_status.border = THIN_BORDER

        cell_start = ws.cell(row=row, column=4, value=task_data.get("start", ""))
        cell_start.alignment = Alignment(horizontal="center")
        cell_start.border = THIN_BORDER

        cell_end = ws.cell(row=row, column=5, value=task_data.get("end", ""))
        cell_end.alignment = Alignment(horizontal="center")
        cell_end.border = THIN_BORDER

        if is_gate:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = GATE_FILL
                ws.cell(row=row, column=col).font = GATE_FONT
        elif is_category:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = CATEGORY_FILL
                ws.cell(row=row, column=col).font = CATEGORY_FONT
        elif is_client:
            cell_task.font = NORMAL_FONT
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = CLIENT_FILL
            status = task_data.get("status", "")
            if status == "Complete":
                cell_status.fill = COMPLETE_FILL
            elif status == "In Progress":
                cell_status.fill = IN_PROGRESS_FILL
            else:
                cell_status.fill = NOT_STARTED_FILL
        elif is_ce:
            cell_task.font = NORMAL_FONT
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = CE_FILL
            cell_status.fill = NOT_STARTED_FILL
        elif is_de:
            cell_task.font = NORMAL_FONT
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = DE_FILL
            cell_status.fill = NOT_STARTED_FILL
        elif is_gtm:
            cell_task.font = NORMAL_FONT
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = GTM_FILL
            cell_status.fill = NOT_STARTED_FILL
        else:
            cell_task.font = NORMAL_FONT
            status = task_data.get("status", "")
            if status == "Complete":
                cell_status.fill = COMPLETE_FILL
            elif status == "In Progress":
                cell_status.fill = IN_PROGRESS_FILL
            elif status == "Blocked":
                cell_status.fill = BLOCKED_FILL
            elif status == "Not Started":
                cell_status.fill = NOT_STARTED_FILL

        row += 1

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)

    print(f"FINAL Gantt v9: {OUTPUT_FILE}")
    print(f"Total lines: {row - 1}")
    print("Project Start: 2026-01-02")
    print("Target (Sign-off): 2026-02-25")
    print("Post-launch (Internal TB): 2026-02-26 to 2026-03-13")

    return OUTPUT_FILE


if __name__ == "__main__":
    create_gantt()
