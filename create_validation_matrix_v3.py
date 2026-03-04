"""
Winter Release FY26 - Validation Matrix v3
Complete validation guide with:
- Full login credentials (email, password, TOTP)
- All company CIDs from Intuit data
- Detailed step-by-step (no redundant login steps)
- Alternative validation methods in extra columns
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Styles
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NA_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Complete feature definitions with detailed instructions (NO LOGIN STEPS - those are in Logins sheet)
FEATURES = [
    {
        "ref": "WR-001",
        "name": "Accounting AI",
        "category": "AI Agents",
        "priority": "P0",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """ACCOUNTING AI - VALIDATION STEPS

WHAT THIS FEATURE DOES:
AI helps categorize bank transactions automatically. Look for sparkle icons next to suggested categories.

NAVIGATION PATH:
Left Menu > Transactions > Bank transactions

STEPS:
1. Click "Transactions" in the left menu
2. Click "Bank transactions"
3. Wait for page to fully load (10-15 seconds)
4. Look at the transaction list
5. Find transactions with a SPARKLE ICON (small star shape)
6. This sparkle means AI is suggesting a category
7. Take screenshot showing sparkle icons

WHAT TO LOOK FOR:
- Sparkle icon next to transaction category = AI active
- "Suggested" label = AI recommendation
- Auto-categorized transactions = AI working

SCREENSHOT CHECKLIST:
[ ] Bank transactions page visible
[ ] At least one sparkle icon visible
[ ] Transaction list showing

PASS CRITERIA:
- Sparkle icons visible on transactions = PASS
- AI suggestions appearing = PASS""",
        "alt_check_1": "Check Settings > Bank transactions > Auto-categorization enabled",
        "alt_check_2": "Run SQL: SELECT COUNT(*) FROM bank_transactions WHERE ai_suggested=1",
        "alt_check_3": "Check Intuit Assist for 'categorize transactions' command",
        "evidence_notes_template": "Banking > Bank Transactions shows AI sparkle icons on [X] transactions.",
        "client_note_template": "Accounting AI is active with auto-categorization enabled.",
    },
    {
        "ref": "WR-002",
        "name": "Sales Tax AI",
        "category": "AI Agents",
        "priority": "P0",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """SALES TAX AI - VALIDATION STEPS

WHAT THIS FEATURE DOES:
AI checks for errors before you file sales tax. Shows warnings and suggestions.

NAVIGATION PATH:
Left Menu > Taxes > Sales tax

STEPS:
1. Click "Taxes" in the left menu
2. Click "Sales tax"
3. Wait for Sales Tax Center to load
4. Look for "Filing Pre-Check" section
5. Look for AI warnings (usually in blue or yellow boxes)
6. Look for "Discrepancy" alerts
7. Take screenshot of pre-check area

WHAT TO LOOK FOR:
- "Pre-Check" or "Review before filing" = AI check
- Warning icons with suggestions = AI working
- Discrepancy alerts = AI finding issues

SCREENSHOT CHECKLIST:
[ ] Sales Tax Center visible
[ ] Pre-check section visible
[ ] Any warnings or alerts visible

PASS CRITERIA:
- Pre-check features visible = PASS
- AI warnings showing = PASS""",
        "alt_check_1": "Check if 'Automated Sales Tax' is enabled in Settings",
        "alt_check_2": "Look for tax rate suggestions on invoices",
        "alt_check_3": "Check Taxes > Payments for AI recommendations",
        "evidence_notes_template": "Taxes > Sales Tax shows AI pre-check features with [describe warnings].",
        "client_note_template": "Sales Tax AI provides pre-filing checks and discrepancy alerts.",
    },
    {
        "ref": "WR-003",
        "name": "Project Management AI",
        "category": "AI Agents",
        "priority": "P0",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": True,
        "step_by_step": """PROJECT MANAGEMENT AI - VALIDATION STEPS

WHAT THIS FEATURE DOES:
AI helps manage projects by showing profitability, suggesting budgets, and providing insights.

NAVIGATION PATH:
Left Menu > Projects > [Select any project]

STEPS:
1. Click "Projects" in the left menu
2. Wait for project list to load
3. Click on any project name to open it
4. Wait for project details (10-20 seconds)
5. Look for:
   - Profitability percentage or graph
   - "Insights" section
   - Sparkle icons for AI suggestions
   - Budget vs Actual comparison
6. Scroll down to see all sections
7. Take screenshot showing project insights

WHAT TO LOOK FOR:
- Profitability % = AI calculated
- Sparkle icons = AI suggestions
- "Insights" tab/section = AI analysis
- Budget recommendations = AI working

SCREENSHOT CHECKLIST:
[ ] Project name visible at top
[ ] Profitability info visible
[ ] Any AI insights or sparkles visible

PASS CRITERIA:
- Profitability tracking visible = PASS
- AI insights showing = PASS""",
        "alt_check_1": "Check Projects list for profitability column",
        "alt_check_2": "Look for 'Project Profitability' report in Reports",
        "alt_check_3": "SQL: SELECT status, COUNT(*) FROM projects GROUP BY status",
        "evidence_notes_template": "Projects > [Name] shows [X]% profitability with AI insights.",
        "client_note_template": "Project Management AI tracks profitability and provides insights.",
    },
    {
        "ref": "WR-004",
        "name": "Finance AI",
        "category": "AI Agents",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """FINANCE AI - VALIDATION STEPS

WHAT THIS FEATURE DOES:
AI provides financial summaries, insights, and recommendations.

NAVIGATION PATH:
Dashboard (Homepage) OR Left Menu > Reports

STEPS:
1. Go to Dashboard (click QuickBooks logo top left)
2. Look for "Intuit Assist" icon (chat bubble, usually bottom right)
3. Look for "Business overview" section with insights
4. If not on Dashboard, go to Reports
5. Look for AI-enhanced report sections
6. Check for "Insights" or "Recommendations" boxes
7. Take screenshot of any AI features

ALTERNATIVE - INTUIT ASSIST:
1. Find the Intuit Assist chat icon
2. Click to open
3. Ask: "What is my cash flow status?"
4. Screenshot the AI response

SCREENSHOT CHECKLIST:
[ ] Dashboard or Reports visible
[ ] Intuit Assist icon or chat visible
[ ] Any AI insights visible

PASS CRITERIA:
- Intuit Assist icon visible = PASS
- Financial insights showing = PASS""",
        "alt_check_1": "Check Dashboard widgets for AI insights",
        "alt_check_2": "Use Intuit Assist: 'Show my revenue trend'",
        "alt_check_3": "Check Reports > Business Snapshot for AI summary",
        "evidence_notes_template": "Finance AI visible via [Dashboard/Intuit Assist]. Shows [insights].",
        "client_note_template": "Finance AI provides automated financial insights and recommendations.",
    },
    {
        "ref": "WR-005",
        "name": "Solutions Specialist",
        "category": "AI Agents",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """SOLUTIONS SPECIALIST - VALIDATION STEPS

WHAT THIS FEATURE DOES:
AI provides personalized recommendations in Business Feed based on your business.

NAVIGATION PATH:
Dashboard > Business Feed (right side)

STEPS:
1. Go to Dashboard (click QuickBooks logo)
2. Look at the RIGHT side of the Dashboard
3. Find "Business Feed" or "Activity Feed" section
4. Scroll through the feed items
5. Look for:
   - "Recommended for you" items
   - Dimension widgets
   - IES-exclusive suggestions
   - Personalized tips
6. Take screenshot of Business Feed

WHAT TO LOOK FOR:
- Personalized recommendations = AI working
- Dimension widgets = IES feature active
- "Suggested" items = AI curated

SCREENSHOT CHECKLIST:
[ ] Dashboard visible
[ ] Business Feed section visible
[ ] At least one recommendation visible

PASS CRITERIA:
- Business Feed with recommendations = PASS""",
        "alt_check_1": "Check for 'Discover' or 'Explore' section on Dashboard",
        "alt_check_2": "Look for industry-specific tips in feed",
        "alt_check_3": "Check Apps page for recommended apps",
        "evidence_notes_template": "Dashboard > Business Feed shows AI recommendations.",
        "client_note_template": "Solutions Specialist provides personalized AI recommendations.",
    },
    {
        "ref": "WR-006",
        "name": "Customer Agent",
        "category": "AI Agents",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """CUSTOMER AGENT - VALIDATION STEPS

WHAT THIS FEATURE DOES:
AI helps manage customer leads by importing from email and suggesting actions.

NAVIGATION PATH:
Left Menu > Sales > Customers > Leads tab

STEPS:
1. Click "Sales" in the left menu
2. Click "Customers"
3. Wait for Customers page to load
4. Look for "Leads" tab at the top
5. Click on "Leads" tab
6. Look for email import options:
   - "Import from Gmail"
   - "Import from Outlook"
   - "Connect email" button
7. Take screenshot showing Leads features

ALTERNATIVE:
1. Look for "Import" button on Customers page
2. Check for lead management options

SCREENSHOT CHECKLIST:
[ ] Customers page visible
[ ] Leads tab visible (if exists)
[ ] Email import options visible

PASS CRITERIA:
- Leads tab visible = PASS
- Email import options = PASS""",
        "alt_check_1": "Check + Create menu for 'Lead' option",
        "alt_check_2": "Look for lead pipeline view",
        "alt_check_3": "SQL: SELECT COUNT(*) FROM customers WHERE lead_status IS NOT NULL",
        "evidence_notes_template": "Sales > Customers shows Leads tab with email import options.",
        "client_note_template": "Customer Agent allows importing leads from Gmail/Outlook.",
    },
    {
        "ref": "WR-007",
        "name": "Intuit Intelligence (Omni)",
        "category": "AI Agents",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """INTUIT INTELLIGENCE - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Conversational AI assistant (chatbot) that answers questions about your business.

NAVIGATION PATH:
Look for chat bubble icon (bottom right of screen)

STEPS:
1. Look at BOTTOM RIGHT corner of any QBO page
2. Find a chat bubble icon or "Intuit Assist" button
3. May also be a sparkle icon
4. Click the icon to open chat window
5. If chat opens, type: "What was my revenue last month?"
6. Wait for AI response
7. Take screenshot of chat window

NOTE: This feature may be in beta. Not visible in all environments.

ALTERNATIVE LOCATIONS:
- Top navigation bar
- Dashboard widget
- Help menu

SCREENSHOT CHECKLIST:
[ ] Intuit Assist icon visible OR
[ ] Chat window open
[ ] AI response showing (if tested)

PASS CRITERIA:
- Intuit Assist icon visible = PASS
- Chat window opens = PASS
- AI responds = PASS""",
        "alt_check_1": "Check Help menu for 'Ask Intuit Assist'",
        "alt_check_2": "Look for AI icon in search bar",
        "alt_check_3": "Check Settings for Intuit Assist toggle",
        "evidence_notes_template": "Intuit Assist [icon visible / chat working]. Beta status: [yes/no].",
        "client_note_template": "Intuit Intelligence provides conversational AI for business queries.",
    },
    {
        "ref": "WR-008",
        "name": "Conversational BI",
        "category": "AI Agents",
        "priority": "P3",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """CONVERSATIONAL BI - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Ask questions in natural language, get reports and charts as answers.

NAVIGATION PATH:
Intuit Assist chat > Ask data question

STEPS:
1. Open Intuit Assist (see WR-007 steps)
2. In the chat, type a data question:
   - "Show me revenue by month"
   - "What are my top 5 customers?"
   - "Compare expenses this year vs last"
3. Press Enter to send
4. Wait for response (10-20 seconds)
5. Look for CHARTS or TABLES in response
6. Take screenshot of question + visual response

WHAT TO LOOK FOR:
- Chart generated from question = BI working
- Table with data = BI working
- Text-only response = Partial (no visualization)

SCREENSHOT CHECKLIST:
[ ] Question visible in chat
[ ] Visual response (chart/table) visible
[ ] Data appears accurate

PASS CRITERIA:
- Visual response generated = PASS
- Data chart showing = PASS""",
        "alt_check_1": "Try different question types (revenue, expenses, customers)",
        "alt_check_2": "Check if results can be saved as report",
        "alt_check_3": "Test date range questions",
        "evidence_notes_template": "Conversational BI query: '[question]' returned [chart type].",
        "client_note_template": "Conversational BI allows natural language queries with visual reports.",
    },
    {
        "ref": "WR-009",
        "name": "Custom KPIs",
        "category": "Reporting",
        "priority": "P0",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """CUSTOM KPIs - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Create and view Key Performance Indicators with custom formulas.

NAVIGATION PATH:
Left Menu > Reports > KPI Scorecard
OR Direct URL: qbo.intuit.com/app/business-intelligence/kpi-scorecard

STEPS:
1. Click "Reports" in left menu
2. Look for "KPIs" or "KPI Scorecard" in list
3. Click on it (or use direct URL above)
4. Wait for page to load (15-20 seconds)
5. You should see KPI tiles with values
6. Look for "Add KPI" or "Create Custom" button
7. Look for KPI library/gallery
8. Take screenshot of KPI Scorecard

WHAT TO LOOK FOR:
- KPI tiles showing values (Revenue, Margin, etc.)
- "Add KPI" button = can customize
- KPI library = multiple options available

SCREENSHOT CHECKLIST:
[ ] KPI Scorecard page visible
[ ] Multiple KPIs with values
[ ] Add/Create button visible

PASS CRITERIA:
- KPI Scorecard loads = PASS
- KPIs with values = PASS
- Custom option available = PASS""",
        "alt_check_1": "Check Dashboard for KPI widgets",
        "alt_check_2": "SQL: SELECT COUNT(*) FROM invoices (data for KPIs)",
        "alt_check_3": "Look for KPIs in Dashboards section",
        "evidence_notes_template": "Reports > KPI Scorecard shows [X] KPIs. Custom option [available].",
        "client_note_template": "Custom KPIs with [X] pre-built metrics and custom formula capability.",
    },
    {
        "ref": "WR-010",
        "name": "Dashboards",
        "category": "Reporting",
        "priority": "P0",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """DASHBOARDS - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Visual dashboards with charts and graphs for business analysis.

NAVIGATION PATH:
Left Menu > Reports > Dashboards
OR Direct URL: qbo.intuit.com/app/reportbuilder

STEPS:
1. Click "Reports" in left menu
2. Look for "Dashboards" option
3. Click on it
4. You should see dashboard templates/gallery
5. Look for options like:
   - Financial Dashboard
   - Sales Dashboard
   - Cash Flow Dashboard
6. Click on any dashboard to open
7. Wait for charts to load
8. Take screenshot of dashboard gallery OR open dashboard

WHAT TO LOOK FOR:
- Dashboard gallery with templates
- Charts with data loading
- Interactive filters
- Drill-down options

SCREENSHOT CHECKLIST:
[ ] Dashboard gallery visible OR
[ ] Open dashboard with charts
[ ] Data populating charts

PASS CRITERIA:
- Dashboard gallery exists = PASS
- Charts showing data = PASS""",
        "alt_check_1": "Try direct URL: qbo.intuit.com/app/reportbuilder",
        "alt_check_2": "Check for 'Custom Reports' section",
        "alt_check_3": "Look for dashboard sharing options",
        "evidence_notes_template": "Reports > Dashboards shows [X] templates. Charts [loading/working].",
        "client_note_template": "Dashboards provide visual analytics with pre-built templates.",
    },
    {
        "ref": "WR-011",
        "name": "3P Data Integrations",
        "category": "Reporting",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """3P DATA INTEGRATIONS - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Connect third-party apps (Salesforce, HubSpot) to bring external data into reports.

NAVIGATION PATH:
Left Menu > Apps

STEPS:
1. Click "Apps" in left menu
2. Wait for Apps page to load
3. Look for "My Apps" section (connected apps)
4. Look for "Find apps" or marketplace section
5. Search for integrations like:
   - Salesforce
   - HubSpot
   - Monday.com
   - Shopify
6. Check for "data" or "reporting" categories
7. Take screenshot of Apps page

WHAT TO LOOK FOR:
- Connected apps in "My Apps"
- Integration marketplace
- Data sync options

SCREENSHOT CHECKLIST:
[ ] Apps page visible
[ ] Integration options visible
[ ] My Apps section visible

PASS CRITERIA:
- Apps page loads = PASS
- Integration options available = PASS""",
        "alt_check_1": "Check Settings > Connected apps",
        "alt_check_2": "Look for 'Data sync' in Settings",
        "alt_check_3": "Search Apps marketplace for 'CRM'",
        "evidence_notes_template": "Apps page shows [X] integration options. Connected: [list].",
        "client_note_template": "Third-party integrations allow connecting CRM and business apps.",
    },
    {
        "ref": "WR-012",
        "name": "Calculated Fields",
        "category": "Reporting",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """CALCULATED FIELDS - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Add custom calculations (formulas) to reports.

NAVIGATION PATH:
Reports > [Any Report] > Customize

STEPS:
1. Click "Reports" in left menu
2. Open any report (Profit and Loss recommended)
3. Click "Customize" button (top right)
4. Customization panel opens
5. Look for:
   - "Add calculated field"
   - "Custom column"
   - "Formula" option
6. May be under "Columns" or "Fields" section
7. Take screenshot of customization panel

WHAT TO LOOK FOR:
- Calculated field option in customize panel
- Formula builder/editor
- Custom column option

SCREENSHOT CHECKLIST:
[ ] Report open
[ ] Customize panel visible
[ ] Calculated field option visible

PASS CRITERIA:
- Calculated field option exists = PASS
- Can add formula = PASS""",
        "alt_check_1": "Try multiple report types for the option",
        "alt_check_2": "Check 'Columns' section in customize",
        "alt_check_3": "Look for 'Add column' button",
        "evidence_notes_template": "Report > Customize shows calculated field option. Formula builder [visible].",
        "client_note_template": "Calculated Fields allows custom formulas in any report.",
    },
    {
        "ref": "WR-013",
        "name": "Management Reports",
        "category": "Reporting",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """MANAGEMENT REPORTS - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Combine multiple financial statements into professional reports for management.

NAVIGATION PATH:
Left Menu > Reports > Management Reports
OR Direct URL: qbo.intuit.com/app/managementreports

STEPS:
1. Click "Reports" in left menu
2. Look for "Management Reports" in the list
3. Click on it
4. Wait for page to load
5. Look for options to:
   - Combine reports
   - Add cover page
   - Add text/notes
   - Customize branding
6. Check for templates
7. Take screenshot

WHAT TO LOOK FOR:
- Management Reports page
- Templates available
- Combined report options
- Cover page customization

SCREENSHOT CHECKLIST:
[ ] Management Reports page visible
[ ] Creation options visible
[ ] Templates available

PASS CRITERIA:
- Page loads = PASS
- Can create combined reports = PASS""",
        "alt_check_1": "Try direct URL: qbo.intuit.com/app/managementreports",
        "alt_check_2": "Check for 'Package' or 'Report Package' option",
        "alt_check_3": "Look in Reports favorites section",
        "evidence_notes_template": "Reports > Management Reports shows [X] templates. Features: [list].",
        "client_note_template": "Management Reports combines financial statements for board presentations.",
    },
    {
        "ref": "WR-014",
        "name": "Benchmarking",
        "category": "Reporting",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """BENCHMARKING - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Compare your business performance against industry benchmarks.

NAVIGATION PATH:
Dashboard OR Reports > Benchmarking

STEPS:
1. Check Dashboard for industry comparison widgets
2. Go to Reports
3. Look for "Benchmarking" or "Industry Comparison"
4. Check KPI Scorecard for benchmark data
5. Look for "vs industry" or "% compared to"
6. Check Performance Center if available

LOCATIONS TO CHECK:
- Dashboard widgets
- KPI Scorecard
- Reports section
- Performance Center

WHAT TO LOOK FOR:
- Industry comparison percentages
- "vs industry" labels
- Benchmark indicators (up/down arrows)

SCREENSHOT CHECKLIST:
[ ] Location where benchmarking found
[ ] Industry comparison visible
[ ] OR note where you searched

PASS CRITERIA:
- Industry comparison visible = PASS
- Benchmark data showing = PASS
- Not found = Document search locations""",
        "alt_check_1": "Check Dashboard for 'How am I doing' widget",
        "alt_check_2": "Look in KPI Scorecard for industry %",
        "alt_check_3": "Search Reports for 'benchmark'",
        "evidence_notes_template": "Benchmarking [found at X / searched: Dashboard, Reports, KPIs].",
        "client_note_template": "Benchmarking compares metrics against industry standards.",
    },
    {
        "ref": "WR-015",
        "name": "Multi-Entity Reports",
        "category": "Reporting",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """MULTI-ENTITY REPORTS - VALIDATION STEPS

WHAT THIS FEATURE DOES:
View consolidated reports across multiple companies in a group.

IMPORTANT: Must switch to CONSOLIDATED VIEW first!

NAVIGATION PATH:
Company Selector (top left) > Consolidated View > Reports

STEPS:
1. Look at TOP LEFT corner (company name)
2. Click on company name
3. Dropdown shows all companies
4. Look for "Consolidated" or "Parent" view
5. Click to switch to Consolidated
6. WAIT for view to change (10-15 seconds)
7. Verify header shows "Consolidated"
8. Go to Reports > Profit and Loss
9. Report should show combined data from ALL companies
10. Take screenshot showing header + report

WHAT TO LOOK FOR:
- Header confirms Consolidated view
- Report shows combined totals
- "All entities" or multi-company data

SCREENSHOT CHECKLIST:
[ ] Header shows Consolidated view
[ ] Report open with data
[ ] Combined entity data visible

PASS CRITERIA:
- Can switch to Consolidated = PASS
- Report shows multi-entity data = PASS""",
        "alt_check_1": "Check company selector for 'All companies' option",
        "alt_check_2": "Look for entity column in reports",
        "alt_check_3": "SQL: SELECT company_type, COUNT(*) FROM invoices GROUP BY company_type",
        "evidence_notes_template": "[Report] in CONSOLIDATED VIEW. Shows data from [X] entities.",
        "client_note_template": "Multi-Entity Reports provide consolidated financial reporting.",
    },
    {
        "ref": "WR-016",
        "name": "Dimension Assignment v2",
        "category": "Dimensions",
        "priority": "P0",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """DIMENSION ASSIGNMENT v2 - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Assign dimension values to transactions with AI suggestions (sparkle icons).

NAVIGATION PATH:
Direct URL: qbo.intuit.com/app/dimensions/assignment
OR Settings (gear) > Dimensions

STEPS:
1. Try direct URL above
2. OR click Gear icon > look for "Dimensions"
3. Look for "Assignment" tab/section
4. Wait for page to load
5. You should see a table of items needing dimensions
6. Look for SPARKLE ICONS = AI suggestions
7. Look for "Suggested" labels
8. Take screenshot showing assignment table

WHAT TO LOOK FOR:
- Assignment table with products/services
- Sparkle icons on suggestions
- Dimension dropdown options

SCREENSHOT CHECKLIST:
[ ] Dimension Assignment page visible
[ ] Items listed in table
[ ] Sparkle icons visible (if present)

PASS CRITERIA:
- Assignment page loads = PASS
- AI sparkles visible = BONUS""",
        "alt_check_1": "Check Settings > Account and Settings > Advanced > Dimensions",
        "alt_check_2": "Look for dimension fields on invoices",
        "alt_check_3": "SQL: SELECT COUNT(*) FROM classifications",
        "evidence_notes_template": "Dimensions > Assignment shows [X] items. AI sparkles: [yes/no].",
        "client_note_template": "Dimension Assignment v2 provides AI-assisted categorization.",
    },
    {
        "ref": "WR-017",
        "name": "Hierarchical Dimension Reporting",
        "category": "Dimensions",
        "priority": "P1",
        "validation_type": "DATA+UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """HIERARCHICAL DIMENSION REPORTING - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Report on dimensions with parent-child hierarchy (Region > City > Store).

OPTION A - DATABASE EVIDENCE (Recommended):

Run SQL on QBO database:
SELECT id, name, parent_id FROM classifications WHERE parent_id IS NOT NULL LIMIT 20;

If rows returned = hierarchy exists. Screenshot the query result.


OPTION B - UI VALIDATION:

NAVIGATION PATH:
Reports > [Any Report] > Customize > Filter by Dimension

STEPS:
1. Go to Reports
2. Open Profit and Loss or similar
3. Click "Customize"
4. Look for dimension filter options
5. When selecting dimension values, look for:
   - Indented items (children under parents)
   - Expandable sections
   - Roll-up totals
6. Take screenshot showing hierarchy

WHAT TO LOOK FOR:
- Parent > Child structure in selection
- Indented dimension values
- Roll-up/drill-down capability

SCREENSHOT CHECKLIST:
[ ] Database query results OR
[ ] Report with hierarchical dimension filter
[ ] Parent-child relationship visible

PASS CRITERIA:
- Database shows parent_id relationships = PASS
- UI shows hierarchical selection = PASS""",
        "alt_check_1": "SQL: SELECT COUNT(*) FROM classifications WHERE parent_id IS NOT NULL",
        "alt_check_2": "Check dimension setup in Settings",
        "alt_check_3": "Look for 'Sub-class' options",
        "evidence_notes_template": "Database: [X] dimensions with parent_id. UI: hierarchy [visible/configured].",
        "client_note_template": "Hierarchical Dimensions allow parent-child organization for roll-up reporting.",
    },
    {
        "ref": "WR-018",
        "name": "Dimensions on Workflow",
        "category": "Dimensions",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """DIMENSIONS ON WORKFLOW - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Use dimensions as conditions in workflow automation rules.

NOTE: May return 404 in Construction environments - that's expected.

NAVIGATION PATH:
Settings (gear) > Workflow automation
OR Direct URL: qbo.intuit.com/app/workflowautomation

STEPS:
1. Click Gear icon (top right)
2. Look for "Workflow automation" or "Workflows"
3. Click on it
4. IF PAGE LOADS:
   - Look at existing rules
   - Click "Create rule" or edit a rule
   - In rule conditions, look for "Dimension" option
   - Check if you can filter by dimension values
5. IF 404 ERROR:
   - Screenshot the 404 page
   - Mark as "NOT AVAILABLE" (expected in some environments)
6. Take screenshot

WHAT TO LOOK FOR:
- Workflow automation page
- Dimension as rule condition option
- OR 404 error (document it)

SCREENSHOT CHECKLIST:
[ ] Workflow page visible OR 404 error
[ ] Dimension option in rules (if available)

PASS CRITERIA:
- Dimension in rule conditions = PASS
- 404 error = Mark NOT AVAILABLE""",
        "alt_check_1": "Try Settings > Manage workflows",
        "alt_check_2": "Check for 'Automation' in left menu",
        "alt_check_3": "Look for rule conditions mentioning Class/Location",
        "evidence_notes_template": "[Workflow shows dimension option] OR [NOT AVAILABLE: 404 in this environment].",
        "client_note_template": "Workflow automation can use dimensions as rule conditions.",
    },
    {
        "ref": "WR-019",
        "name": "Dimensions on Balance Sheet",
        "category": "Dimensions",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """DIMENSIONS ON BALANCE SHEET - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Filter Balance Sheet reports by dimension values (Class, Location, etc).

NAVIGATION PATH:
Reports > Balance Sheet > Customize > Filter

STEPS:
1. Click "Reports" in left menu
2. Find "Balance Sheet" report
3. Click to open it
4. Wait for report to generate
5. Click "Customize" button (top right)
6. Look for "Filter" section
7. In filters, look for:
   - "Class" (type of dimension)
   - "Location"
   - "Department"
   - Any custom dimensions
8. Click on dimension dropdown to see values
9. Select a value and run report
10. Take screenshot of customization panel

WHAT TO LOOK FOR:
- Dimension filter options in customize
- Dropdown with dimension values
- Report updates when filtered

SCREENSHOT CHECKLIST:
[ ] Balance Sheet open
[ ] Customize panel visible
[ ] Dimension filter options shown

PASS CRITERIA:
- Dimension filters available = PASS
- Can filter report = PASS""",
        "alt_check_1": "Check P&L report for same dimension filters",
        "alt_check_2": "Look for 'Segment' or 'Location' columns",
        "alt_check_3": "Try 'Balance Sheet by Class' report",
        "evidence_notes_template": "Balance Sheet > Customize shows dimension filters: [list dimensions].",
        "client_note_template": "Balance Sheet can be filtered by dimensions for segment analysis.",
    },
    {
        "ref": "WR-020",
        "name": "Parallel Approval",
        "category": "Workflow",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """PARALLEL APPROVAL - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Allow multiple approvers to review simultaneously (not sequentially).

NOTE: May return 404 in Construction environments.

NAVIGATION PATH:
Settings (gear) > Workflow automation
OR Direct URL: qbo.intuit.com/app/workflowautomation

STEPS:
1. Click Gear icon (top right)
2. Click "Workflow automation"
3. IF PAGE LOADS:
   - Look for approval workflow rules
   - Create or edit an approval rule
   - Look for MULTIPLE approver option
   - Check for "Parallel" or "Any approver" choice:
     * "All must approve" = sequential
     * "Any can approve" = parallel
4. IF 404 ERROR:
   - Screenshot the 404
   - Mark as NOT AVAILABLE
5. Take screenshot

WHAT TO LOOK FOR:
- Multiple approver fields
- "Any can approve" option
- Parallel approval configuration

SCREENSHOT CHECKLIST:
[ ] Workflow page visible OR 404
[ ] Multiple approvers option (if available)
[ ] Parallel/Any option visible

PASS CRITERIA:
- Parallel approval option = PASS
- 404 = NOT AVAILABLE""",
        "alt_check_1": "Look for 'Approval rules' in Settings",
        "alt_check_2": "Check for 'Approval required' on bills/expenses",
        "alt_check_3": "Look for user role 'Approver'",
        "evidence_notes_template": "[Workflow shows parallel approval] OR [NOT AVAILABLE: 404].",
        "client_note_template": "Parallel Approval allows simultaneous review by multiple approvers.",
    },
    {
        "ref": "WR-021",
        "name": "Desktop Migration",
        "category": "Migration",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": False,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """DESKTOP MIGRATION - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Migrate data from QuickBooks Desktop to QuickBooks Online.

NOTE: Full demo requires FRESH tenant. Can only verify option exists.

NAVIGATION PATH:
Settings (gear) > Import data

STEPS:
1. Click Gear icon (top right)
2. Look for "Import data" or "Import"
3. Click on it
4. Look for Desktop migration option:
   - "Import from QuickBooks Desktop"
   - "Desktop migration"
   - "Convert from Desktop"
5. May also be under "Tools" menu
6. Note what file types supported (.qbw, .qbb)
7. Take screenshot of import options

WHAT TO LOOK FOR:
- Desktop import option
- Supported file types
- Migration workflow

SCREENSHOT CHECKLIST:
[ ] Import page visible
[ ] Desktop migration option visible

PASS CRITERIA:
- Desktop migration option exists = PASS

NOTE: Full migration requires fresh tenant.""",
        "alt_check_1": "Check Settings > Tools > Import",
        "alt_check_2": "Look for 'Migration' in Help menu",
        "alt_check_3": "Search Settings for 'Desktop'",
        "evidence_notes_template": "Settings > Import shows Desktop Migration option. [Visible/requires fresh tenant].",
        "client_note_template": "Desktop Migration allows importing from QuickBooks Desktop.",
    },
    {
        "ref": "WR-022",
        "name": "DFY Migration",
        "category": "Migration",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": False,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """DFY (DONE-FOR-YOU) MIGRATION - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Professional migration service - Intuit handles the data migration for you.

NOTE: Full demo requires FRESH tenant.

NAVIGATION PATH:
Settings (gear) > Import data

STEPS:
1. Click Gear icon (top right)
2. Look for "Import data"
3. Click on it
4. Look for DFY option:
   - "Done for you" migration
   - "Professional migration"
   - "We'll do it for you"
5. May be presented as premium option
6. Note service details if visible
7. Take screenshot

WHAT TO LOOK FOR:
- DFY/Professional migration option
- Service description
- Pricing information (if shown)

SCREENSHOT CHECKLIST:
[ ] Import options visible
[ ] DFY option visible (if exists)

PASS CRITERIA:
- DFY option visible = PASS
- Only self-service = Note subscription level""",
        "alt_check_1": "Check for 'Professional services' link",
        "alt_check_2": "Look in Help for migration services",
        "alt_check_3": "Contact Intuit for DFY availability",
        "evidence_notes_template": "Settings > Import shows DFY Migration option. [Visible/subscription dependent].",
        "client_note_template": "Done-For-You Migration provides professional migration services.",
    },
    {
        "ref": "WR-023",
        "name": "Feature Compatibility",
        "category": "Migration",
        "priority": "P3",
        "validation_type": "N/A",
        "tco_applicable": False,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """FEATURE COMPATIBILITY - VALIDATION STEPS

WHAT THIS IS:
Documentation showing which Desktop features are compatible with QBO.

THIS IS DOCUMENTATION - NO UI VALIDATION NEEDED

VALIDATION METHOD:
- Reference Intuit's official compatibility documentation
- No screenshot required
- Mark as N/A for UI validation

EVIDENCE:
Note: "Feature Compatibility is documentation only - refer to Intuit migration guide"

EXPECTED RESULT:
N/A - Documentation feature""",
        "alt_check_1": "N/A - Documentation only",
        "alt_check_2": "Refer to Intuit migration docs",
        "alt_check_3": "N/A",
        "evidence_notes_template": "N/A - Documentation feature. Refer to Intuit migration compatibility guide.",
        "client_note_template": "Feature Compatibility documentation available through Intuit resources.",
    },
    {
        "ref": "WR-024",
        "name": "Certified Payroll Report",
        "category": "Construction",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": False,
        "construction_sales_applicable": True,
        "construction_events_applicable": True,
        "step_by_step": """CERTIFIED PAYROLL REPORT - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Generate WH-347 certified payroll reports for government construction projects.

NOTE: CONSTRUCTION-SPECIFIC feature. Validate in Construction environment.

NAVIGATION PATH:
Reports > Payroll > Certified Payroll
OR Payroll > Reports > Certified Payroll

STEPS:
1. Go to Reports
2. Look in "Payroll" category
3. OR search for "Certified Payroll"
4. OR check Payroll > Reports
5. Find "Certified Payroll Report" or "WH-347"
6. Click to open
7. Select date range and project (if required)
8. Generate the report
9. Verify WH-347 format
10. Take screenshot

WHAT TO LOOK FOR:
- WH-347 format
- Employee wages, hours, deductions
- Government compliance fields
- Project association

SCREENSHOT CHECKLIST:
[ ] Certified Payroll Report visible
[ ] WH-347 header/format visible
[ ] Employee data showing

PASS CRITERIA:
- WH-347 report available = PASS
- Government format correct = PASS""",
        "alt_check_1": "Check Payroll > Tax forms for WH-347",
        "alt_check_2": "Look for 'Prevailing wage' reports",
        "alt_check_3": "SQL: SELECT company_type, COUNT(*) FROM employees GROUP BY company_type",
        "evidence_notes_template": "Certified Payroll Report (WH-347) in Construction environment. Format: [compliant].",
        "client_note_template": "Certified Payroll generates WH-347 reports for government projects.",
    },
    {
        "ref": "WR-025",
        "name": "Sales Order",
        "category": "Construction",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": False,
        "construction_sales_applicable": True,
        "construction_events_applicable": True,
        "step_by_step": """SALES ORDER - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Create sales orders before invoicing (order-to-invoice workflow).

NAVIGATION PATH:
+ (Create) button > Sales Order
OR Sales menu > Sales Orders

STEPS:
1. Click "+" (Create) button
2. Look for "Sales Order" option
3. OR check Sales menu > Sales Orders
4. OR check Settings > Sales settings
5. May need to enable in Settings:
   - Gear > Account and Settings > Sales
   - Look for "Sales Order" toggle
6. If found, click to view Sales Order form
7. Note available fields
8. Take screenshot

WHAT TO LOOK FOR:
- Sales Order in Create menu
- Sales Order creation form
- Order-to-invoice workflow

SCREENSHOT CHECKLIST:
[ ] Sales Order option visible in menu
[ ] OR Sales Order form visible
[ ] OR Settings showing Sales Order option

PASS CRITERIA:
- Sales Order available = PASS
- May need enabling in Settings""",
        "alt_check_1": "Check Settings > Sales for order settings",
        "alt_check_2": "Look for 'Order' in + Create menu",
        "alt_check_3": "Check Sales transactions list for orders",
        "evidence_notes_template": "Sales Order feature [in Create menu / in Settings / form visible].",
        "client_note_template": "Sales Orders support order-to-cash workflow before invoicing.",
    },
    {
        "ref": "WR-026",
        "name": "Multi-Entity Payroll Hub",
        "category": "Payroll",
        "priority": "P1",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """MULTI-ENTITY PAYROLL HUB - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Centralized employee management across multiple companies.

IMPORTANT: Must be in CONSOLIDATED VIEW!

NAVIGATION PATH:
Company Selector > Consolidated View > Payroll > Employees

STEPS:
1. Look at TOP LEFT (company name)
2. Click on company name
3. Select "Consolidated" or "Parent" view
4. WAIT for view to change
5. Verify header shows Consolidated
6. Click "Payroll" in left menu
7. Click "Employees" or "Employee Hub"
8. You should see employees from ALL companies
9. Look for entity/company column
10. Take screenshot

WHAT TO LOOK FOR:
- Header confirms Consolidated view
- Employee list from multiple companies
- Entity column showing company names

SCREENSHOT CHECKLIST:
[ ] Header shows Consolidated
[ ] Payroll > Employees visible
[ ] Multiple entity employees visible

PASS CRITERIA:
- Multi-entity employee list = PASS
- Entity column visible = PASS""",
        "alt_check_1": "Check for 'All entities' filter on employees",
        "alt_check_2": "Look for entity dropdown in Payroll",
        "alt_check_3": "SQL: SELECT company_type, COUNT(*) FROM employees GROUP BY company_type",
        "evidence_notes_template": "Payroll Hub in CONSOLIDATED VIEW shows employees from [X] entities.",
        "client_note_template": "Multi-Entity Payroll Hub provides centralized employee management.",
    },
    {
        "ref": "WR-027",
        "name": "Garnishments",
        "category": "Payroll",
        "priority": "P2",
        "validation_type": "UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": False,
        "step_by_step": """GARNISHMENTS - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Configure automatic wage garnishments (child support, tax levies) for employees.

NAVIGATION PATH:
Payroll > Employees > [Employee] > Deductions/Garnishments

STEPS:
1. Click "Payroll" in left menu
2. Click "Employees"
3. Wait for list to load
4. Click on any employee name
5. Wait for profile to load
6. Scroll through profile sections
7. Look for:
   - "Deductions" section
   - "Garnishments" section
   - "Withholdings" section
8. May be under "Pay" or "Taxes" tab
9. Check for garnishment types:
   - Child support
   - Tax levies
   - Creditor garnishments
10. Take screenshot

WHAT TO LOOK FOR:
- Garnishment section in profile
- Garnishment type options
- Configuration fields

SCREENSHOT CHECKLIST:
[ ] Employee profile open
[ ] Garnishments section visible
[ ] Options available

PASS CRITERIA:
- Garnishment section exists = PASS
- Can configure types = PASS""",
        "alt_check_1": "Check employee 'Additional pay/deductions'",
        "alt_check_2": "Look for 'Court-ordered' deductions",
        "alt_check_3": "Check Payroll Settings for garnishment setup",
        "evidence_notes_template": "Payroll > Employees > [Name] shows Garnishments section. Types: [list].",
        "client_note_template": "Garnishments feature allows configuring wage garnishments.",
    },
    {
        "ref": "WR-028",
        "name": "Assignments in QBTime",
        "category": "Payroll",
        "priority": "P2",
        "validation_type": "DATA+UI",
        "tco_applicable": True,
        "construction_sales_applicable": True,
        "construction_events_applicable": True,
        "step_by_step": """ASSIGNMENTS IN QBTIME - VALIDATION STEPS

WHAT THIS FEATURE DOES:
Track time entries with project/job assignments for labor costing.

OPTION A - DATABASE EVIDENCE (Quick):

Run SQL:
SELECT COUNT(*) as total_time_entries FROM time_entries;

Large number (e.g., 62,997) = time tracking active.


OPTION B - UI VALIDATION:

NAVIGATION PATH:
Left Menu > Time OR Payroll > Time

STEPS:
1. Click "Time" in left menu
2. OR Payroll > Time entries
3. Wait for time entries page
4. Look at the time entry list
5. Check for columns:
   - Employee
   - Customer/Project
   - Service item
   - Hours
6. Look for "Assignment" or "Project" column
7. Check for job costing info
8. Take screenshot

WHAT TO LOOK FOR:
- Time entries with project assignments
- Job/customer association
- Hours by project

SCREENSHOT CHECKLIST:
[ ] Database query result OR
[ ] Time entries page visible
[ ] Assignments/projects visible

PASS CRITERIA:
- Database shows time entries = PASS
- UI shows assignments = PASS""",
        "alt_check_1": "SQL: SELECT COUNT(*) FROM time_entries",
        "alt_check_2": "Check for 'Timesheet' in left menu",
        "alt_check_3": "Look for time by project reports",
        "evidence_notes_template": "Database: [X] time entries. UI: Assignments to projects [visible].",
        "client_note_template": "QBTime Assignments tracks time by project for labor costing.",
    },
    {
        "ref": "WR-029",
        "name": "Enhanced Amendments (CA)",
        "category": "Payroll",
        "priority": "P3",
        "validation_type": "N/A",
        "tco_applicable": False,
        "construction_sales_applicable": False,
        "construction_events_applicable": False,
        "step_by_step": """ENHANCED AMENDMENTS (CA) - VALIDATION STEPS

WHAT THIS FEATURE DOES:
California-specific payroll amendment handling.

IMPORTANT: Requires CALIFORNIA tenant!

VALIDATION STATUS: NOT APPLICABLE
Current environments are not California-specific.

IF YOU HAVE CA TENANT:
1. Login to California tenant
2. Go to Payroll > Taxes
3. Look for "Amendments" or "Corrections"
4. Check for CA-specific workflow
5. Screenshot amendment options

FOR CURRENT ENVIRONMENTS:
Mark as N/A - Requires California tenant

EXPECTED RESULT:
N/A - CA tenant required""",
        "alt_check_1": "N/A - Requires CA tenant",
        "alt_check_2": "Check if tenant is configured for California",
        "alt_check_3": "Contact team for CA environment access",
        "evidence_notes_template": "N/A - Requires California-specific tenant.",
        "client_note_template": "Enhanced Amendments (CA) is California-specific. Requires CA tenant.",
    },
]

# Complete login information with credentials and all CIDs
LOGINS = [
    {
        "environment": "TCO",
        "email": "quickbooks-tco-tbxdemo@tbxofficial.com",
        "password": "TestBox!23",
        "totp_secret": "RP4MFT45ZY5EYGMRIPEANJA6YFKHV5O7",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341455130122367",
                "name": "Traction Control Outfitters (Parent)",
                "type": "Parent",
                "priority": "P0",
                "dataset": "TCO",
            },
            {
                "cid": "9341455130196737",
                "name": "Global Tread Distributors",
                "type": "Child",
                "priority": "P0",
                "dataset": "TCO",
            },
            {
                "cid": "9341455130166501",
                "name": "Apex Tire & Auto Retail",
                "type": "Child",
                "priority": "P0",
                "dataset": "TCO",
            },
            {
                "cid": "9341455130182073",
                "name": "RoadReady Service Solutions",
                "type": "Child",
                "priority": "P0",
                "dataset": "TCO",
            },
        ],
        "notes": "TCO multi-entity demo. Use Parent for consolidated, Children for individual company features.",
    },
    {
        "environment": "CONSTRUCTION SALES",
        "email": "quickbooks-test-account@tbxofficial.com",
        "password": "TestBox123!",
        "totp_secret": "23CIWY3QYCOXJRKZYG6YKR7HUYVPLPEL",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341454156620895",
                "name": "Keystone Construction (Par.) - PRIMARY SALES",
                "type": "Parent",
                "priority": "P0",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156620204",
                "name": "Keystone Terra (Ch.) - PRIMARY SALES",
                "type": "Child",
                "priority": "P0",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156643813",
                "name": "Keystone Ironcraft - PRIMARY SALES",
                "type": "Child",
                "priority": "P1",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156621045",
                "name": "Keystone BlueCraft (Ch.)",
                "type": "Child",
                "priority": "P0",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156640324",
                "name": "KeyStone Stonecraft",
                "type": "Child",
                "priority": "P1",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156634620",
                "name": "KeyStone Canopy",
                "type": "Child",
                "priority": "P1",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156644492",
                "name": "KeyStone Ecocraft",
                "type": "Child",
                "priority": "P1",
                "dataset": "Construction",
            },
            {
                "cid": "9341454156629550",
                "name": "KeyStone Volt",
                "type": "Child",
                "priority": "P1",
                "dataset": "Construction",
            },
        ],
        "notes": "Construction Sales demo. Use Parent for consolidated, PRIMARY marked companies for main demos.",
    },
    {
        "environment": "CONSTRUCTION EVENTS",
        "email": "quickbooks-test-account-qsp@tbxofficial.com",
        "password": "TestBox123!",
        "totp_secret": "23CIWY3QYCOXJRKZYG6YKR7HUYVPLPEL",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341454796804202",
                "name": "Keystone Construction (Event) - PRIMARY EVENTS",
                "type": "Parent",
                "priority": "P0",
                "dataset": "Construction",
            },
            {
                "cid": "9341454842738078",
                "name": "Keystone Terra (Event) - PRIMARY EVENTS",
                "type": "Child",
                "priority": "P0",
                "dataset": "Construction",
            },
            {
                "cid": "9341454842756096",
                "name": "Keystone BlueCraft (Event) - PRIMARY EVENTS",
                "type": "Child",
                "priority": "P0",
                "dataset": "Construction",
            },
        ],
        "notes": "Construction Events demo environment. Use for event demos only.",
    },
    {
        "environment": "CONSTRUCTION BI",
        "email": "quickbooks-tbx-product-team-test@tbxofficial.com",
        "password": "TBD - Check with team",
        "totp_secret": "TBD",
        "url": "https://qbo.intuit.com",
        "companies": [
            {
                "cid": "9341455531293719",
                "name": "Keystone Construction (BI)",
                "type": "Parent",
                "priority": "",
                "dataset": "Construction",
            },
            {
                "cid": "9341455531294375",
                "name": "Keystone Terra (BI)",
                "type": "Child",
                "priority": "",
                "dataset": "Construction",
            },
            {
                "cid": "9341455531274694",
                "name": "Keystone BlueCraft (BI)",
                "type": "Child",
                "priority": "",
                "dataset": "Construction",
            },
        ],
        "notes": "Product team BI testing environment. Confirm credentials with team.",
    },
]


def create_workbook():
    wb = Workbook()

    # ========== SHEET 1: TCO VALIDATION ==========
    create_environment_sheet(wb, "TCO Validation", "TCO", [f for f in FEATURES if f["tco_applicable"]])

    # ========== SHEET 2: CONSTRUCTION SALES VALIDATION ==========
    ws_cs = wb.create_sheet("Construction Sales")
    create_environment_sheet_on_ws(
        ws_cs,
        "CONSTRUCTION SALES",
        [f for f in FEATURES if f["construction_sales_applicable"]],
    )

    # ========== SHEET 3: CONSTRUCTION EVENTS VALIDATION ==========
    ws_ce = wb.create_sheet("Construction Events")
    create_environment_sheet_on_ws(
        ws_ce,
        "CONSTRUCTION EVENTS",
        [f for f in FEATURES if f["construction_events_applicable"]],
    )

    # ========== SHEET 4: STEP BY STEP GUIDE ==========
    create_step_by_step_sheet(wb)

    # ========== SHEET 5: LOGINS ==========
    create_logins_sheet(wb)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    return wb


def create_environment_sheet(wb, sheet_name, env_name, features):
    ws = wb.active
    ws.title = sheet_name
    create_environment_sheet_on_ws(ws, env_name, features)


def create_environment_sheet_on_ws(ws, env_name, features):
    # Headers - Now includes alternative checks from column D onwards
    headers = [
        "Ref",
        "Feature Name",
        "Category",
        "Alt Check 1",
        "Alt Check 2",
        "Alt Check 3",
        "Status",
        "Evidence Link",
        "Client Notes",
        "Validator",
        "Date",
    ]

    # Title row
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"{env_name} - WINTER RELEASE FY26 VALIDATION"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    row = 4
    for feature in features:
        ws.cell(row=row, column=1, value=feature["ref"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=feature["name"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=feature["category"]).border = THIN_BORDER

        # Alternative checks (columns D, E, F)
        alt1 = ws.cell(row=row, column=4, value=feature.get("alt_check_1", ""))
        alt1.border = THIN_BORDER
        alt1.alignment = Alignment(wrap_text=True)

        alt2 = ws.cell(row=row, column=5, value=feature.get("alt_check_2", ""))
        alt2.border = THIN_BORDER
        alt2.alignment = Alignment(wrap_text=True)

        alt3 = ws.cell(row=row, column=6, value=feature.get("alt_check_3", ""))
        alt3.border = THIN_BORDER
        alt3.alignment = Alignment(wrap_text=True)

        # Status dropdown placeholder
        status_cell = ws.cell(row=row, column=7, value="")
        status_cell.border = THIN_BORDER

        # Evidence link placeholder
        ws.cell(row=row, column=8, value="").border = THIN_BORDER

        # Client notes template
        notes_cell = ws.cell(row=row, column=9, value=feature["client_note_template"])
        notes_cell.border = THIN_BORDER
        notes_cell.alignment = Alignment(wrap_text=True)

        # Validator placeholder
        ws.cell(row=row, column=10, value="").border = THIN_BORDER

        # Date placeholder
        ws.cell(row=row, column=11, value="").border = THIN_BORDER

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 45
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 12


def create_step_by_step_sheet(wb):
    ws = wb.create_sheet("Step by Step Guide")

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "DETAILED VALIDATION INSTRUCTIONS (No redundant login steps - see Logins sheet)"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Ref", "Feature", "Step-by-Step Instructions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Data
    row = 4
    for feature in FEATURES:
        ws.cell(row=row, column=1, value=feature["ref"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=feature["name"]).border = THIN_BORDER

        instructions_cell = ws.cell(row=row, column=3, value=feature["step_by_step"])
        instructions_cell.alignment = Alignment(wrap_text=True, vertical="top")
        instructions_cell.border = THIN_BORDER

        # Set row height based on content
        ws.row_dimensions[row].height = 350

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 110


def create_logins_sheet(wb):
    ws = wb.create_sheet("Logins and Access")

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "COMPLETE LOGIN CREDENTIALS AND COMPANY IDs"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    row = 3
    for login in LOGINS:
        # Environment header
        env_cell = ws.cell(row=row, column=1, value=login["environment"])
        env_cell.font = Font(bold=True, size=12)
        env_cell.fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        # Credentials
        ws.cell(row=row, column=1, value="Email:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=login["email"])
        row += 1

        ws.cell(row=row, column=1, value="Password:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=login["password"])
        row += 1

        ws.cell(row=row, column=1, value="TOTP Secret:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=login["totp_secret"])
        row += 1

        ws.cell(row=row, column=1, value="URL:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=login["url"])
        row += 1

        # Companies header
        ws.cell(row=row, column=1, value="").font = Font(bold=True)
        row += 1

        # Companies table header
        company_headers = ["CID", "Company Name", "Type", "Priority", "Dataset"]
        for col, header in enumerate(company_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        row += 1

        # Companies list
        for company in login["companies"]:
            ws.cell(row=row, column=1, value=company["cid"])
            ws.cell(row=row, column=2, value=company["name"])
            ws.cell(row=row, column=3, value=company["type"])
            ws.cell(row=row, column=4, value=company["priority"])
            ws.cell(row=row, column=5, value=company["dataset"])
            row += 1

        # Notes
        row += 1
        ws.cell(row=row, column=1, value="Notes:").font = Font(bold=True)
        notes_cell = ws.cell(row=row, column=2, value=login["notes"])
        notes_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 2

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15


if __name__ == "__main__":
    wb = create_workbook()
    output_path = "C:/Users/adm_r/intuit-boom/docs/WINTER_RELEASE_VALIDATION_MATRIX_v3.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    print("Sheets: TCO Validation, Construction Sales, Construction Events, Step by Step Guide, Logins and Access")
    print("\nImprovements in v3:")
    print("- Complete login credentials (email, password, TOTP)")
    print("- All CIDs from Intuit data")
    print("- Removed redundant LOGIN steps from step-by-step")
    print("- Added Alt Check 1, 2, 3 columns for alternative validation methods")
