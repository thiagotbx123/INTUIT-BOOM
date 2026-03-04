"""
Winter Release FY26 - Feature Check Plan v3
Same sequence as v2, adds Evidence column with exact deliverables per feature
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Feature Check Plan"

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
blocked_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
blocked_font = Font(name="Calibri", size=10, color="9C0006")
na_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
p0_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
normal_font = Font(name="Calibri", size=10)
wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = [
    "Seq",
    "ID",
    "Feature",
    "Priority",
    "Flag\nOn?",
    "Nav Group",
    "Step-by-Step (What User Does)",
    "Evidence to Collect",
    "Status",
    "Notes",
]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = thin

features = [
    # ── 1. REPORTS ──
    {
        "seq": 1,
        "id": 16,
        "name": "KPIs",
        "pri": "P0",
        "flag": "YES",
        "nav": "1. Reports",
        "steps": (
            '1. Click "Reports" in left nav\n'
            '2. Click "KPIs" to open KPI Scorecard\n'
            '3. Click "Manage KPIs" (top right)\n'
            "4. Scroll through KPI Library\n"
            '5. Click "Create KPI"\n'
            "6. In formula builder, open account dropdown > verify CoA accounts appear\n"
            "7. Select time period (current/previous)\n"
            "8. Check if 3P data variables appear as options\n"
            "9. Save custom KPI\n"
            "10. Return to KPI Scorecard > verify custom KPI shows\n"
            "11. Add a KPI that compares against previous year/period (line or bar chart)"
        ),
        "evidence": (
            "PRINT 1: KPI Scorecard page showing KPIs with data\n"
            "PRINT 2: KPI Library (Manage KPIs) showing standard + custom entries\n"
            'PRINT 3: "Create KPI" formula builder with CoA accounts in dropdown\n'
            "PRINT 4: Time period selector (current/previous) visible\n"
            "PRINT 5: 3P data variables listed as options\n"
            "PRINT 6: Custom KPI saved and visible on KPI Scorecard\n"
            "PRINT 7: KPI comparison chart (previous year/period) rendered"
        ),
    },
    {
        "seq": 2,
        "id": 17,
        "name": "Dashboards",
        "pri": "P0",
        "flag": "YES",
        "nav": "1. Reports",
        "steps": (
            '1. From Reports, click "Dashboards"\n'
            "2. Look for 3 NEW pre-built dashboards:\n"
            "   - Sales Performance (CRM)\n"
            "   - Payroll\n"
            "   - Time Tracking\n"
            "3. Click Edit on a dashboard\n"
            '4. In KPI list, look for "Create KPI" CTA button\n'
            "5. Click it > verify builder opens inline\n"
            "6. Edit a chart widget\n"
            '7. Look for "Compare" option\n'
            '8. Select "Previous Period" or "Previous Year"\n'
            "9. Verify comparison renders correctly\n"
            "10. Create/open a Sales Performance Dashboard to tell a story"
        ),
        "evidence": (
            "PRINT 1: Dashboard gallery showing 3 NEW pre-built dashboards highlighted\n"
            'PRINT 2: Dashboard editor with "Create KPI" CTA button visible\n'
            "PRINT 3: Inline KPI builder open from within dashboard editor\n"
            'PRINT 4: Chart widget with "Compare" dropdown showing Previous Period/Year\n'
            "PRINT 5: Chart rendering comparison data (side-by-side or overlay)\n"
            "PRINT 6: Sales Performance Dashboard populated with data telling a story"
        ),
    },
    {
        "seq": 3,
        "id": 18,
        "name": "3P Data (Advanced App Integrations)",
        "pri": "P0",
        "flag": "YES",
        "nav": "1. Reports",
        "steps": (
            "1. Go to Reports > KPIs > Manage KPIs\n"
            '2. Look for KPIs with "Connect to App" option\n'
            "3. Verify CRM connectors: Monday.com, HubSpot, Salesforce\n"
            "4. Verify Workforce connectors: Gusto, Hubstaff, Clockify\n"
            "5. Count 3P KPI library entries:\n"
            "   - CRM: 12 KPIs\n"
            "   - Payroll: 13 KPIs\n"
            "   - Time: 6 KPIs\n"
            "6. Go to Reports > Dashboards\n"
            "7. Check for 3 new pre-built 3P dashboards"
        ),
        "evidence": (
            'PRINT 1: KPI Library showing "Connect to App" CTA on a 3P KPI\n'
            "PRINT 2: CRM connectors list (Monday.com, HubSpot, Salesforce)\n"
            "PRINT 3: Workforce connectors list (Gusto, Hubstaff, Clockify)\n"
            "PRINT 4: 3P KPI Library section showing CRM KPIs (count visible)\n"
            "PRINT 5: 3P KPI Library section showing Payroll KPIs (count visible)\n"
            "PRINT 6: 3P KPI Library section showing Time KPIs (count visible)\n"
            "PRINT 7: Dashboard gallery showing 3P-powered dashboard templates"
        ),
    },
    {
        "seq": 4,
        "id": 20,
        "name": "Management Reports",
        "pri": "P0",
        "flag": "YES",
        "nav": "1. Reports",
        "steps": (
            "1. Go to Reports > Management Reports\n"
            "2. Create/open a single-entity report with:\n"
            "   - Footer configured\n"
            '   - Smart chip comparing period ("this period" / "last period")\n'
            "   - Custom KPI added as chart widget\n"
            "3. Test email sharing: Share > Email as PDF\n"
            "4. Switch to Consolidated View\n"
            "5. Add monthly performance report to Consolidated View\n"
            "6. Show KPI charts and select which companies to include\n"
            "7. Verify non-admin RBAC access (log in as non-admin if possible)"
        ),
        "evidence": (
            "PRINT 1: Management Report with footer visible at bottom\n"
            'PRINT 2: Smart chip showing "this period" / "last period" placeholder\n'
            "PRINT 3: Custom KPI chart widget embedded in report\n"
            "PRINT 4: Email sharing dialog (Share > Email as PDF)\n"
            "PRINT 5: PDF preview of the report (post-email or export)\n"
            "PRINT 6: Consolidated View - Management Report with ME KPIs in chart\n"
            "PRINT 7: Company selector showing which entities are included\n"
            "PRINT 8: Non-admin user view (if accessible)\n"
            "NOTE: AI-summary for KPIs/reports NOT shipping Feb (descoped Jan 9)"
        ),
    },
    {
        "seq": 5,
        "id": 21,
        "name": "New Modern Reports",
        "pri": "P1",
        "flag": "YES",
        "nav": "1. Reports",
        "steps": (
            "1. Go to Reports > Standard Reports\n"
            "2. Open P&L report\n"
            "3. Check new UI: Cash vs Accrual toggle in header\n"
            "4. Check Display & Compare dropdowns\n"
            "5. Look for Customization in dedicated section\n"
            "6. Test page-level scroll (hover left/right of report)\n"
            "7. Set Display columns by month\n"
            "8. Test single-click Pivot functionality\n"
            "9. Open Balance Sheet > drill down into zero balance accounts\n"
            "10. Verify banded rows (from Nov release) still working"
        ),
        "evidence": (
            "PRINT 1: Report header showing Cash vs Accrual toggle\n"
            "PRINT 2: Display & Compare dropdowns expanded\n"
            "PRINT 3: Customization section in dedicated area\n"
            "PRINT 4: Report with monthly columns displayed\n"
            "PRINT 5: Pivot result after single-click\n"
            "PRINT 6: Balance Sheet with zero balance account drill-down open\n"
            "PRINT 7: Report showing banded rows styling"
        ),
    },
    {
        "seq": 6,
        "id": 13,
        "name": "Finance AI",
        "pri": "P1",
        "flag": "YES",
        "nav": "1. Reports",
        "steps": (
            "1. Go to Reports > Standard Reports (or BI > Standard Reports in Fusion)\n"
            '2. Look for banner: "Finance AI generated a financial summary for <Previous Month>"\n'
            '3. Click "Review summary" CTA > financial summary opens in browser\n'
            '4. On any KPI, click "Drill down on this insight"\n'
            "5. Verify 3 follow-up question suggestions appear\n"
            "6. Click a suggestion pill > verify drill-down data loads\n"
            "7. Go to Dashboard > Business Feed\n"
            "8. Look for DFY proactive summary card\n"
            '9. Click "Review summary" on the card'
        ),
        "evidence": (
            "PRINT 1: Standard Reports page with Finance AI banner visible\n"
            'PRINT 2: Financial summary page opened after clicking "Review summary"\n'
            'PRINT 3: KPI with "Drill down on this insight" link visible\n'
            "PRINT 4: 3 follow-up question suggestion pills displayed\n"
            "PRINT 5: Drill-down data loaded after clicking suggestion pill\n"
            "PRINT 6: Business Feed showing DFY proactive summary card\n"
            "PRINT 7: Summary opened from Business Feed card"
        ),
    },
    {
        "seq": 7,
        "id": 19,
        "name": "Calculated Fields on Reports",
        "pri": "P1",
        "flag": "NO",
        "nav": "1. Reports",
        "steps": (
            "BLOCKER: Feature Flag is OFF - need Intuit to enable\n"
            "---\n"
            "1. Open P&L or Balance Sheet (must be modern view)\n"
            "2. Click Customize > Calculated Fields\n"
            "3. Add custom formula COLUMN (e.g., profit margin %)\n"
            "4. Add custom formula ROW\n"
            "5. Save with new color\n"
            "6. Test on Balance Sheet, Trial Balance, AP/AR aging\n"
            "7. Verify formulas save and persist\n"
            "8. Test across at least 3 report types (44 supported)"
        ),
        "evidence": (
            "BLOCKED - Collect after flag is enabled:\n"
            'PRINT 1: Customize panel showing "Calculated Fields" option\n'
            "PRINT 2: Formula builder with custom COLUMN formula entered\n"
            "PRINT 3: Report showing custom column rendered with values\n"
            "PRINT 4: Formula builder with custom ROW formula entered\n"
            "PRINT 5: Report showing custom row rendered with values\n"
            "PRINT 6: Report saved with custom color applied\n"
            "PRINT 7: Same report reopened showing formulas persisted\n"
            "PRINT 8: At least 3 different reports with calculated fields working"
        ),
    },
    # ── 2. REPORTS CONSOLIDATED ──
    {
        "seq": 8,
        "id": 24,
        "name": "New ME Reports",
        "pri": "P1",
        "flag": "YES",
        "nav": "2. Reports (Consolidated)",
        "steps": (
            "1. Switch to Consolidated View (top bar entity switcher)\n"
            "2. Go to Reports > Standard Reports\n"
            "3. Scroll to Consolidated Reports section\n"
            "4. Find and open 5 NEW reports:\n"
            "   a. Invoice List\n"
            "   b. Transaction List by date\n"
            "   c. Transaction detail by account\n"
            "   d. Deposit detail\n"
            "   e. Check detail\n"
            "5. Open each report and verify data populates\n"
            "6. Test filter/display options"
        ),
        "evidence": (
            'PRINT 1: Consolidated View active (entity switcher showing "Consolidated")\n'
            "PRINT 2: Standard Reports page showing Consolidated Reports section with 5 new reports listed\n"
            "PRINT 3: Invoice List report opened with data populated\n"
            "PRINT 4: Transaction List by date report with data\n"
            "PRINT 5: Transaction detail by account report with data\n"
            "PRINT 6: Deposit detail report with data\n"
            "PRINT 7: Check detail report with data\n"
            "PRINT 8: Filter/display options on one of the reports (showing same options as SE)"
        ),
    },
    # ── 3. SETTINGS & DIMENSIONS ──
    {
        "seq": 9,
        "id": 22,
        "name": "Smart Dimension Assignment v2",
        "pri": "P1",
        "flag": "YES",
        "nav": "3. Settings & Dimensions",
        "steps": (
            "1. Go to Settings > Dimensions\n"
            '2. Click "Dimension Assignment" dropdown > verify tool opens\n'
            '3. Check for "Save for Later" option\n'
            '4. Click "All Products and Services" tab\n'
            "5. Select multiple items > Assign Dimensions > test bulk assignment\n"
            "6. NEW: Go to Accounting > Fixed Assets\n"
            '7. Click "Dimension Assignment" dropdown > verify works here too\n'
            "8. NEW: Go to Dashboard > Business Feed\n"
            "9. Look for Dimension assignment widget\n"
            '10. Click "Review" on Cost Group AI pop-up > validate suggestions\n'
            "11. Verify AI recommendations still working (sparkle icons)"
        ),
        "evidence": (
            'PRINT 1: Settings > Dimensions page with "Dimension Assignment" dropdown\n'
            "PRINT 2: Dimension Assignment tool opened\n"
            'PRINT 3: "Save for Later" option visible\n'
            'PRINT 4: "All Products and Services" tab with items listed\n'
            "PRINT 5: Bulk selection of items with Assign Dimensions dialog\n"
            'PRINT 6: NEW - Fixed Assets page showing "Dimension Assignment" dropdown\n'
            "PRINT 7: NEW - Business Feed with Dimension assignment widget\n"
            "PRINT 8: AI recommendations with sparkle icons on suggestions"
        ),
    },
    {
        "seq": 10,
        "id": 23,
        "name": "Dimensions on Workflow Automation",
        "pri": "P1",
        "flag": "YES",
        "nav": "3. Settings & Dimensions",
        "steps": (
            "1. Click gear icon > Manage Workflows (under Tools)\n"
            "2. Click + Custom workflow\n"
            "3. Select record type and action > Next > name workflow\n"
            "4. In conditions: verify Dimensions appear as selectable option\n"
            "5. Create REMINDER workflow:\n"
            '   Bill > "Marketing" dept dimension > amount > $5K > notify Manager\n'
            "6. Save and turn on\n"
            "7. Create APPROVAL workflow:\n"
            '   Invoice > "California" region dimension > overdue > notify lead\n'
            "8. Save and turn on\n"
            "9. Verify both appear in workflow list"
        ),
        "evidence": (
            "PRINT 1: Manage Workflows page\n"
            "PRINT 2: New workflow setup - conditions section showing Dimensions as option\n"
            "PRINT 3: REMINDER workflow - dimension condition set (e.g., Bill > Marketing > $5K)\n"
            "PRINT 4: REMINDER workflow saved and turned on (green toggle)\n"
            "PRINT 5: APPROVAL workflow - dimension condition set (e.g., Invoice > California > overdue)\n"
            "PRINT 6: APPROVAL workflow saved and turned on\n"
            "PRINT 7: Workflow list showing both new workflows active"
        ),
    },
    {
        "seq": 11,
        "id": 25,
        "name": "Parallel Approval in Workflow Automation",
        "pri": "P1",
        "flag": "YES",
        "nav": "3. Settings & Dimensions",
        "steps": (
            "1. Still in Manage Workflows > + Custom workflow\n"
            '2. Select record > action = "Approval" > Next > name it\n'
            '3. Under "Do this" > "Request Approval"\n'
            "4. Add 2+ parallel approvers to same step (up to 5)\n"
            "5. Set minimum approvers required\n"
            "6. Create hybrid: Step 1 Sequential + Step 2 Parallel\n"
            '7. Under "Notify via" check Task and/or Email\n'
            "8. Save and turn on\n"
            "9. Run workflow against transactions:\n"
            "    - 1 pending approval\n"
            "    - 1 approved by parallel approver\n"
            "    - 1 declined\n"
            '10. Check "History" in Manage Workflows\n'
            "11. Check audit trail on transaction form (hover info icon)"
        ),
        "evidence": (
            "PRINT 1: Approval workflow setup with 2+ parallel approvers listed\n"
            'PRINT 2: "Minimum approvers required" field set\n'
            "PRINT 3: Hybrid workflow: Step 1 Sequential + Step 2 Parallel visible\n"
            'PRINT 4: "Notify via" options (Task + Email) checked\n'
            "PRINT 5: Workflow saved and turned on\n"
            'PRINT 6: Transaction in "Pending Approval" status\n'
            'PRINT 7: Transaction "Approved" by parallel approver\n'
            'PRINT 8: Transaction "Declined" by approver\n'
            "PRINT 9: Manage Workflows > History section showing workflow runs\n"
            "PRINT 10: Transaction form with audit trail visible (hover info icon)"
        ),
    },
    {
        "seq": 12,
        "id": 35,
        "name": "Platform Customization",
        "pri": "P2",
        "flag": "N/A",
        "nav": "3. Settings & Dimensions",
        "steps": (
            "1. Check Settings area for Platform Customization options\n"
            "2. Note: After industry selection, user receives Guided Tour:\n"
            "   - Construction > construction-tailored Guided Tour\n"
            "   - Non-construction > generic Guided Tour\n"
            "3. Only applicable on NEW IES account sign-in at Set Up screen\n"
            "4. Document available customization options"
        ),
        "evidence": (
            "PRINT 1: Settings page showing any Platform Customization options\n"
            "NOTE: Only testable on new account setup. If not accessible:\n"
            'REPORT: Document as "N/A - requires new IES account setup"\n'
            "If accessible on fresh tenant:\n"
            "PRINT 2: Industry selection screen\n"
            "PRINT 3: Construction-tailored Guided Tour (first screen)\n"
            "PRINT 4: Generic Guided Tour (first screen) for comparison"
        ),
    },
    # ── 4. TAXES & BANKING ──
    {
        "seq": 13,
        "id": 11,
        "name": "Sales Tax AI (EA w/ Beta tag)",
        "pri": "P1",
        "flag": "-",
        "nav": "4. Taxes & Banking",
        "steps": (
            "1. Go to Business Feed > look for Sales Tax AI review prompt\n"
            "   OR Taxes > Sales Tax > Sales tax returns\n"
            '2. Click dropdown next to "View Summary" in Action column\n'
            "3. Observe Filing Pre-Check: P&L vs Sales Tax Liability reconciliation\n"
            "4. Review task list of impacted transactions\n"
            "5. Check AI-generated explanations and suggested next steps\n"
            "6. Choose action: fix transaction or mark as safe to ignore\n"
            "7. Check tabs: For Review, Resolved, Ignored\n"
            "8. Look for EA badge + Beta tag\n"
            "\n"
            "PREREQ: Income LI assigned to expense account triggers the agent.\n"
            "CAUTION: Turning on sales tax requires business address + agencies."
        ),
        "evidence": (
            "PRINT 1: Sales Tax page or Business Feed showing Sales Tax AI entry point\n"
            "PRINT 2: Filing Pre-Check showing P&L vs STL income comparison\n"
            "PRINT 3: Task list of impacted transactions with AI explanations\n"
            "PRINT 4: AI-generated explanation detail on a specific transaction\n"
            'PRINT 5: "For Review" tab with items listed\n'
            'PRINT 6: "Resolved" tab after fixing a transaction\n'
            'PRINT 7: "Ignored" tab after marking safe to ignore\n'
            "PRINT 8: EA (Early Access) badge + Beta tag visible on feature"
        ),
    },
    {
        "seq": 14,
        "id": 10,
        "name": "Accounting AI",
        "pri": "P2",
        "flag": "-",
        "nav": "4. Taxes & Banking",
        "steps": (
            "1. Go to Accounting > Bank Transactions\n"
            "2. Look for Ready-to-Post banner\n"
            "   (Appears when bank account has 3+ data-backed matches)\n"
            '3. Click "Review" to see best predictions\n'
            "4. Review and batch post transactions\n"
            "5. Note: From/To is required for Ready-to-Post\n"
            "6. Test Ready-to-Post filter\n"
            "\n"
            "PREREQ: Need 3+ transactions (Invoice, Expense, or Bill)\n"
            "that are data-backed within IES. UAT currently only has payroll."
        ),
        "evidence": (
            "PRINT 1: Bank Transactions page showing Ready-to-Post banner\n"
            'PRINT 2: "Review" view showing AI best predictions for batch\n'
            "PRINT 3: Batch post confirmation or result\n"
            "PRINT 4: Ready-to-Post filter applied in transaction list\n"
            "IF BANNER NOT PRESENT:\n"
            'REPORT: "Ready-to-Post banner requires 3+ data-backed matches.\n'
            "Currently insufficient data. Need refreshed bank transactions with\n"
            'Invoices/Expenses/Bills linked."'
        ),
    },
    # ── 5. SALES & ESTIMATES ──
    {
        "seq": 15,
        "id": 34,
        "name": "Project Estimates: Automation & Integrations",
        "pri": "P1",
        "flag": "N/A",
        "nav": "5. Sales & Estimates",
        "steps": (
            "1. Go to Sales > Estimates\n"
            "2. Create estimate against a project\n"
            "3. Verify Project Estimate form opens automatically\n"
            "4. Go to existing estimate list\n"
            '5. Click Actions > "Convert to Project Estimate"\n'
            "6. Go to Estimates > Manage settings\n"
            '7. Find "Auto-convert project-based estimates" toggle\n'
            "8. Toggle ON > create new estimate against project\n"
            "9. Verify auto-conversion happens\n"
            "10. Check for first-time user explanation popup"
        ),
        "evidence": (
            "PRINT 1: New estimate form auto-opening as Project Estimate (when project selected)\n"
            'PRINT 2: Estimate list > Actions menu showing "Convert to Project Estimate"\n'
            "PRINT 3: Converted Project Estimate opened\n"
            'PRINT 4: Estimates > Manage settings showing "Auto-convert" toggle\n'
            "PRINT 5: Toggle set to ON\n"
            "PRINT 6: First-time user explanation popup (if displayed)\n"
            "PRINT 7: New estimate auto-converted to PE after toggle ON"
        ),
    },
    {
        "seq": 16,
        "id": 31,
        "name": "Proposals",
        "pri": "P1",
        "flag": "YES",
        "nav": "5. Sales & Estimates",
        "steps": (
            "1. Go to All Apps > Customer Hub > Proposals\n"
            '2. Click "Create Proposal"\n'
            "3. Select customer > click Continue\n"
            "4. Use WYSIWYG editor: add text, images, logos, formatting\n"
            "5. Add element blocks of different formats\n"
            '6. Click "Continue to Send" > preview\n'
            "7. Verify customer details auto-populate\n"
            "8. Click Send\n"
            "9. Test Estimate > Proposal flow (line items carry over)\n"
            "10. Test Proposal > Estimate conversion on acceptance\n"
            "11. Verify e-signature on client portal\n"
            "12. Check draft/in-progress/completed views\n"
            "\n"
            "DATA: Create examples in Pending, Declined, and Accepted statuses"
        ),
        "evidence": (
            "PRINT 1: Proposals page in Customer Hub (list view)\n"
            'PRINT 2: "Create Proposal" > customer selection screen\n'
            "PRINT 3: WYSIWYG editor with text, images, logos added\n"
            "PRINT 4: Element blocks of different formats in editor\n"
            'PRINT 5: "Continue to Send" preview with customer details auto-populated\n'
            "PRINT 6: Sent proposal confirmation\n"
            "PRINT 7: Estimate > Proposal flow showing line items carried over\n"
            "PRINT 8: Client portal view with e-signature option\n"
            "PRINT 9: Proposal list showing different statuses (Pending, Declined, Accepted)"
        ),
    },
    {
        "seq": 17,
        "id": 41,
        "name": "Sales Orders",
        "pri": "P1",
        "flag": "YES",
        "nav": "5. Sales & Estimates",
        "steps": (
            "1. Non-linked SO: +Create > Sales Order > Customer > Products > Save\n"
            "2. SO from Estimate: All Apps > Customer Hub > Estimates\n"
            '   > Accepted estimate > dropdown > "Convert to Sales Order" > Save\n'
            '3. SO to Invoice: Open SO > dropdown next to "Review and Send"\n'
            '   > "Convert to invoice"\n'
            '4. SO to PO: Open SO > dropdown > "Convert to purchase order"\n'
            "5. Verify various fulfillment statuses\n"
            "6. Check reporting on Sales Orders\n"
            "\n"
            "DATA: Create examples linked to estimates, POs, invoices\n"
            "in various fulfillment statuses"
        ),
        "evidence": (
            "PRINT 1: Non-linked Sales Order created and saved\n"
            'PRINT 2: Estimate > "Convert to Sales Order" option in dropdown\n'
            "PRINT 3: Sales Order created from Estimate with line items\n"
            'PRINT 4: SO > dropdown showing "Convert to invoice" option\n'
            "PRINT 5: Invoice created from SO\n"
            'PRINT 6: SO > dropdown showing "Convert to purchase order" option\n'
            "PRINT 7: PO created from SO\n"
            "PRINT 8: Sales Order list showing various fulfillment statuses\n"
            "PRINT 9: Sales Order report or report including SO data\n"
            "ACCESS: Verify via Inventory > Sales Order AND Sales & Get Paid > Sales Order"
        ),
    },
    # ── 6. PRODUCTS & SERVICES ──
    {
        "seq": 18,
        "id": 28,
        "name": "Cost Groups",
        "pri": "P0",
        "flag": "YES",
        "nav": "6. Products & Services",
        "steps": (
            "1. Go to All Apps > Sales & Get Paid > Products & Services\n"
            "2. Verify Cost Group column (next to Category)\n"
            "3. Click Filter > verify filter by cost group\n"
            "4. Click Edit on product > select OOB cost group > Save\n"
            '   OR: Click "Review" on Cost Group AI pop-up > validate > Approve all\n'
            "5. Click More > Manage cost groups\n"
            "6. Verify pre-populated: Labor, Materials, Equipment, Sub, Misc\n"
            "7. Create a new cost group\n"
            "8. Bulk select items > Assign cost group\n"
            "9. Add new product > verify Cost Group dropdown\n"
            '10. Go to Project > Budget > Group by > verify "Cost Groups"\n'
            '11. Project Overview > Cost breakdown widget > "View Report"'
        ),
        "evidence": (
            "PRINT 1: Products & Services page with Cost Group column visible\n"
            "PRINT 2: Filter by cost group applied (filtered list)\n"
            "PRINT 3: Product edit showing Cost Group dropdown with OOB groups\n"
            "PRINT 4: AI pop-up for Cost Group review/approval (if visible)\n"
            "PRINT 5: Manage Cost Groups page showing 5 pre-populated (Labor, Materials, Equipment, Sub, Misc)\n"
            "PRINT 6: New cost group created\n"
            'PRINT 7: Bulk selection with "Assign cost group" action\n'
            "PRINT 8: New product form with Cost Group dropdown\n"
            'PRINT 9: Budget > "Group by" dropdown showing "Cost Groups" option\n'
            "PRINT 10: Budget grouped by cost groups\n"
            "PRINT 11: Project Overview > Cost breakdown widget\n"
            'REPORT: "View Report" from cost breakdown widget opened'
        ),
    },
    {
        "seq": 19,
        "id": 40,
        "name": "Moving Average Cost Method",
        "pri": "P2",
        "flag": "N/A",
        "nav": "6. Products & Services",
        "steps": (
            "1. Check Products & Services > inventory valuation settings\n"
            "2. Look for Moving Average Cost option (alongside FIFO)\n"
            "3. Note: New IES customers can select MAC instead of FIFO\n"
            "4. DT migrators seamlessly maintain MAC\n"
            "5. Check inventory reports with MAC valuation\n"
            "6. No dedicated UI page - methodology selection only"
        ),
        "evidence": (
            "PRINT 1: Inventory valuation settings showing MAC option (if visible)\n"
            "PRINT 2: Inventory report using MAC valuation (if applicable)\n"
            "IF NOT VISIBLE in existing environment:\n"
            'REPORT: "MAC selection available only on new IES account setup or DT migration.\n'
            'No dedicated UI page. Accounting methodology, not a feature page."'
        ),
    },
    # ── 7. PROJECTS ──
    {
        "seq": 20,
        "id": 27,
        "name": "Project Phases v2",
        "pri": "P0",
        "flag": "YES",
        "nav": "7. Projects",
        "steps": (
            "1. Go to All Apps > Projects > select a project\n"
            "2. Click Edit > expand Phases section\n"
            "3. Click Add Phase > define:\n"
            "   - Name, Description, Status, Start Date, End Date\n"
            "   - Add at least 3 phases contextual to project\n"
            "   - Dates sequential (some overlap OK)\n"
            "4. Save and close\n"
            "5. Open Project Details > verify Phases widget shows progress\n"
            "6. Mark a phase complete from widget\n"
            "7. Upload doc > verify PM Agent auto-generates phases\n"
            '8. Go to Budget > "Group By" dropdown > select "Phases"\n'
            "9. Drag products/services under appropriate phases\n"
            "10. Save and Publish budget\n"
            '11. Click "Convert to project estimate"\n'
            "12. Verify phases auto-populate in estimate\n"
            "13. Show/hide phases on estimate\n"
            "14. Open Bill/PO > verify Phase field selectable"
        ),
        "evidence": (
            "PRINT 1: Project Edit > Phases section expanded\n"
            "PRINT 2: Add Phase form with Name, Description, Status, Start/End dates filled\n"
            "PRINT 3: Project with 3+ phases defined (phase list view)\n"
            "PRINT 4: Project Details > Phases widget showing progress\n"
            "PRINT 5: Phase marked as complete from widget\n"
            "PRINT 6: Doc upload > PM Agent auto-generated phases\n"
            'PRINT 7: Budget > "Group By" dropdown showing "Phases" option\n'
            "PRINT 8: Budget with products/services organized under phases\n"
            'PRINT 9: Budget published > "Convert to project estimate" button\n'
            "PRINT 10: Project Estimate with phases auto-populated\n"
            "PRINT 11: Estimate with phases shown/hidden toggle\n"
            "PRINT 12: Bill or PO form with Phase field selectable"
        ),
    },
    {
        "seq": 21,
        "id": 30,
        "name": "AIA Billing",
        "pri": "P0",
        "flag": "YES",
        "nav": "7. Projects",
        "steps": (
            "1. Enable: Account & Settings > Sales > Progress Invoicing ON\n"
            "2. Open Project with phases (from step 20)\n"
            "3. Budget > Create Estimate from budget\n"
            "4. Verify phases group line items in estimate\n"
            "5. Apply line-item markups\n"
            "6. Select PDF View > verify phases grouping on PDF\n"
            "7. Verify AIA columns: Est. Total, Invoiced, Amount, Remaining\n"
            "8. Click Add to project > Invoice\n"
            "9. Set billing: % of milestone OR $ amount per line item\n"
            "10. Save > create 2nd progress invoice\n"
            "11. Verify sequential: builds on previous (Invoiced to Date updates)\n"
            "12. Check PDF renders with AIA calculations\n"
            "13. Check client portal view"
        ),
        "evidence": (
            "PRINT 1: Account & Settings > Sales > Progress Invoicing toggle ON\n"
            "PRINT 2: Project Estimate with phases grouping line items\n"
            "PRINT 3: Line-item markups applied\n"
            "PRINT 4: PDF View showing phases grouping on document\n"
            "PRINT 5: AIA columns visible: Est. Total, Invoiced, Amount, Remaining\n"
            "PRINT 6: First progress invoice with % or $ per line item set\n"
            'PRINT 7: Second progress invoice showing "Invoiced to Date" updated (sequential)\n'
            "PRINT 8: PDF export of AIA invoice with calculations\n"
            "PRINT 9: Client portal view of AIA invoice (if accessible)"
        ),
    },
    {
        "seq": 22,
        "id": 29,
        "name": "Budgets v3",
        "pri": "P1",
        "flag": "NO",
        "nav": "7. Projects",
        "steps": (
            "BLOCKER: Feature Flag is OFF - need Intuit to enable\n"
            "---\n"
            "1. Open project > Budget Widget\n"
            '2. Click "Create new budget" or drag-drop spreadsheet\n'
            "3. Import budget > verify import works\n"
            "4. Verify 0-click setup for ADV migrators\n"
            "5. Check budget overrun insights widget\n"
            "6. Export to Excel > verify\n"
            "7. Print to PDF > verify\n"
            "8. Check versioning across project cost reports\n"
            "9. Create Estimate from Budget > verify linking\n"
            '10. Verify "Group by Phase" option\n'
            '11. Verify "Group by Cost Group" option'
        ),
        "evidence": (
            "BLOCKED - Collect after flag is enabled:\n"
            'PRINT 1: Budget Widget with "Create new budget" button\n'
            "PRINT 2: Drag-drop spreadsheet import\n"
            "PRINT 3: Budget imported and populated\n"
            "PRINT 4: 0-click setup for ADV migrators (if applicable)\n"
            "PRINT 5: Budget overrun insights widget\n"
            "EXPORT 1: Excel export of budget\n"
            "EXPORT 2: PDF print of budget\n"
            "PRINT 6: Budget versioning in project cost reports\n"
            "PRINT 7: Budget > Estimate linking\n"
            'PRINT 8: "Group by Phase" view\n'
            'PRINT 9: "Group by Cost Group" view'
        ),
    },
    {
        "seq": 23,
        "id": 32,
        "name": "Negative Change Orders",
        "pri": "P1",
        "flag": "YES",
        "nav": "7. Projects",
        "steps": (
            "1. Open project with existing estimate\n"
            "2. Click Add to project > Change Orders\n"
            "3. Select project estimate\n"
            "4. Add items (positive) - verify works\n"
            "5. Remove items (negative) - verify net-negative CO allowed\n"
            "6. Verify system blocks if total estimate < $0\n"
            "7. Note: Discounts disabled for negative totals\n"
            "8. Save and send to customer\n"
            "9. Accept CO > verify project estimate income adjusted\n"
            "10. Check Change Order Reports for negative values\n"
            "11. Verify sales tax behaves like refund"
        ),
        "evidence": (
            "PRINT 1: Change Order creation with items being added (positive)\n"
            "PRINT 2: Change Order with items removed (negative values shown)\n"
            "PRINT 3: Net-negative Change Order allowed and saved\n"
            "PRINT 4: System blocking CO when total estimate would go < $0\n"
            "PRINT 5: Change Order sent to customer\n"
            "PRINT 6: Accepted CO > project estimate income adjusted (before/after)\n"
            "PRINT 7: Change Order Report showing negative values\n"
            "PRINT 8: Sales tax on negative CO showing refund-like behavior"
        ),
    },
    {
        "seq": 24,
        "id": 12,
        "name": "Project Management AI",
        "pri": "P1",
        "flag": "NO",
        "nav": "7. Projects",
        "steps": (
            "BLOCKER: Feature Flag is OFF - need Intuit to enable\n"
            'NOTE: "Not yet working in UAT environment" (2/11-JW)\n'
            "---\n"
            "1. +Create > New Project\n"
            "2. Verify AI agent panel appears on left with insights\n"
            "3. Check cost, price, profitability insights\n"
            "4. Create estimate > verify PM agent helps\n"
            "5. Upload doc > verify phases auto-generated\n"
            "6. Attach estimates: ensure open estimates exist for customer\n"
            "   > Create project with same customer\n"
            "   > Transactions tab > banner to attach estimate (converting to PE)\n"
            "7. Project closing > verify AI-generated summary\n"
            "8. Verify improved UI for AI insights on project dashboard"
        ),
        "evidence": (
            "BLOCKED + NOT WORKING UAT (2/11) - Collect when available:\n"
            "PRINT 1: New Project form with AI agent panel on left side\n"
            "PRINT 2: AI insights showing cost/price/profitability data\n"
            "PRINT 3: PM agent helping with estimate creation\n"
            "PRINT 4: Doc uploaded > phases auto-generated by AI\n"
            "PRINT 5: Transactions tab > banner to attach recently created estimate\n"
            "PRINT 6: Estimate attached and converted to Project Estimate\n"
            "PRINT 7: Project closing > AI-generated summary\n"
            "PRINT 8: Improved UI for AI insights on project dashboard"
        ),
    },
    {
        "seq": 25,
        "id": 33,
        "name": "PM AI Enhancements",
        "pri": "P2",
        "flag": "N/A",
        "nav": "7. Projects",
        "steps": (
            "1. Same area as PM AI (ID 12)\n"
            "2. Verify proactive cost/price/profitability insights on project setup\n"
            "3. Verify recently created estimates banner on Transactions tab\n"
            "4. Verify improved UI for AI insights on project dashboard\n"
            "5. Compare with PM AI base - note specific enhancements"
        ),
        "evidence": (
            "PRINT 1: Proactive insights on project setup (compare with base PM AI)\n"
            "PRINT 2: Estimates banner on Transactions tab\n"
            "PRINT 3: Improved UI for AI insights on project dashboard\n"
            "REPORT: Document specific enhancements vs base PM AI (ID 12)"
        ),
    },
    # ── 8. PURCHASES & INVENTORY ──
    {
        "seq": 26,
        "id": 39,
        "name": "Item Receipt",
        "pri": "P1",
        "flag": "YES",
        "nav": "8. Purchases & Inventory",
        "steps": (
            "1. Enable: Settings > Account & Settings > Sales tab\n"
            '   > Products & Services > Edit > "Use item receipts to update qty" ON > Save\n'
            "2. Go to Inventory > Create item receipt\n"
            "3. Select vendor\n"
            "4. Select purchase order (if linked) > OK\n"
            "5. Or enter items manually in detail area\n"
            "6. Save Item Receipt\n"
            "7. Test partial receipt: receive subset of items\n"
            "8. Verify inventory quantities update\n"
            "9. Create Bill from Item Receipt\n"
            "10. Check PO status updates\n"
            "\n"
            "DATA: 1 IR standalone + 1 IR linked to PO (contextual vendor/items)"
        ),
        "evidence": (
            'PRINT 1: Account & Settings > Sales > "Use item receipts" toggle ON\n'
            'PRINT 2: "Create item receipt" form with vendor selected\n'
            "PRINT 3: Item Receipt linked to PO (PO reference visible)\n"
            "PRINT 4: Item Receipt saved (standalone, no PO)\n"
            "PRINT 5: Partial receipt - subset of items received\n"
            "PRINT 6: Inventory quantity BEFORE and AFTER receipt (showing update)\n"
            "PRINT 7: Bill created from Item Receipt\n"
            "PRINT 8: PO status updated after receipt (partial/complete)"
        ),
    },
    # ── 9. PAYROLL & TIME ──
    {
        "seq": 27,
        "id": 37,
        "name": "Garnishments",
        "pri": "P1",
        "flag": "YES",
        "nav": "9. Payroll & Time",
        "steps": (
            "1. Go to Payroll > Employees\n"
            "2. Select employee (one in most recent payroll)\n"
            "3. Navigate to Deductions and Contributions > Start or Edit\n"
            '4. Click "Add garnishment"\n'
            "5. Select Garnishment type\n"
            "6. Add details: Agency (from Vendor list), amount, frequency\n"
            '   - If agency not in vendors: "Add vendor" from dropdown\n'
            "7. Save garnishment\n"
            "8. Preview paycheck > verify garnishment deduction\n"
            "9. Run payroll for this employee\n"
            "10. Check Payroll Tax Center for garnishment liability\n"
            "11. Check payroll reports for garnishment data"
        ),
        "evidence": (
            "PRINT 1: Employee profile > Deductions and Contributions section\n"
            'PRINT 2: "Add garnishment" form with type, agency, amount, frequency\n'
            "PRINT 3: Garnishment saved and listed in employee deductions\n"
            "PRINT 4: Paycheck preview showing garnishment deduction line\n"
            "PRINT 5: Payroll run completed with garnishment applied\n"
            "PRINT 6: Payroll Tax Center showing garnishment liability\n"
            "REPORT: Payroll report showing garnishment data for the employee"
        ),
    },
    {
        "seq": 28,
        "id": 38,
        "name": "Time 2.0 - Assignments in QB Time",
        "pri": "P1",
        "flag": "YES",
        "nav": "9. Payroll & Time",
        "steps": (
            "1. Accounts and Settings > Time Settings > Timesheet fields\n"
            '2. Click Edit > for a field > Actions > "Assign Customers"\n'
            "3. Select customers to assign > Save\n"
            "4. Time > Assignments > Customer tab\n"
            '5. Actions dropdown > "Assign Workers"\n'
            "6. Assign workers to a project with lots of data\n"
            "7. Assign 2+ employees to same project\n"
            "8. Assign certain service items to some, others to different employees\n"
            "9. Time > Assignments > Worker tab\n"
            "10. Verify assignments from worker perspective\n"
            "11. Verify workers see only relevant projects/customers\n"
            "12. Check time entries against assignments\n"
            "13. Check reporting on assignments"
        ),
        "evidence": (
            'PRINT 1: Time Settings > Timesheet fields with "Assign Customers" option\n'
            "PRINT 2: Customer assignment drawer with customers selected\n"
            "PRINT 3: Time > Assignments > Customer tab showing project list\n"
            'PRINT 4: "Assign Workers" drawer with workers selected for a project\n'
            "PRINT 5: Multiple workers assigned to same project\n"
            "PRINT 6: Service items assigned to specific employees\n"
            "PRINT 7: Worker tab showing assignments from worker perspective\n"
            "PRINT 8: Worker view showing only relevant projects/customers (filtered)\n"
            "REPORT: Assignments report or time entries linked to assignments"
        ),
    },
    {
        "seq": 29,
        "id": 36,
        "name": "ME Employee Hub",
        "pri": "P2",
        "flag": "NO",
        "nav": "9. Payroll & Time",
        "steps": (
            "BLOCKER: Feature Flag is OFF - need Intuit to enable\n"
            "---\n"
            "1. Enter Consolidated View\n"
            "2. Left nav > Workforce tab\n"
            '3. Select "Employee Hub"\n'
            "4. View unified table of W2 employees across all entities\n"
            "5. Click employee Name > auto-switch to their entity\n"
            "6. Verify profile page opens\n"
            "7. Check cross-entity data aggregation"
        ),
        "evidence": (
            "BLOCKED - Collect after flag is enabled:\n"
            "PRINT 1: Consolidated View > Workforce tab visible\n"
            "PRINT 2: Employee Hub showing unified table across entities\n"
            "PRINT 3: Employee list with entity column (showing which entity each belongs to)\n"
            "PRINT 4: Click employee > auto-switched to their entity context\n"
            "PRINT 5: Individual employee profile page opened"
        ),
    },
    # ── 10. DASHBOARD & MISC ──
    {
        "seq": 30,
        "id": 15,
        "name": "Solutions Specialist - COA Standardization",
        "pri": "P1",
        "flag": "-",
        "nav": "10. Dashboard & Misc",
        "steps": (
            "1. Go to Dashboard > Business Feed\n"
            "2. Look for COA standardization recommendations\n"
            "3. Same location where customers standardize COA + cleanup vendor lists\n"
            "4. Check Missing Attachments Checklist\n"
            "5. Check Unbilled Expense Checklist enhancements\n"
            "6. Check Accounting Mismatch Checklist enhancements\n"
            "7. Verify memo column, sorting by discrepancy magnitude\n"
            "\n"
            "NOTE: Automated intercompany mapping descoped to limited availability"
        ),
        "evidence": (
            "PRINT 1: Business Feed showing COA standardization widget/recommendation\n"
            "PRINT 2: COA standardization recommendation detail\n"
            "PRINT 3: Missing Attachments Checklist\n"
            "PRINT 4: Unbilled Expense Checklist (showing enhancements)\n"
            "PRINT 5: Accounting Mismatch Checklist (showing enhancements)\n"
            "PRINT 6: Memo column visible + sorting by discrepancy magnitude"
        ),
    },
    {
        "seq": 31,
        "id": 26,
        "name": "DFY Migration Experience",
        "pri": "P2",
        "flag": "N/A",
        "nav": "10. Dashboard & Misc",
        "steps": (
            "NOT TESTABLE in TestBox environment\n"
            "---\n"
            "1. Migration flow happens in QB Desktop\n"
            "2. No data needed within IES\n"
            "3. Document as N/A\n"
            "4. If accessible, check post-migration checklists:\n"
            "   a. Missing Attachments\n"
            "   b. Unbilled Expense\n"
            "   c. Accounting Mismatch"
        ),
        "evidence": (
            'REPORT: "N/A - Migration flow requires QB Desktop.\n'
            'Not testable in TestBox/IES environment."\n'
            "IF post-migration checklists visible:\n"
            "PRINT 1: Missing Attachments Checklist\n"
            "PRINT 2: Unbilled Expense Checklist\n"
            "PRINT 3: Accounting Mismatch Checklist"
        ),
    },
    # ── EXCLUDED ──
    {
        "seq": 32,
        "id": 14,
        "name": "Customer AI",
        "pri": "N/A",
        "flag": "-",
        "nav": "EXCLUDED",
        "steps": ("EXCLUDED: Decision NOT to launch to IES (Jan 21, 2026)\nDo not include in demo validation."),
        "evidence": (
            "NO EVIDENCE NEEDED\n"
            'REPORT: "Feature excluded from Feb release per decision Jan 21, 2026.\n'
            'Not launching in IES."'
        ),
    },
]

# Write data
for i, f in enumerate(features):
    row = i + 2
    vals = [
        f["seq"],
        f["id"],
        f["name"],
        f["pri"],
        f["flag"],
        f["nav"],
        f["steps"],
        f["evidence"],
        "NOT STARTED",
        "",
    ]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin
        c.alignment = wrap
        c.font = normal_font

    # Center small cols
    for col in [1, 2, 4, 5, 9]:
        ws.cell(row=row, column=col).alignment = center_wrap

    # P0 highlight
    if f["pri"] == "P0":
        ws.cell(row=row, column=4).fill = p0_fill
        ws.cell(row=row, column=4).font = Font(name="Calibri", size=10, bold=True, color="CC0000")

    # Blocked
    if f["flag"] == "NO":
        for col in [5, 7, 8, 9]:
            ws.cell(row=row, column=col).fill = blocked_fill
        ws.cell(row=row, column=5).font = blocked_font

    # Excluded
    if f["pri"] == "N/A":
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = na_fill

# Column widths
widths = {1: 5, 2: 5, 3: 34, 4: 8, 5: 8, 6: 22, 7: 65, 8: 65, 9: 14, 10: 25}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Row heights
for row in range(2, len(features) + 2):
    ws.row_dimensions[row].height = 200
ws.row_dimensions[1].height = 30

ws.freeze_panes = "D2"
ws.auto_filter.ref = f"A1:J{len(features) + 1}"

output = r"C:\Users\adm_r\Downloads\Winter_Release_FY26_Feature_Check_Plan_v3.xlsx"
wb.save(output)
print(f"SAVED: {output}")
print(f"Total features: {len(features)}")

# Count evidence items
total_evidence = 0
for f in features:
    lines = [
        line.strip()
        for line in f["evidence"].split("\n")
        if line.strip().startswith("PRINT") or line.strip().startswith("EXPORT") or line.strip().startswith("REPORT")
    ]
    total_evidence += len(lines)
print(f"Total evidence items: {total_evidence}")
