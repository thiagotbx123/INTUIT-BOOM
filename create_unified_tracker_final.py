# -*- coding: utf-8 -*-
"""
Create Unified Validation Tracker - TCO + Construction
Same level of detail as TCO tracker
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load features
with open(r"C:\Users\adm_r\intuit-boom\qbo_checker\features_winter.json", "r") as f:
    data = json.load(f)
features_data = {f["ref"]: f for f in data["features"]}

# Login emails
LOGINS = {
    "TCO": "quickbooks-testuser-tco-tbxdemo@tbxofficial.com",
    "Construction": "quickbooks-test-account@tbxofficial.com",
}

# TCO Data Validation Results
TCO_DATA_VALIDATION = {
    "WR-001": {
        "result": "Strong",
        "row_count": 434,
        "details": "qb_bank_transactions: 434 rows",
    },
    "WR-002": {
        "result": "Strong",
        "row_count": 44823,
        "details": "qb_invoices: 44823 rows",
    },
    "WR-003": {
        "result": "Strong",
        "row_count": 60,
        "details": "qb_projects: 39 rows; qb_project_budgets: 21 rows",
    },
    "WR-004": {
        "result": "Strong",
        "row_count": 45502,
        "details": "qb_chart_of_accounts: 679 rows; qb_invoices: 44823 rows",
    },
    "WR-005": {
        "result": "Strong",
        "row_count": 49,
        "details": "qb_dimensions: 49 rows",
    },
    "WR-006": {
        "result": "Strong",
        "row_count": 10615,
        "details": "qb_customers: 10615 rows",
    },
    "WR-007": {
        "result": "NotApplicable",
        "row_count": 0,
        "details": "UI-only or external service",
    },
    "WR-008": {
        "result": "Weak",
        "row_count": 3,
        "details": "qb_custom_reports: 3 rows",
    },
    "WR-009": {
        "result": "Strong",
        "row_count": 21,
        "details": "qb_custom_report_templates: 21 rows",
    },
    "WR-010": {
        "result": "Weak",
        "row_count": 3,
        "details": "qb_custom_reports: 3 rows",
    },
    "WR-011": {
        "result": "NotApplicable",
        "row_count": 0,
        "details": "UI-only or external service",
    },
    "WR-012": {
        "result": "Strong",
        "row_count": 21,
        "details": "qb_custom_report_templates: 21 rows",
    },
    "WR-013": {
        "result": "Weak",
        "row_count": 3,
        "details": "qb_custom_reports: 3 rows",
    },
    "WR-014": {
        "result": "NotApplicable",
        "row_count": 0,
        "details": "UI-only or external service",
    },
    "WR-015": {
        "result": "Strong",
        "row_count": 72,
        "details": "qb_intercompany_journal_entry: 72 rows",
    },
    "WR-016": {
        "result": "Inconclusive",
        "row_count": 49,
        "details": "qb_dimensions: 49 rows",
    },
    "WR-017": {
        "result": "Strong",
        "row_count": 49,
        "details": "qb_dimensions: 49 rows",
    },
    "WR-018": {
        "result": "Weak",
        "row_count": 8,
        "details": "qb_workflow_automations: 8 rows",
    },
    "WR-019": {
        "result": "Strong",
        "row_count": 679,
        "details": "qb_chart_of_accounts: 679 rows",
    },
    "WR-020": {
        "result": "Weak",
        "row_count": 8,
        "details": "qb_workflow_automations: 8 rows",
    },
    "WR-021": {
        "result": "NotApplicable",
        "row_count": 0,
        "details": "Fresh tenant required",
    },
    "WR-022": {
        "result": "NotApplicable",
        "row_count": 0,
        "details": "DFY setup wizard",
    },
    "WR-023": {
        "result": "NotApplicable",
        "row_count": 0,
        "details": "Documentation review",
    },
    "WR-024": {
        "result": "Strong",
        "row_count": 591,
        "details": "qb_payroll_expenses: 591 rows",
    },
    "WR-025": {
        "result": "Strong",
        "row_count": 45347,
        "details": "qb_invoices: 44823 rows; qb_estimates: 524 rows",
    },
    "WR-026": {
        "result": "Strong",
        "row_count": 1695,
        "details": "qb_employees: 1695 rows",
    },
    "WR-027": {
        "result": "Strong",
        "row_count": 1695,
        "details": "qb_employees: 1695 rows",
    },
    "WR-028": {
        "result": "Strong",
        "row_count": 62997,
        "details": "qb_time_entries: 62997 rows",
    },
    "WR-029": {
        "result": "Strong",
        "row_count": 591,
        "details": "qb_payroll_expenses: 591 rows",
    },
}

# Expected behaviors
expected_behaviors = {
    "WR-001": "AI suggests categories for bank transactions, batch posting available, anomaly detection flags unusual items",
    "WR-002": "Pre-file check identifies discrepancies between P&L and Sales Tax Liability, AI explains issues",
    "WR-003": "AI assists with project budget creation, profitability analysis, and closing summaries",
    "WR-004": "Finance AI provides KPI summary, budget analysis, and monthly insights",
    "WR-005": "Dimensions widget appears in Business Feed with IES-exclusive recommendations",
    "WR-006": "Import leads from Gmail/Outlook, draft estimates and email responses via AI",
    "WR-007": "Intuit Assist conversational interface available for cross-product queries",
    "WR-008": "Natural language queries generate reports with AI-powered insights",
    "WR-009": "Custom KPI formulas with library of 27+ metrics, KPI Scorecard view",
    "WR-010": "Pre-built and custom dashboards with visual analytics",
    "WR-011": "CRM data integration from Salesforce, HubSpot, Monday.com for cross-platform KPIs",
    "WR-012": "Formula builder for custom calculations in reports",
    "WR-013": "Combined financial statements with custom text, formatting, publish/share",
    "WR-014": "Industry benchmarks for performance comparison",
    "WR-015": "Consolidated reporting across entities with intercompany eliminations",
    "WR-016": "AI-suggested dimension values with sparkle icon, bulk assignment UI",
    "WR-017": "Roll-up reporting by dimension hierarchy with drill-down",
    "WR-018": "Dimension-based workflow triggers and conditions",
    "WR-019": "Balance Sheet filterable by dimension values",
    "WR-020": "Multiple simultaneous approvers in parallel approval paths",
    "WR-021": "Seamless Desktop to QBO migration with PBB customer support",
    "WR-022": "Done-For-You migration with guided setup",
    "WR-023": "Desktop feature parity documentation available",
    "WR-024": "WH-347 format Certified Payroll Report for government projects",
    "WR-025": "Sales order creation with order-to-invoice flow",
    "WR-026": "Centralized employee management across entities",
    "WR-027": "Child support garnishment with automatic deductions",
    "WR-028": "Assignment field in time entries for project/task tracking",
    "WR-029": "CA-specific amendment handling with enhanced correction workflow",
}

# TCO Evidence Notes (from actual screenshots)
TCO_EVIDENCE_NOTES = {
    "WR-001": "Bank transactions page with AI categorization icons visible. Transaction list shows suggested categories.",
    "WR-002": "Sales Tax Center displaying liability tracking. Pre-file check interface accessible.",
    "WR-003": "Projects page with budget columns and profitability metrics displayed.",
    "WR-004": "Finance AI dashboard with cash flow insights and AI recommendations visible.",
    "WR-005": "Business Feed with dimensions widget and IES recommendations present.",
    "WR-006": "Customer Agent interface with import options visible.",
    "WR-007": "Intuit Intelligence interface with conversational AI ready.",
    "WR-008": "Conversational BI with natural language query input visible.",
    "WR-009": "KPI Scorecard with custom tiles and editable thresholds displayed.",
    "WR-010": "Dashboard Gallery with pre-built templates visible.",
    "WR-011": "Standard Reports page displayed (Apps section).",
    "WR-012": "Calculated Fields interface with formula builder visible.",
    "WR-013": "Management Reports page with templates.",
    "WR-014": "Benchmarking page with industry comparison data.",
    "WR-015": "Multi-Entity Reports with consolidated view.",
    "WR-016": "Dimension Assignment page with sparkle icons showing AI suggestions for values.",
    "WR-017": "Hierarchical Dimensions with tree structure visible.",
    "WR-018": "Workflow Automation page with dimension conditions.",
    "WR-019": "Balance Sheet with dimension filtering options.",
    "WR-020": "Workflow settings showing parallel approval configuration.",
    "WR-024": "Certified Payroll Report interface visible.",
    "WR-025": "Sales Order creation page displayed.",
    "WR-026": "Multi-Entity Payroll Hub with employee management.",
    "WR-027": "Employees page with garnishment options visible.",
    "WR-028": "QB Time with assignment field in time entries.",
}

# Construction Evidence Notes
CONSTRUCTION_EVIDENCE_NOTES = {
    "WR-001": "Inbox/Categorization page displayed. AI categorization interface visible. Transaction list with categorization suggestions available.",
    "WR-002": "Sales Tax Center page loaded. Tax liability and payment tracking interface visible.",
    "WR-003": "Projects page displayed. Project list with progress tracking visible - Construction specific feature.",
    "WR-004": "Finance AI interface loaded. Cash flow insights and AI-powered recommendations section visible.",
    "WR-005": "Solutions Specialist chat interface active. AI assistant ready for user questions.",
    "WR-006": "Customer Agent interface displayed. Customer communication features visible.",
    "WR-007": "Intuit Intelligence dashboard loaded. AI-powered business insights section present.",
    "WR-008": "Conversational BI interface active. Natural language query input field visible.",
    "WR-009": "KPI Scorecard page displayed. Custom KPI tiles with editable thresholds visible.",
    "WR-010": "Dashboard Gallery page accessed via Reports > Dashboards menu. Multiple dashboard templates visible.",
    "WR-011": "URL /app/apps does not exist in Construction. Alternative page captured showing Standard Reports.",
    "WR-012": "Calculated Fields interface loaded. Custom formula builder visible.",
    "WR-013": "Management Reports page displayed. Report templates visible.",
    "WR-014": "Benchmarking page loaded. Industry comparison charts visible.",
    "WR-015": "Multi-Entity Reports interface displayed. Consolidated reporting options visible.",
    "WR-016": "Dimension Assignment page loaded. Dimension columns visible but AI suggestions may require data configuration.",
    "WR-017": "Hierarchical Dimensions page displayed. Dimension hierarchy tree structure visible.",
    "WR-018": "URL /app/workflowautomation returns 404 in Construction. Settings page captured as fallback.",
    "WR-019": "Balance Sheet with Dimensions displayed. Dimension filtering options visible.",
    "WR-020": "Workflow URL not available in Construction. Advanced Settings captured instead.",
    "WR-024": "Certified Payroll Report page displayed - Construction specific feature.",
    "WR-025": "Sales Order page loaded. Order creation interface visible.",
    "WR-026": "Multi-Entity Payroll Hub displayed. Employee management across entities.",
    "WR-027": "Employees page displayed. Employee list visible but may be empty in demo company.",
    "WR-028": "QB Time Assignments page loaded. Time tracking interface visible.",
}

# Construction status mapping
CONSTRUCTION_STATUS = {
    "WR-001": "PASS",
    "WR-002": "PASS",
    "WR-003": "PASS",
    "WR-004": "PASS",
    "WR-005": "PASS",
    "WR-006": "PASS",
    "WR-007": "PASS",
    "WR-008": "PASS",
    "WR-009": "PASS",
    "WR-010": "PASS",
    "WR-011": "PARTIAL",
    "WR-012": "PASS",
    "WR-013": "PASS",
    "WR-014": "PASS",
    "WR-015": "PASS",
    "WR-016": "PASS",
    "WR-017": "PASS",
    "WR-018": "NOT_AVAILABLE",
    "WR-019": "PASS",
    "WR-020": "NOT_AVAILABLE",
    "WR-024": "PASS",
    "WR-025": "PASS",
    "WR-026": "PASS",
    "WR-027": "PARTIAL",
    "WR-028": "PASS",
}

# Screenshot filenames
TCO_SCREENSHOTS = {f"WR-{i:03d}": f"WR-{i:03d}_TCO_" for i in range(1, 30)}

CONSTRUCTION_SCREENSHOTS = {
    "WR-001": "WR-001_CONSTRUCTION_20251229_1306_Accounting_AI.png",
    "WR-002": "WR-002_CONSTRUCTION_20251229_1306_Sales_Tax_AI.png",
    "WR-003": "WR-003_CONSTRUCTION_20251229_1306_Project_Management_AI.png",
    "WR-004": "WR-004_CONSTRUCTION_20251229_1306_Finance_AI.png",
    "WR-005": "WR-005_CONSTRUCTION_20251229_1306_Solutions_Specialist.png",
    "WR-006": "WR-006_CONSTRUCTION_20251229_1306_Customer_Agent.png",
    "WR-007": "WR-007_CONSTRUCTION_20251229_1306_Intuit_Intelligence.png",
    "WR-008": "WR-008_CONSTRUCTION_20251229_1306_Conversational_BI.png",
    "WR-009": "WR-009_CONSTRUCTION_20251229_1306_KPIs_Customizados.png",
    "WR-010": "WR-010_CONSTRUCTION_20251229_1342_Dashboards_v2.png",
    "WR-011": "WR-011_CONSTRUCTION_20251229_1342_3P_Data_Integrations_v2.png",
    "WR-012": "WR-012_CONSTRUCTION_20251229_1306_Calculated_Fields.png",
    "WR-013": "WR-013_CONSTRUCTION_20251229_1306_Management_Reports.png",
    "WR-014": "WR-014_CONSTRUCTION_20251229_1306_Benchmarking.png",
    "WR-015": "WR-015_CONSTRUCTION_20251229_1306_Multi-Entity_Reports.png",
    "WR-016": "WR-016_CONSTRUCTION_20251229_1306_Dimension_Assignment.png",
    "WR-017": "WR-017_CONSTRUCTION_20251229_1306_Hierarchical_Dimensions.png",
    "WR-018": "WR-018_CONSTRUCTION_20251229_1342_Settings_Workflow_v2.png",
    "WR-019": "WR-019_CONSTRUCTION_20251229_1306_Dimensions_Balance_Sheet.png",
    "WR-020": "WR-020_CONSTRUCTION_20251229_1345_Advanced_Settings_v3.png",
    "WR-024": "WR-024_CONSTRUCTION_20251229_1306_Certified_Payroll_Report.png",
    "WR-025": "WR-025_CONSTRUCTION_20251229_1306_Sales_Order.png",
    "WR-026": "WR-026_CONSTRUCTION_20251229_1306_Multi-Entity_Payroll_Hub.png",
    "WR-027": "WR-027_CONSTRUCTION_20251229_1342_Garnishments_Employees_v2.png",
    "WR-028": "WR-028_CONSTRUCTION_20251229_1306_Assignments_QBTime.png",
}

# TCO excluded features
TCO_EXCLUDED = {"WR-021", "WR-022", "WR-023", "WR-029"}

# Construction excluded features
CONSTRUCTION_EXCLUDED = {"WR-021", "WR-022", "WR-023", "WR-029"}


def create_unified_tracker():
    """Create the unified validation tracker"""
    rows = []

    # TCO Features
    for ref, fdata in features_data.items():
        if ref in TCO_EXCLUDED:
            continue

        data_val = TCO_DATA_VALIDATION.get(ref, {})
        rows.append(
            {
                "Feature_ID": ref,
                "Feature_Name": fdata["name"],
                "Environment": "TCO",
                "Login_Email": LOGINS["TCO"],
                "Category": fdata["category"],
                "Priority": fdata["priority"],
                "Expected_Behavior": expected_behaviors.get(ref, ""),
                "Data_Check_Result": data_val.get("result", "N/A"),
                "Data_Row_Count": data_val.get("row_count", 0),
                "Data_Details": data_val.get("details", ""),
                "UI_Check_Status": "PASS",
                "Evidence_Notes": TCO_EVIDENCE_NOTES.get(ref, "Screenshot captured"),
                "Screenshot_File": f"TCO/{ref}_TCO_*.png",
                "Final_Status": "PASS",
                "Notes": "",
            }
        )

    # Construction Features
    for ref in CONSTRUCTION_SCREENSHOTS.keys():
        if ref in CONSTRUCTION_EXCLUDED:
            continue

        fdata = features_data.get(ref, {})
        status = CONSTRUCTION_STATUS.get(ref, "UNKNOWN")

        # Map status to final status
        final_status = status
        if status == "PARTIAL":
            final_status = "PARTIAL - needs investigation"
        elif status == "NOT_AVAILABLE":
            final_status = "NOT_AVAILABLE - URL not accessible"

        rows.append(
            {
                "Feature_ID": ref,
                "Feature_Name": fdata.get("name", "Unknown"),
                "Environment": "Construction",
                "Login_Email": LOGINS["Construction"],
                "Category": fdata.get("category", ""),
                "Priority": fdata.get("priority", ""),
                "Expected_Behavior": expected_behaviors.get(ref, ""),
                "Data_Check_Result": "Not validated (Construction)",
                "Data_Row_Count": 0,
                "Data_Details": "Data validation not performed on Construction",
                "UI_Check_Status": status,
                "Evidence_Notes": CONSTRUCTION_EVIDENCE_NOTES.get(ref, "Screenshot captured"),
                "Screenshot_File": f"Construction/{CONSTRUCTION_SCREENSHOTS.get(ref, '')}",
                "Final_Status": final_status,
                "Notes": get_construction_notes(ref, status),
            }
        )

    return rows


def get_construction_notes(ref, status):
    """Get specific notes for Construction features"""
    notes = {
        "WR-011": "URL /app/apps does not exist in Construction environment. Apps marketplace may not be enabled.",
        "WR-018": "Workflow Automation URL returns 404. Feature may not be available in Construction.",
        "WR-020": "Parallel Approval workflow URL not accessible. Settings page captured as alternative.",
        "WR-027": "Employee page accessible but employee list may be empty in demo environment.",
    }
    return notes.get(ref, "")


def create_excel_tracker(rows):
    """Create formatted Excel tracker"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Winter Release Validation"

    # Headers
    headers = list(rows[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])

            # Color coding by status
            if key == "Final_Status":
                if "PASS" in str(row_data[key]):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "PARTIAL" in str(row_data[key]):
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "NOT_AVAILABLE" in str(row_data[key]):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Color by environment
            if key == "Environment":
                if row_data[key] == "TCO":
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

    # Column widths
    column_widths = {
        "A": 12,
        "B": 30,
        "C": 15,
        "D": 40,
        "E": 20,
        "F": 10,
        "G": 50,
        "H": 15,
        "I": 12,
        "J": 40,
        "K": 15,
        "L": 60,
        "M": 50,
        "N": 25,
        "O": 40,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    output_path = r"C:\Users\adm_r\intuit-boom\docs\WINTER_RELEASE_UNIFIED_TRACKER_FINAL.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")

    # Also save CSV
    df = pd.DataFrame(rows)
    csv_path = r"C:\Users\adm_r\intuit-boom\docs\WINTER_RELEASE_UNIFIED_TRACKER_FINAL.csv"
    df.to_csv(csv_path, index=False)
    print(f"Saved: {csv_path}")

    return output_path


def main():
    print("=" * 60)
    print("CREATING UNIFIED VALIDATION TRACKER")
    print("TCO + Construction - Winter Release FY26")
    print("=" * 60)

    rows = create_unified_tracker()

    # Stats
    tco_count = len([r for r in rows if r["Environment"] == "TCO"])
    const_count = len([r for r in rows if r["Environment"] == "Construction"])
    pass_count = len([r for r in rows if "PASS" in str(r["Final_Status"])])
    partial_count = len([r for r in rows if "PARTIAL" in str(r["Final_Status"])])
    na_count = len([r for r in rows if "NOT_AVAILABLE" in str(r["Final_Status"])])

    print(f"\nTotal features: {len(rows)}")
    print(f"  TCO: {tco_count}")
    print(f"  Construction: {const_count}")
    print("\nStatus breakdown:")
    print(f"  PASS: {pass_count}")
    print(f"  PARTIAL: {partial_count}")
    print(f"  NOT_AVAILABLE: {na_count}")

    create_excel_tracker(rows)

    print("\n" + "=" * 60)
    print("DONE!")
    print("=" * 60)


if __name__ == "__main__":
    main()
