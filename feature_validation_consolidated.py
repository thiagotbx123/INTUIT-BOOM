import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUTPUT_FILE = r"C:\Users\adm_r\intuit-boom\FEATURE_VALIDATION_CONSOLIDATED.xlsx"

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
PHASE_FILL = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
ROW_FILL = PatternFill(start_color="EBF4FA", end_color="EBF4FA", fill_type="solid")
IN_PROGRESS_FILL = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

HEADER_FONT = Font(name="Cambria", bold=True, color="FFFFFF", size=11)
PHASE_FONT = Font(name="Cambria", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="Cambria", size=10)
STATUS_FONT = Font(name="Cambria", size=9, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Feature Validation"

headers = ["#", "Task / Deliverable", "Owner", "Status", "Start", "End"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 10

data = [
    ["2.2", "PHASE: FEATURE VALIDATION", "", "", "", "", True],
    [
        "",
        "AI AGENTS validation (WR-001 to WR-008)",
        "FDE",
        "In Progress",
        "Jan 20",
        "Jan 30",
        False,
    ],
    [
        "",
        "REPORTING validation (WR-009 to WR-015)",
        "FDE",
        "In Progress",
        "Jan 22",
        "Feb 03",
        False,
    ],
    [
        "",
        "DIMENSIONS validation (WR-016 to WR-019)",
        "FDE",
        "In Progress",
        "Jan 27",
        "Feb 03",
        False,
    ],
    [
        "",
        "WORKFLOW validation (WR-020)",
        "FDE",
        "In Progress",
        "Jan 29",
        "Feb 03",
        False,
    ],
    [
        "",
        "MIGRATION validation (WR-021 to WR-023)",
        "FDE",
        "In Progress",
        "Jan 29",
        "Feb 04",
        False,
    ],
    [
        "",
        "CONSTRUCTION validation (WR-024 to WR-025)",
        "FDE",
        "In Progress",
        "Jan 30",
        "Feb 04",
        False,
    ],
    [
        "",
        "PAYROLL validation (WR-026 to WR-029)",
        "FDE",
        "In Progress",
        "Jan 31",
        "Feb 05",
        False,
    ],
]

for row_idx, row_data in enumerate(data, 2):
    is_phase = row_data[6]

    for col_idx, value in enumerate(row_data[:6], 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")

        if is_phase:
            cell.fill = PHASE_FILL
            cell.font = PHASE_FONT
        else:
            cell.font = NORMAL_FONT
            if col_idx == 4:  # Status
                cell.fill = IN_PROGRESS_FILL
                cell.font = STATUS_FONT
            else:
                cell.fill = ROW_FILL

ws.freeze_panes = "A2"
wb.save(OUTPUT_FILE)
print(f"Created: {OUTPUT_FILE}")
